from django.shortcuts import get_object_or_404
from bokitas.models import Proyecto, Beneficiario, Menor, Familia, AntropBef, AntropMenor, Medicamento, Jornada, Nutricional
from django.db.models import Q
from django.http.response import HttpResponse
from django.views.generic.base import TemplateView
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import Reference, BarChart3D

# Create your views here.

class exportar_proyecto(TemplateView):
    def get(self, request, *args, **kwargs):
        proyecto = request.GET.get('proyecto')
        workbook = Workbook()
        bandera = True
        fecha = datetime.now()
        fecha_fin = fecha.strftime('%d-%m-%Y - hora: %H:%m')
        
        hojas = ["REGISTRO BENEFICIARIOS","REGISTRO MENORES","ANTROPOMETRICOS MENORES","ANTROPOMETRICO EMBARAZADAS","ANTROPOMETRICO LACTANTE","MEDICAMENTOS Y PRODUCTOS"]

        for hoja in hojas: 
            if bandera:
                worksheet = workbook.active
                worksheet.title = hoja
                bandera = False
                worksheet.merge_cells('A1:C1')
                first_cell = worksheet['A1']
                first_cell.value = "Fecha: " + fecha_fin
            else:
                worksheet = workbook.create_sheet(hoja)

    #*********  Crear encabezado en la hoja  *************
            worksheet.merge_cells('A2:D2')
            second_cell = worksheet['A2']
            second_cell.value = "Bokitas"
            second_cell.font  = Font(name = 'Tw Cen MT Condensed Extra Bold', size = 52, bold = True, color="6F42C1")

            worksheet.merge_cells('A3:D3')
            third_cell = worksheet['A3']
            third_cell.value = "Bokitas Fundation    www.bokitas.org"
            third_cell.font  = Font(name = 'Tw Cen MT Condensed Extra Bold', size = 12, bold = True, color="6F42C1")
        
#***********************  HOJA DE DATOS DE LOS BENEFICIARIOS  ********************************

    #*********  Registro de Datos de los Beneficiario  *************
        worksheet = workbook.worksheets[0]
        worksheet.merge_cells('A4:W6')
        fourth_cell = worksheet['A4']
        fourth_cell.value = "REGISTRO DE LOS BENEFICIARIOS"
        fourth_cell.font  = Font(name = 'Tahoma', size = 16, bold = True, color="333399")
        fourth_cell.alignment = Alignment(horizontal="center", vertical="center")      

        beneficiarios = Beneficiario.objects.filter(proyecto=proyecto).order_by('cedula')
        
        #**************  Obtener el total de beneficiarios  ***************

        total_beneficiarios = beneficiarios.count()
        total_embarazadas = beneficiarios.values("embarazada").filter(embarazada="SI").count()
        total_lactantes = beneficiarios.values("lactante").filter(lactante="SI").count()


        titulos = ['PROYECTO','CÉDULA','NOMBRE','APELLIDO','SEXO','FECHA NAC.','EDAD','MESES','NACIONALIDAD','NUM HIJOS','EMBARAZADA','LACTANDO','ESTADO CIVIL','EDUCACIÓN','PROFESIÓN','LABORAL','TELEFONO','CORREO','DIRECCIÓN','CIUDAD','ESTADO','ESTATUS','NÚMERO DE CUENTA']
        row_num = 7
        thin = Side(border_style="thin", color="000000")
        double = Side(border_style="double", color="000000")

    #********* asigna el titulo a las columnas  ************************
        for col_num, column_title in enumerate(titulos, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.fill = PatternFill("solid", fgColor="E2D9F3")
            cell.border = Border(top=thin, left=thin, right=thin, bottom=double)
            cell.font  = Font(bold=True, size = 14, color="333399")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            adjusted_width = (len(cell.value) + 10) * 1.2
            worksheet.column_dimensions[get_column_letter(col_num)].width = adjusted_width

    #**************  Agrega la data a las celdas
        for beneficiario in beneficiarios:
            row_num += 1
            if beneficiario.fecha_nac:
                datos = [beneficiario.proyecto,beneficiario.cedula,beneficiario.nombre,beneficiario.apellido,beneficiario.sexo,beneficiario.fecha_nac.strftime('%d-%m-%Y'),beneficiario.edad,beneficiario.meses,beneficiario.nacionalidad,beneficiario.num_hijos,beneficiario.embarazada,beneficiario.lactante,beneficiario.estado_civil,beneficiario.educacion,beneficiario.profesion,beneficiario.laboral,beneficiario.telefono,beneficiario.correo,beneficiario.direccion,beneficiario.ciudad,beneficiario.estado,beneficiario.estatus,beneficiario.numero_cuenta]
            else:
                datos = [beneficiario.proyecto,beneficiario.cedula,beneficiario.nombre,beneficiario.apellido,beneficiario.sexo,beneficiario.fecha_nac,beneficiario.edad,beneficiario.meses,beneficiario.nacionalidad,beneficiario.num_hijos,beneficiario.embarazada,beneficiario.lactante,beneficiario.estado_civil,beneficiario.educacion,beneficiario.profesion,beneficiario.laboral,beneficiario.telefono,beneficiario.correo,beneficiario.direccion,beneficiario.ciudad,beneficiario.estado,beneficiario.estatus,beneficiario.numero_cuenta]


            for col_num, cell_value in enumerate(datos, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = str(cell_value)
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                if col_num == 6 or col_num == 7 or col_num == 8 or col_num == 9 or col_num == 10 or col_num == 11 or col_num == 12 or col_num == 22:
                    cell.alignment = Alignment(horizontal="center")

        #**************  Agrega el total de beneficiarios  ************************
        row_num += 2
        worksheet.merge_cells('A'+str(row_num+1)+':B'+str(row_num+1))
        cell = worksheet.cell(row=row_num+1, column=1)
        cell.value = "Total de Beneficiarios: " +' ' + str(total_beneficiarios)
        cell2 = worksheet.cell(row=row_num+2, column=1)
        cell2.value = "Total de Embarazadas: " +' ' + str(total_embarazadas)
        cell3 = worksheet.cell(row=row_num+3, column=1)
        cell3.value = "Total de Lactantes: " +' ' + str(total_lactantes)

#***********************  HOJA DE DATOS DE LOS MENORES  ********************************

    #*********  Registro de Datos de los Menores  *************
        worksheet = workbook.worksheets[1]
        worksheet.merge_cells('A4:P6')
        fourth_cell = worksheet['A4']
        fourth_cell.value = "REGISTROS DEL PERFIL DE MENORES"
        fourth_cell.font  = Font(name = 'Tahoma', size = 16, bold = True, color="333399")
        fourth_cell.alignment = Alignment(horizontal="center", vertical="center")      

        menores = Menor.objects.filter(proyecto=proyecto).order_by('-cedula_bef')

        #**************  Obtener el total de Menores  ***************
        total_menores = menores.count()
        total_activos = menores.values("estatus").filter(estatus="ACTIVO").count()
        total_altas = menores.values("estatus").filter(estatus="ALTA").count()
        total_egresos = menores.values("estatus").filter(estatus="EGRESO").count()
        total_desincorporados = menores.values("estatus").filter(estatus="DESINCORPORADO").count()
        total_menores_5 = menores.values("edad").filter(edad__lte=5).count()
        total_menores_2 = menores.values("edad").filter(edad__lte=2).count()

        titulos = ['PROYECTO','REPRESENTANTE','PARENTESCO','CÉDULA','NOMBRE','APELLIDO','SEXO','FECHA NAC.','EDAD','MESES','PESO ACTUAL','TALLA ACTUAL','DIAGNOSTICO PESO','DIAGNOSTICO TALLA','FECHA INGRESO','ESTATUS']
        row_num = 7
        thin = Side(border_style="thin", color="000000")
        double = Side(border_style="double", color="000000")

    #********* asigna el titulo a las columnas  ************************
        for col_num, column_title in enumerate(titulos, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.fill = PatternFill("solid", fgColor="E2D9F3")
            cell.border = Border(top=thin, left=thin, right=thin, bottom=double)
            cell.font  = Font(bold=True, size = 14, color="333399")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            adjusted_width = (len(cell.value) + 10) * 1.2
            worksheet.column_dimensions[get_column_letter(col_num)].width = adjusted_width

    #**************  Agrega la data a las celdas
        for menor in menores:
            row_num += 1
            datos = [menor.proyecto,menor.cedula_bef,menor.parentesco,menor.cedula,menor.nombre,menor.apellido,menor.sexo,menor.fecha_nac.strftime('%d-%m-%Y'),menor.edad,menor.meses,menor.peso_actual,menor.talla_actual,menor.diagnostico_actual,menor.diagnostico_talla_actual,menor.fecha_ing_proyecto.strftime('%d-%m-%Y'),menor.estatus]

            for col_num, cell_value in enumerate(datos, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = str(cell_value)
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                if col_num == 3 or col_num == 8 or col_num == 9 or col_num == 10 or col_num == 11 or col_num == 12 or col_num == 15 or col_num == 16:
                    cell.alignment = Alignment(horizontal="center")

        #**************  Agrega el total de Menores  ************************
        row_num += 2
        worksheet.merge_cells('A'+str(row_num+1)+':B'+str(row_num+1))
        cell = worksheet.cell(row=row_num+1, column=1)
        cell.value = "Total de Menores: " +' ' + str(total_menores)
        cell2 = worksheet.cell(row=row_num+2, column=1)
        cell2.value = "Total de Activos: " +' ' + str(total_activos)
        cell3 = worksheet.cell(row=row_num+3, column=1)
        cell3.value = "Total de Altas: " +' ' + str(total_altas)
        cell4 = worksheet.cell(row=row_num+4, column=1)
        cell4.value = "Total de Egresos: " +' ' + str(total_egresos)
        cell5 = worksheet.cell(row=row_num+5, column=1)
        cell5.value = "Total de Desincorporados: " +' ' + str(total_desincorporados)
        cell6 = worksheet.cell(row=row_num+6, column=1)
        cell6.value = "Total de Menores de 5 años: " +' ' + str(total_menores_5)
        cell7 = worksheet.cell(row=row_num+7, column=1)
        cell7.value = "Total de Menores de 2 años: " +' ' + str(total_menores_2)


#*************  HOJA DE DATOS ANTROPOMETRICOS DEL MENORES  *********************

    #*********  Registro de Datos antropometricos Menores  *************
        worksheet = workbook.worksheets[2]
        worksheet.merge_cells('A4:W6')
        fourth_cell = worksheet['A4']
        fourth_cell.value = "REGISTRO DATOS ANTROPOMETRICOS DEL MENOR"
        fourth_cell.font  = Font(name = 'Tahoma', size = 16, bold = True, color="333399")
        fourth_cell.alignment = Alignment(horizontal="center", vertical="center")      

        imc_menores = AntropMenor.objects.filter(proyecto_id=proyecto).order_by('-cedula_bef','cedula_id').select_related('cedula')

        titulos = ['PROYECTO','REPRESENTANTE','CÉDULA','NOMBRE','APELLIDO','SEXO','FECHA NAC.','EDAD','MESES','PESO ACTUAL','TALLA ACTUAL','FECHA JORNADA','EDAD','MESES','PESO','TALLA','CBI','PTR','PSE','CC','IMC','DIAGNOSTICO PESO','DIAGNOSTICO TALLA']
        row_num = 7
        thin = Side(border_style="thin", color="000000")
        double = Side(border_style="double", color="000000")

    #********* asigna el titulo a las columnas  ************************
        for col_num, column_title in enumerate(titulos, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.fill = PatternFill("solid", fgColor="E2D9F3")
            cell.border = Border(top=thin, left=thin, right=thin, bottom=double)
            cell.font  = Font(bold=True, size = 14, color="333399")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            adjusted_width = (len(cell.value) + 10) * 1.2
            worksheet.column_dimensions[get_column_letter(col_num)].width = adjusted_width

    #**************  Agrega la data a las celdas
        xCedula = 0

        for imc in imc_menores:
            row_num += 1
            datos = [imc.proyecto,imc.cedula_bef,imc.cedula.cedula,imc.cedula.nombre,imc.cedula.apellido,imc.cedula.sexo,imc.cedula.fecha_nac.strftime('%d-%m-%Y'),imc.edad,imc.meses,imc.cedula.peso_actual,imc.cedula.talla_actual,imc.jornada.jornada.strftime('%d-%m-%Y'),imc.edad,imc.meses,imc.peso,imc.talla,imc.cbi,imc.ptr,imc.pse,imc.cc,imc.imc,imc.diagnostico,imc.diagnostico_talla]

            if imc.cedula.cedula != xCedula:
                for col_num, cell_value in enumerate(datos, 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.value = str(cell_value)
                    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                    cell.alignment = Alignment(horizontal="center")
                    if col_num == 1 or col_num == 2 or col_num == 4 or col_num == 5 or col_num == 22 or col_num == 23:
                        cell.alignment = Alignment(horizontal="left")

                xCedula = imc.cedula.cedula
            else:
                for col_num, cell_value in enumerate(datos, 1):
                    if col_num > 11:
                        cell = worksheet.cell(row=row_num, column=col_num)
                        cell.value = str(cell_value)
                        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                        cell.alignment = Alignment(horizontal="center")
                        if col_num == 22 or col_num == 23:
                            cell.alignment = Alignment(horizontal="left")


#*************  HOJA DE DATOS ANTROPOMETRICOS EMBARAZADAS  *********************

    #*********  Registro de Datos antropometricos Embarazadas  *************
        worksheet = workbook.worksheets[3]
        worksheet.merge_cells('A4:P6')
        fourth_cell = worksheet['A4']
        fourth_cell.value = "REGISTRO DATOS ANTROPOMETRICOS EMBARAZADAS"
        fourth_cell.font  = Font(name = 'Tahoma', size = 16, bold = True, color="333399")
        fourth_cell.alignment = Alignment(horizontal="center", vertical="center")      

        imc_embarazos = AntropBef.objects.filter(proyecto=proyecto, embarazo_lactando = 'EMBARAZADA').order_by('-cedula_bef').select_related('cedula_bef')

        titulos = ['PROYECTO','CÉDULA','NOMBRE','APELLIDO','FECHA NAC.','FECHA JORNADA','EDAD','MESES','MESES EMBARAZO','PESO','TALLA','CBI','IMC','DIAGNOSTICO PESO','RIESGO','OBSERVACIONES']
        row_num = 7
        thin = Side(border_style="thin", color="000000")
        double = Side(border_style="double", color="000000")

    #********* asigna el titulo a las columnas  ************************
        for col_num, column_title in enumerate(titulos, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.fill = PatternFill("solid", fgColor="E2D9F3")
            cell.border = Border(top=thin, left=thin, right=thin, bottom=double)
            cell.font  = Font(bold=True, size = 14, color="333399")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            adjusted_width = (len(cell.value) + 10) * 1.2
            worksheet.column_dimensions[get_column_letter(col_num)].width = adjusted_width

    #**************  Agrega la data a las celdas
        xCedula = 0

        for imc in imc_embarazos:
            row_num += 1
            if beneficiario.fecha_nac:
                datos = [imc.cedula_bef.proyecto,imc.cedula_bef.cedula,imc.cedula_bef.nombre,imc.cedula_bef.apellido,imc.cedula_bef.fecha_nac.strftime('%d-%m-%Y'),imc.jornada.jornada.strftime('%d-%m-%Y'),imc.edad,imc.meses,imc.tiempo_gestacion,imc.peso,imc.talla,imc.cbi,imc.imc,imc.diagnostico,imc.riesgo,imc.observacion]
            else:
                datos = [imc.cedula_bef.proyecto,imc.cedula_bef.cedula,imc.cedula_bef.nombre,imc.cedula_bef.apellido,imc.cedula_bef.fecha_nac,imc.jornada.jornada.strftime('%d-%m-%Y'),imc.edad,imc.meses,imc.tiempo_gestacion,imc.peso,imc.talla,imc.cbi,imc.imc,imc.diagnostico,imc.riesgo,imc.observacion]

            if imc.cedula_bef.cedula != xCedula:
                for col_num, cell_value in enumerate(datos, 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.value = str(cell_value)
                    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                    cell.alignment = Alignment(horizontal="center")
                    if col_num == 1 or col_num == 2 or col_num == 3 or col_num == 4:
                        cell.alignment = Alignment(horizontal="left")

                xCedula = imc.cedula_bef.cedula
            else:
                for col_num, cell_value in enumerate(datos, 1):
                    if col_num > 7:
                        cell = worksheet.cell(row=row_num, column=col_num)
                        cell.value = str(cell_value)
                        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                        cell.alignment = Alignment(horizontal="center")


#*************  HOJA DE DATOS ANTROPOMETRICOS LACTANTES  *********************

    #*********  Registro de Datos antropometricos Lactante  *************
        worksheet = workbook.worksheets[4]
        worksheet.merge_cells('A4:P6')
        fourth_cell = worksheet['A4']
        fourth_cell.value = "REGISTRO DATOS ANTROPOMETRICOS LACTANTES"
        fourth_cell.font  = Font(name = 'Tahoma', size = 16, bold = True, color="333399")
        fourth_cell.alignment = Alignment(horizontal="center", vertical="center")      

        imc_embarazos = AntropBef.objects.filter(proyecto=proyecto, embarazo_lactando = 'LACTANDO').order_by('-cedula_bef').select_related('cedula_bef')

        titulos = ['PROYECTO','CÉDULA','NOMBRE','APELLIDO','FECHA NAC.','FECHA JORNADA','EDAD','MESES','MESES LACTANDO','PESO','TALLA','CBI','IMC','DIAGNOSTICO PESO','RIESGO','OBSERVACIONES']
        row_num = 7
        thin = Side(border_style="thin", color="000000")
        double = Side(border_style="double", color="000000")

    #********* asigna el titulo a las columnas  ************************
        for col_num, column_title in enumerate(titulos, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.fill = PatternFill("solid", fgColor="E2D9F3")
            cell.border = Border(top=thin, left=thin, right=thin, bottom=double)
            cell.font  = Font(bold=True, size = 14, color="333399")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            adjusted_width = (len(cell.value) + 10) * 1.2
            worksheet.column_dimensions[get_column_letter(col_num)].width = adjusted_width

    #**************  Agrega la data a las celdas
        xCedula = 0

        for imc in imc_embarazos:
            row_num += 1
            if beneficiario.fecha_nac:
                datos = [imc.cedula_bef.proyecto,imc.cedula_bef.cedula,imc.cedula_bef.nombre,imc.cedula_bef.apellido,imc.cedula_bef.fecha_nac.strftime('%d-%m-%Y'),imc.jornada.jornada.strftime('%d-%m-%Y'),imc.edad,imc.meses,imc.tiempo_gestacion,imc.peso,imc.talla,imc.cbi,imc.imc,imc.diagnostico,imc.riesgo,imc.observacion]
            else:
                datos = [imc.cedula_bef.proyecto,imc.cedula_bef.cedula,imc.cedula_bef.nombre,imc.cedula_bef.apellido,imc.cedula_bef.fecha_nac,imc.jornada.jornada.strftime('%d-%m-%Y'),imc.edad,imc.meses,imc.tiempo_gestacion,imc.peso,imc.talla,imc.cbi,imc.imc,imc.diagnostico,imc.riesgo,imc.observacion]

            if imc.cedula_bef.cedula != xCedula:
                for col_num, cell_value in enumerate(datos, 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.value = str(cell_value)
                    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                    cell.alignment = Alignment(horizontal="center")
                    if col_num == 1 or col_num == 2 or col_num == 3 or col_num == 4:
                        cell.alignment = Alignment(horizontal="left")

                xCedula = imc.cedula_bef.cedula
            else:
                for col_num, cell_value in enumerate(datos, 1):
                    if col_num > 7:
                        cell = worksheet.cell(row=row_num, column=col_num)
                        cell.value = str(cell_value)
                        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                        cell.alignment = Alignment(horizontal="center")

#*************  HOJA DE DATOS ENTREGA MEDICAMENTOS PRODUCTOS  *********************

    #*********  Registro de Datos DATOS ENTREGA MEDICAMENTOS  *************
        worksheet = workbook.worksheets[5]
        worksheet.merge_cells('A4:H6')
        fourth_cell = worksheet['A4']
        fourth_cell.value = "REGISTRO DE ENTREGA DE MEDICAMENTOS Y PRODUCTOS"
        fourth_cell.font  = Font(name = 'Tahoma', size = 16, bold = True, color="333399")
        fourth_cell.alignment = Alignment(horizontal="center", vertical="center")      

        medicamentos = Medicamento.objects.filter(proyecto=proyecto).order_by('-cedula_bef').select_related('cedula_bef')

        titulos = ['PROYECTO','CÉDULA','NOMBRE','APELLIDO','FECHA JORNADA','NOMBRE DEL PRODUCTO','DESCRIPCION DEL PRODUCTO','CANTIDAD']
        row_num = 7
        thin = Side(border_style="thin", color="000000")
        double = Side(border_style="double", color="000000")

    #********* asigna el titulo a las columnas  ************************
        for col_num, column_title in enumerate(titulos, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.fill = PatternFill("solid", fgColor="E2D9F3")
            cell.border = Border(top=thin, left=thin, right=thin, bottom=double)
            cell.font  = Font(bold=True, size = 14, color="333399")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            adjusted_width = (len(cell.value) + 10) * 1.2
            worksheet.column_dimensions[get_column_letter(col_num)].width = adjusted_width

    #**************  Agrega la data a las celdas

        for med in medicamentos:
            row_num += 1
            datos = [med.cedula_bef.proyecto,med.cedula_bef.cedula,med.cedula_bef.nombre,med.cedula_bef.apellido,med.jornada,med.nombre,med.descripcion,med.cantidad]

            for col_num, cell_value in enumerate(datos, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = str(cell_value)
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                cell.alignment = Alignment(horizontal="center")
                if col_num == 1 or col_num == 2 or col_num == 3 or col_num == 4:
                    cell.alignment = Alignment(horizontal="left")


    #*********  Establecer el nombre del Archivo *******
        nombre_archvo = "Reporte_por_Proyecto.xlsx"
    
    #*********  Definir el tipo de respuesta que se va a dar ***********
        response = HttpResponse(content_type = "application/ms-excel")
        contenido = "attachment; filename = {0}".format(nombre_archvo)
        response["Content-Disposition"] = contenido
        workbook.save(response)
        return response
    


class exp_beneficiario_detalle(TemplateView):
    def get(self, request,  pk, *args, **kwargs):
        workbook = Workbook()
        fecha = datetime.now()
        fecha_fin = fecha.strftime('%d-%m-%Y - hora: %H:%m') 
        worksheet = workbook.active
        worksheet.title = "REGISTRO BENEFICIARIO"
        worksheet.merge_cells('A1:C1')
        first_cell = worksheet['A1']
        first_cell.value = "Fecha: " + fecha_fin

    #*********  Crear encabezado en la hoja  *************
        worksheet.merge_cells('A2:D2')
        second_cell = worksheet['A2']
        second_cell.value = "Bokitas"
        second_cell.font  = Font(name = 'Tw Cen MT Condensed Extra Bold', size = 52, bold = True, color="6F42C1")

        worksheet.merge_cells('A3:D3')
        third_cell = worksheet['A3']
        third_cell.value = "Bokitas Fundation    www.bokitas.org"
        third_cell.font  = Font(name = 'Tw Cen MT Condensed Extra Bold', size = 12, bold = True, color="6F42C1")
        
#***********************  HOJA DE DATOS DE LOS BENEFICIARIOS  ********************************

    #*********  Registro de Datos de los Beneficiario  *************
        worksheet = workbook.worksheets[0]
       # worksheet = workbook['REGISTRO BENEFICIARIOS']
        worksheet.merge_cells('A4:W6')
        fourth_cell = worksheet['A4']
        fourth_cell.value = "REGISTRO DEL BENEFICIARIO"
        fourth_cell.font  = Font(name = 'Tahoma', size = 16, bold = True, color="333399")
        fourth_cell.alignment = Alignment(horizontal="center", vertical="center")      

        beneficiarios = Beneficiario.objects.filter(id=pk)
        
        titulos = ['PROYECTO','CÉDULA','NOMBRE','APELLIDO','SEXO','FECHA NAC.','EDAD','MESES','NACIONALIDAD','NUM HIJOS','EMBARAZADA','LACTANDO','ESTADO CIVIL','EDUCACIÓN','PROFESIÓN','LABORAL','TELEFONO','CORREO','DIRECCIÓN','CIUDAD','ESTADO','ESTATUS','NÚMERO DE CUENTA']
        row_num = 7
        thin = Side(border_style="thin", color="000000")
        double = Side(border_style="double", color="000000")

    #********* asigna el titulo a las columnas  ************************
        for col_num, column_title in enumerate(titulos, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.fill = PatternFill("solid", fgColor="E2D9F3")
            cell.border = Border(top=thin, left=thin, right=thin, bottom=double)
            cell.font  = Font(bold=True, size = 14, color="333399")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            adjusted_width = (len(cell.value) + 10) * 1.2
            worksheet.column_dimensions[get_column_letter(col_num)].width = adjusted_width

    #**************  Agrega la data a las celdas
        for beneficiario in beneficiarios:
            row_num += 1
            if beneficiario.fecha_nac:
                datos = [beneficiario.proyecto,beneficiario.cedula,beneficiario.nombre,beneficiario.apellido,beneficiario.sexo,beneficiario.fecha_nac.strftime('%d-%m-%Y'),beneficiario.edad,beneficiario.meses,beneficiario.nacionalidad,beneficiario.num_hijos,beneficiario.embarazada,beneficiario.lactante,beneficiario.estado_civil,beneficiario.educacion,beneficiario.profesion,beneficiario.laboral,beneficiario.telefono,beneficiario.correo,beneficiario.direccion,beneficiario.ciudad,beneficiario.estado,beneficiario.estatus,beneficiario.numero_cuenta]
            else:
                datos = [beneficiario.proyecto,beneficiario.cedula,beneficiario.nombre,beneficiario.apellido,beneficiario.sexo,beneficiario.fecha_nac,beneficiario.edad,beneficiario.meses,beneficiario.nacionalidad,beneficiario.num_hijos,beneficiario.embarazada,beneficiario.lactante,beneficiario.estado_civil,beneficiario.educacion,beneficiario.profesion,beneficiario.laboral,beneficiario.telefono,beneficiario.correo,beneficiario.direccion,beneficiario.ciudad,beneficiario.estado,beneficiario.estatus,beneficiario.numero_cuenta]

            for col_num, cell_value in enumerate(datos, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = str(cell_value)
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                if col_num == 6 or col_num == 7 or col_num == 8 or col_num == 9 or col_num == 10 or col_num == 11 or col_num == 12 or col_num == 22:
                    cell.alignment = Alignment(horizontal="center")


    
    #*********  Registro de Datos antropometricos Menores  *************

        row_num = row_num + 3
        xx = 'A'+ str(row_num)
        yy = 'V'+ str(row_num + 2)
        worksheet.merge_cells(xx +':'+ yy)
        fourth_cell = worksheet[xx]
        fourth_cell.value = "REGISTRO DE HIJOS"
        fourth_cell.font  = Font(name = 'Tahoma', size = 16, bold = True, color="333399")
        fourth_cell.alignment = Alignment(horizontal="center", vertical="center")      

        imc_menores = AntropMenor.objects.filter(cedula_bef_id = pk).order_by('cedula_id').select_related('cedula')

        titulos = ['PROYECTO','CÉDULA','NOMBRE','APELLIDO','SEXO','FECHA NAC.','EDAD','MESES','PESO ACTUAL','TALLA ACTUAL','FECHA JORNADA','EDAD','MESES','PESO','TALLA','CBI','PTR','PSE','CC','IMC','DIAGNOSTICO PESO','DIAGNOSTICO TALLA']
        thin = Side(border_style="thin", color="000000")
        double = Side(border_style="double", color="000000")
        row_num = row_num + 3
    #********* asigna el titulo a las columnas  ************************
        for col_num, column_title in enumerate(titulos, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            if col_num > 10:
                cell.fill = PatternFill("solid", fgColor="9EEAF9")
            else:
                cell.fill = PatternFill("solid", fgColor="E2D9F3")
            cell.border = Border(top=thin, left=thin, right=thin, bottom=double)
            cell.font  = Font(bold=True, size = 14, color="333399")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if col_num == 21 or col_num == 22: 
                adjusted_width = (len(cell.value) + 10) * 1.2
                worksheet.column_dimensions[get_column_letter(col_num)].width = adjusted_width

    #**************  Agrega la data a las celdas
        xCedula = 0

        for imc in imc_menores:
            row_num += 1
            datos = [imc.proyecto,imc.cedula.cedula,imc.cedula.nombre,imc.cedula.apellido,imc.cedula.sexo,imc.cedula.fecha_nac.strftime('%d-%m-%Y'),imc.edad,imc.meses,imc.cedula.peso_actual,imc.cedula.talla_actual,imc.jornada.jornada.strftime('%d-%m-%Y'),imc.edad,imc.meses,imc.peso,imc.talla,imc.cbi,imc.ptr,imc.pse,imc.cc,imc.imc,imc.diagnostico,imc.diagnostico_talla]
            
            if imc.cedula.cedula != xCedula:
                for col_num, cell_value in enumerate(datos, 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.value = str(cell_value)
                    cell.border = Border(top=double, left=thin, right=thin, bottom=thin)
                    cell.alignment = Alignment(horizontal="center")
                    if col_num == 1 or col_num == 2 or col_num == 3 or col_num == 4 or col_num == 5 or col_num == 21 or col_num == 22:
                        cell.alignment = Alignment(horizontal="left")

                xCedula = imc.cedula.cedula
            else:
                for col_num, cell_value in enumerate(datos, 1):
                    if col_num > 10:
                        cell = worksheet.cell(row=row_num, column=col_num)
                        cell.value = str(cell_value)
                        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                        cell.alignment = Alignment(horizontal="center")
                        if col_num == 21 or col_num == 22:
                            cell.alignment = Alignment(horizontal="left")


    #*********  Registro de Datos antropometricos Beneficiario *************
    
        row_num = row_num + 3
        xx = 'A'+ str(row_num)
        yy = 'O'+ str(row_num + 2)
        worksheet.merge_cells(xx +':'+ yy)
        fourth_cell = worksheet[xx]
        fourth_cell.value = "REGISTRO ANTROPOMÉTRICO DEL BENEFICIARIO"
        fourth_cell.font  = Font(name = 'Tahoma', size = 16, bold = True, color="333399")
        fourth_cell.alignment = Alignment(horizontal="center", vertical="center")      

        imc_beneficiario = AntropBef.objects.filter(cedula_bef_id = pk).order_by('cedula_bef_id').select_related('cedula_bef')

        titulos = ['PROYECTO','CÉDULA','NOMBRE','APELLIDO','SEXO','FECHA NAC.','FECHA JORNADA','ESTADO','EDAD','MESES','PESO','TALLA','CBI','IMC','DIAGNOSTICO PESO']
        thin = Side(border_style="thin", color="000000")
        double = Side(border_style="double", color="000000")
        row_num = row_num + 3
    #********* asigna el titulo a las columnas  ************************
        for col_num, column_title in enumerate(titulos, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            if col_num > 6:
                cell.fill = PatternFill("solid", fgColor="9EEAF9")
            else:
                cell.fill = PatternFill("solid", fgColor="E2D9F3")
            cell.border = Border(top=thin, left=thin, right=thin, bottom=double)
            cell.font  = Font(bold=True, size = 14, color="333399")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if col_num == 21 or col_num == 22: 
                adjusted_width = (len(cell.value) + 10) * 1.2
                worksheet.column_dimensions[get_column_letter(col_num)].width = adjusted_width

    #**************  Agrega la data a las celdas
        xCedula = 0

        for imc in imc_beneficiario:
            row_num += 1
            datos = [imc.cedula_bef.proyecto,imc.cedula_bef.cedula,imc.cedula_bef.nombre,imc.cedula_bef.apellido,imc.cedula_bef.sexo,imc.cedula_bef.fecha_nac.strftime('%d-%m-%Y'),imc.jornada.jornada.strftime('%d-%m-%Y'),imc.embarazo_lactando,imc.edad,imc.meses,imc.peso,imc.talla,imc.cbi,imc.imc,imc.diagnostico]
            
            if imc.cedula_bef != xCedula:
                for col_num, cell_value in enumerate(datos, 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.value = str(cell_value)
                    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                    cell.alignment = Alignment(horizontal="left")
                    if col_num == 6 or col_num == 7 or col_num == 8 or col_num == 9 or col_num == 10 or col_num == 11 or col_num == 12 or col_num == 13 or col_num == 14:
                        cell.alignment = Alignment(horizontal="center")

                xCedula = imc.cedula_bef
            else:
                for col_num, cell_value in enumerate(datos, 1):
                    if col_num > 6:
                        cell = worksheet.cell(row=row_num, column=col_num)
                        cell.value = str(cell_value)
                        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                        cell.alignment = Alignment(horizontal="center")
                        if col_num == 15:
                            cell.alignment = Alignment(horizontal="left")


    #*********  Registro entrega de Medicamentos / Productos al Beneficiario *************
    
        row_num = row_num + 3
        xx = 'A'+ str(row_num)
        yy = 'H'+ str(row_num + 2)
        worksheet.merge_cells(xx +':'+ yy)
        fourth_cell = worksheet[xx]
        fourth_cell.value = "REGISTRO DE ENTREGA DE MEDICAMENTOS / PRODUCTOS"
        fourth_cell.font  = Font(name = 'Tahoma', size = 16, bold = True, color="333399")
        fourth_cell.alignment = Alignment(horizontal="center", vertical="center")      

        medicamentos = Medicamento.objects.filter(cedula_bef_id = pk).order_by('cedula_bef_id').select_related('cedula_bef')

        titulos = ['PROYECTO','CÉDULA','NOMBRE','APELLIDO','FECHA JORNADA','NOMBRE PROD.',' DESCRIPCIÓN ','CANTIDAD']
        thin = Side(border_style="thin", color="000000")
        double = Side(border_style="double", color="000000")
        row_num = row_num + 3
    #********* asigna el titulo a las columnas  ************************
        for col_num, column_title in enumerate(titulos, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.fill = PatternFill("solid", fgColor="E2D9F3")
            cell.border = Border(top=thin, left=thin, right=thin, bottom=double)
            cell.font  = Font(bold=True, size = 14, color="333399")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if col_num == 5 or col_num == 7: 
                adjusted_width = (len(cell.value) + 10) * 1.2
                worksheet.column_dimensions[get_column_letter(col_num)].width = adjusted_width

    #**************  Agrega la data a las celdas
        xCedula = 0

        for med in medicamentos:
            row_num += 1
            datos = [med.cedula_bef.proyecto,med.cedula_bef.cedula,med.cedula_bef.nombre,med.cedula_bef.apellido,med.jornada.jornada.strftime('%d-%m-%Y'),med.nombre,med.descripcion,med.cantidad]
            
            
            for col_num, cell_value in enumerate(datos, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = str(cell_value)
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                cell.alignment = Alignment(horizontal="left")
                if col_num == 5 or col_num == 8:
                    cell.alignment = Alignment(horizontal="center")



    #*********  Registro Familiares del Beneficiario *************
    
        row_num = row_num + 3
        xx = 'A'+ str(row_num)
        yy = 'L'+ str(row_num + 2)
        worksheet.merge_cells(xx +':'+ yy)
        fourth_cell = worksheet[xx]
        fourth_cell.value = "REGISTRO DE FAMILIARES DEL BENEFICIARIO"
        fourth_cell.font  = Font(name = 'Tahoma', size = 16, bold = True, color="333399")
        fourth_cell.alignment = Alignment(horizontal="center", vertical="center")      

        familiares = Familia.objects.filter(cedula_bef_id = pk)

        titulos = ['PARENTESCO','CÉDULA','NOMBRE','APELLIDO','SEXO','FECHA NAC.','EDAD','MESES','ESTADO CIVIL',' EDUCACIÓN ','PROFESIÓN','LABORAL']
        thin = Side(border_style="thin", color="000000")
        double = Side(border_style="double", color="000000")
        row_num = row_num + 3
    #********* asigna el titulo a las columnas  ************************
        for col_num, column_title in enumerate(titulos, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.fill = PatternFill("solid", fgColor="E2D9F3")
            cell.border = Border(top=thin, left=thin, right=thin, bottom=double)
            cell.font  = Font(bold=True, size = 14, color="333399")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if col_num == 5 or col_num == 7: 
                adjusted_width = (len(cell.value) + 10) * 1.2
                worksheet.column_dimensions[get_column_letter(col_num)].width = adjusted_width

    #**************  Agrega la data a las celdas
        xCedula = 0

        for fam in familiares:
            row_num += 1
            datos = [fam.parentesco,fam.cedula,fam.nombre,fam.apellido,fam.sexo,fam.fecha_nac.strftime('%d-%m-%Y'),fam.edad,fam.meses,fam.estado_civil,fam.educacion,fam.profesion,fam.laboral]
            
            for col_num, cell_value in enumerate(datos, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = str(cell_value)
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                cell.alignment = Alignment(horizontal="left")
                if col_num == 6 or col_num == 7 or col_num == 8:
                    cell.alignment = Alignment(horizontal="center")


    #*********  Establecer el nombre del Archivo *******
        nombre_archvo = "Rep_Beneficiario_detalle.xlsx"
    
    #*********  Definir el tipo de respuesta que se va a dar ***********
        response = HttpResponse(content_type = "application/ms-excel")
        contenido = "attachment; filename = {0}".format(nombre_archvo)
        response["Content-Disposition"] = contenido
        workbook.save(response)
        return response
    


class listado_beneficiario(TemplateView):
    def get(self, request, *args, **kwargs):
        proyecto = request.GET.get('proyecto')
        workbook = Workbook()
        fecha = datetime.now()
        fecha_fin = fecha.strftime('%d-%m-%Y - hora: %H:%m') 
        worksheet = workbook.active
        worksheet.title = "LISTADO DE BENEFICIARIOS"
        worksheet.merge_cells('A1:C1')
        first_cell = worksheet['A1']
        first_cell.value = "Fecha: " + fecha_fin

    #*********  Crear encabezado en la hoja  *************
        worksheet.merge_cells('A2:D2')
        second_cell = worksheet['A2']
        second_cell.value = "Bokitas"
        second_cell.font  = Font(name = 'Tw Cen MT Condensed Extra Bold', size = 52, bold = True, color="6F42C1")

        worksheet.merge_cells('A3:D3')
        third_cell = worksheet['A3']
        third_cell.value = "Bokitas Fundation    www.bokitas.org"
        third_cell.font  = Font(name = 'Tw Cen MT Condensed Extra Bold', size = 12, bold = True, color="6F42C1")
        
#***********************  HOJA DE DATOS DE LOS BENEFICIARIOS  ********************************

    #*********  Registro de Datos de los Beneficiario  *************
        
        worksheet = workbook.worksheets[0]
        worksheet.merge_cells('A4:W6')
        fourth_cell = worksheet['A4']
        fourth_cell.value = "REGISTRO DEL BENEFICIARIO"
        fourth_cell.font  = Font(name = 'Tahoma', size = 16, bold = True, color="333399")
        fourth_cell.alignment = Alignment(horizontal="center", vertical="center") 

        beneficiarios = Beneficiario.objects.filter(proyecto=proyecto)

        #**************  Obtener el total de beneficiarios  ***************
        total_beneficiarios = beneficiarios.count()
        total_embarazadas = beneficiarios.values("embarazada").filter(embarazada="SI").count()
        total_lactantes = beneficiarios.values("lactante").filter(lactante="SI").count()

        titulos = ['PROYECTO','CÉDULA','NOMBRE','APELLIDO','SEXO','FECHA NAC.','EDAD','MESES','NACIONALIDAD','NUM HIJOS','EMBARAZADA','LACTANDO','ESTADO CIVIL','EDUCACIÓN','PROFESIÓN','LABORAL','TELEFONO','CORREO','DIRECCIÓN','CIUDAD','ESTADO','ESTATUS','NÚMERO DE CUENTA']
        row_num = 7
        thin = Side(border_style="thin", color="000000")
        double = Side(border_style="double", color="000000")

    #********* asigna el titulo a las columnas  ************************
        for col_num, column_title in enumerate(titulos, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.fill = PatternFill("solid", fgColor="E2D9F3")
            cell.border = Border(top=thin, left=thin, right=thin, bottom=double)
            cell.font  = Font(bold=True, size = 14, color="333399")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            adjusted_width = (len(cell.value) + 10) * 1.2
            worksheet.column_dimensions[get_column_letter(col_num)].width = adjusted_width

    #**************  Agrega la data a las celdas
        for beneficiario in beneficiarios:
            row_num += 1
            if beneficiario.fecha_nac:
                datos = [beneficiario.proyecto,beneficiario.cedula,beneficiario.nombre,beneficiario.apellido,beneficiario.sexo,beneficiario.fecha_nac.strftime('%d-%m-%Y'),beneficiario.edad,beneficiario.meses,beneficiario.nacionalidad,beneficiario.num_hijos,beneficiario.embarazada,beneficiario.lactante,beneficiario.estado_civil,beneficiario.educacion,beneficiario.profesion,beneficiario.laboral,beneficiario.telefono,beneficiario.correo,beneficiario.direccion,beneficiario.ciudad,beneficiario.estado,beneficiario.estatus,beneficiario.numero_cuenta]
            else:
                datos = [beneficiario.proyecto,beneficiario.cedula,beneficiario.nombre,beneficiario.apellido,beneficiario.sexo,beneficiario.fecha_nac,beneficiario.edad,beneficiario.meses,beneficiario.nacionalidad,beneficiario.num_hijos,beneficiario.embarazada,beneficiario.lactante,beneficiario.estado_civil,beneficiario.educacion,beneficiario.profesion,beneficiario.laboral,beneficiario.telefono,beneficiario.correo,beneficiario.direccion,beneficiario.ciudad,beneficiario.estado,beneficiario.estatus,beneficiario.numero_cuenta]

            for col_num, cell_value in enumerate(datos, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = str(cell_value)
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                if col_num == 6 or col_num == 7 or col_num == 8 or col_num == 9 or col_num == 10 or col_num == 11 or col_num == 12 or col_num == 22:
                    cell.alignment = Alignment(horizontal="center")

        #**************  Agrega el total de beneficiarios  ************************
        row_num += 2
        worksheet.merge_cells('A'+str(row_num+1)+':B'+str(row_num+1))
        cell = worksheet.cell(row=row_num+1, column=1)
        cell.value = "Total de Beneficiarios: " +' ' + str(total_beneficiarios)
        cell2 = worksheet.cell(row=row_num+2, column=1)
        cell2.value = "Total de Embarazadas: " +' ' + str(total_embarazadas)
        cell3 = worksheet.cell(row=row_num+3, column=1)
        cell3.value = "Total de Lactantes: " +' ' + str(total_lactantes)


    #*********  Establecer el nombre del Archivo *******
        nombre_archvo = "Rep_listado_Beneficiarios.xlsx"
    
    #*********  Definir el tipo de respuesta que se va a dar ***********
        response = HttpResponse(content_type = "application/ms-excel")
        contenido = "attachment; filename = {0}".format(nombre_archvo)
        response["Content-Disposition"] = contenido
        workbook.save(response)
        return response
    


class listado_menores(TemplateView):
    def get(self, request, *args, **kwargs):
        proyecto = request.GET.get('proyecto')
        workbook = Workbook()
        fecha = datetime.now()
        fecha_fin = fecha.strftime('%d-%m-%Y - hora: %H:%m') 
        worksheet = workbook.active
        worksheet.title = "LISTADO DE MENORES"
        worksheet.merge_cells('A1:C1')
        first_cell = worksheet['A1']
        first_cell.value = "Fecha: " + fecha_fin

    #*********  Crear encabezado en la hoja  *************
        worksheet.merge_cells('A2:D2')
        second_cell = worksheet['A2']
        second_cell.value = "Bokitas"
        second_cell.font  = Font(name = 'Tw Cen MT Condensed Extra Bold', size = 52, bold = True, color="6F42C1")

        worksheet.merge_cells('A3:D3')
        third_cell = worksheet['A3']
        third_cell.value = "Bokitas Fundation    www.bokitas.org"
        third_cell.font  = Font(name = 'Tw Cen MT Condensed Extra Bold', size = 12, bold = True, color="6F42C1")
        
#***********************  HOJA DE DATOS DE LOS MENORES  ********************************

    #*********  Registro de Datos de los Menores  *************
        worksheet = workbook.worksheets[0]
        worksheet.merge_cells('A4:P6')
        fourth_cell = worksheet['A4']
        fourth_cell.value = "REGISTROS DEL PERFIL DE MENORES"
        fourth_cell.font  = Font(name = 'Tahoma', size = 16, bold = True, color="333399")
        fourth_cell.alignment = Alignment(horizontal="center", vertical="center")      

        menores = Menor.objects.filter(proyecto=proyecto).order_by('-cedula_bef')

        #**************  Obtener el total de Menores  ***************
        total_menores = menores.count()
        total_activos = menores.values("estatus").filter(estatus="ACTIVO").count()
        total_altas = menores.values("estatus").filter(estatus="ALTA").count()
        total_egresos = menores.values("estatus").filter(estatus="EGRESO").count()
        total_desincorporados = menores.values("estatus").filter(estatus="DESINCORPORADO").count()
        total_menores_5 = menores.values("edad").filter(edad__lte=5).count()
        total_menores_2 = menores.values("edad").filter(edad__lte=2).count()

        titulos = ['PROYECTO','REPRESENTANTE','PARENTESCO','CÉDULA','NOMBRE','APELLIDO','SEXO','FECHA NAC.','EDAD','MESES','PESO ACTUAL','TALLA ACTUAL','DIAGNOSTICO PESO','DIAGNOSTICO TALLA','FECHA INGRESO','ESTATUS']
        row_num = 7
        thin = Side(border_style="thin", color="000000")
        double = Side(border_style="double", color="000000")

    #********* asigna el titulo a las columnas  ************************
        for col_num, column_title in enumerate(titulos, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.fill = PatternFill("solid", fgColor="E2D9F3")
            cell.border = Border(top=thin, left=thin, right=thin, bottom=double)
            cell.font  = Font(bold=True, size = 14, color="333399")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            adjusted_width = (len(cell.value) + 10) * 1.2
            worksheet.column_dimensions[get_column_letter(col_num)].width = adjusted_width

    #**************  Agrega la data a las celdas
        for menor in menores:
            row_num += 1
            datos = [menor.proyecto,menor.cedula_bef,menor.parentesco,menor.cedula,menor.nombre,menor.apellido,menor.sexo,menor.fecha_nac.strftime('%d-%m-%Y'),menor.edad,menor.meses,menor.peso_actual,menor.talla_actual,menor.diagnostico_actual,menor.diagnostico_talla_actual,menor.fecha_ing_proyecto.strftime('%d-%m-%Y'),menor.estatus]

            for col_num, cell_value in enumerate(datos, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = str(cell_value)
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                if col_num == 3 or col_num == 8 or col_num == 9 or col_num == 10 or col_num == 11 or col_num == 12 or col_num == 15 or col_num == 16:
                    cell.alignment = Alignment(horizontal="center")

                #**************  Agrega el total de Menores  ************************
        row_num += 2
        worksheet.merge_cells('A'+str(row_num+1)+':B'+str(row_num+1))
        cell = worksheet.cell(row=row_num+1, column=1)
        cell.value = "Total de Menores: " +' ' + str(total_menores)
        cell2 = worksheet.cell(row=row_num+2, column=1)
        cell2.value = "Total de Activos: " +' ' + str(total_activos)
        cell3 = worksheet.cell(row=row_num+3, column=1)
        cell3.value = "Total de Altas: " +' ' + str(total_altas)
        cell4 = worksheet.cell(row=row_num+4, column=1)
        cell4.value = "Total de Egresos: " +' ' + str(total_egresos)
        cell5 = worksheet.cell(row=row_num+5, column=1)
        cell5.value = "Total de Desincorporados: " +' ' + str(total_desincorporados)
        cell6 = worksheet.cell(row=row_num+6, column=1)
        cell6.value = "Total de Menores de 5 años: " +' ' + str(total_menores_5)
        cell7 = worksheet.cell(row=row_num+7, column=1)
        cell7.value = "Total de Menores de 2 años: " +' ' + str(total_menores_2)


    #*********  Establecer el nombre del Archivo *******
        nombre_archvo = "Rep_listado_Menores.xlsx"
    
    #*********  Definir el tipo de respuesta que se va a dar ***********
        response = HttpResponse(content_type = "application/ms-excel")
        contenido = "attachment; filename = {0}".format(nombre_archvo)
        response["Content-Disposition"] = contenido
        workbook.save(response)
        return response
    


class exportar_jornada(TemplateView):
    def get(self, request, *args, **kwargs):
        proyecto = request.GET.get('proyecto')
        jornada = request.GET.get('jornada')
        workbook = Workbook()
        bandera = True
        fecha = datetime.now()
        fecha_fin = fecha.strftime('%d-%m-%Y - hora: %H:%m')
        
        hojas = ["REGISTRO MENORES","ANTROPOMETRICO MENOR","REGISTRO DE EMBARAZADAS","REGISTRO DE LACTANTES","MEDICAMENTOS Y PRODUCTOS"]

        for hoja in hojas: 
            if bandera:
                worksheet = workbook.active
                worksheet.title = hoja
                bandera = False
                worksheet.merge_cells('A1:C1')
                first_cell = worksheet['A1']
                first_cell.value = "Fecha: " + fecha_fin
            else:
                worksheet = workbook.create_sheet(hoja)

    #*********  Crear encabezado en la hoja  *************
            worksheet.merge_cells('A2:D2')
            second_cell = worksheet['A2']
            second_cell.value = "Bokitas"
            second_cell.font  = Font(name = 'Tw Cen MT Condensed Extra Bold', size = 52, bold = True, color="6F42C1")

            worksheet.merge_cells('A3:D3')
            third_cell = worksheet['A3']
            third_cell.value = "Bokitas Fundation    www.bokitas.org"
            third_cell.font  = Font(name = 'Tw Cen MT Condensed Extra Bold', size = 12, bold = True, color="6F42C1")
        
#***********************  HOJA DE DATOS DE LOS MENORES  ********************************

    #*********  Registro de Datos de los Menores  *************
        worksheet = workbook.worksheets[0]
        worksheet.merge_cells('A4:P6')
        fourth_cell = worksheet['A4']
        fourth_cell.value = "REGISTROS DEL PERFIL DE MENORES"
        fourth_cell.font  = Font(name = 'Tahoma', size = 16, bold = True, color="333399")
        fourth_cell.alignment = Alignment(horizontal="center", vertical="center")      

        menores = Menor.objects.filter(proyecto=proyecto,jornada=jornada).order_by('-cedula_bef')

    #**************  Obtener el total de Menores  ***************
        total_menores = menores.count()
        total_activos = menores.values("estatus").filter(estatus="ACTIVO").count()
        total_altas = menores.values("estatus").filter(estatus="ALTA").count()
        total_egresos = menores.values("estatus").filter(estatus="EGRESO").count()
        total_desincorporados = menores.values("estatus").filter(estatus="DESINCORPORADO").count()
        total_menores_5 = menores.values("edad").filter(edad__lte=5).count()
        total_menores_2 = menores.values("edad").filter(edad__lte=2).count()

        titulos = ['PROYECTO','REPRESENTANTE','PARENTESCO','CÉDULA','NOMBRE','APELLIDO','SEXO','FECHA NAC.','EDAD','MESES','PESO ACTUAL','TALLA ACTUAL','DIAGNOSTICO PESO','DIAGNOSTICO TALLA','FECHA INGRESO','ESTATUS']
        row_num = 7
        thin = Side(border_style="thin", color="000000")
        double = Side(border_style="double", color="000000")

    #********* asigna el titulo a las columnas  ************************
        for col_num, column_title in enumerate(titulos, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.fill = PatternFill("solid", fgColor="E2D9F3")
            cell.border = Border(top=thin, left=thin, right=thin, bottom=double)
            cell.font  = Font(bold=True, size = 14, color="333399")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            adjusted_width = (len(cell.value) + 10) * 1.2
            worksheet.column_dimensions[get_column_letter(col_num)].width = adjusted_width

    #**************  Agrega la data a las celdas
        for menor in menores:
            row_num += 1
            datos = [menor.proyecto,menor.cedula_bef,menor.parentesco,menor.cedula,menor.nombre,menor.apellido,menor.sexo,menor.fecha_nac.strftime('%d-%m-%Y'),menor.edad,menor.meses,menor.peso_actual,menor.talla_actual,menor.diagnostico_actual,menor.diagnostico_talla_actual,menor.fecha_ing_proyecto.strftime('%d-%m-%Y'),menor.estatus]

            for col_num, cell_value in enumerate(datos, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = str(cell_value)
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                if col_num == 3 or col_num == 8 or col_num == 9 or col_num == 10 or col_num == 11 or col_num == 12 or col_num == 15 or col_num == 16:
                    cell.alignment = Alignment(horizontal="center")

        #**************  Agrega el total de Menores  ************************
        row_num += 2
        worksheet.merge_cells('A'+str(row_num+1)+':B'+str(row_num+1))
        cell = worksheet.cell(row=row_num+1, column=1)
        cell.value = "Total de Menores: " +' ' + str(total_menores)
        cell2 = worksheet.cell(row=row_num+2, column=1)
        cell2.value = "Total de Activos: " +' ' + str(total_activos)
        cell3 = worksheet.cell(row=row_num+3, column=1)
        cell3.value = "Total de Altas: " +' ' + str(total_altas)
        cell4 = worksheet.cell(row=row_num+4, column=1)
        cell4.value = "Total de Egresos: " +' ' + str(total_egresos)
        cell5 = worksheet.cell(row=row_num+5, column=1)
        cell5.value = "Total de Desincorporados: " +' ' + str(total_desincorporados)
        cell6 = worksheet.cell(row=row_num+6, column=1)
        cell6.value = "Total de Menores de 5 años: " +' ' + str(total_menores_5)
        cell7 = worksheet.cell(row=row_num+7, column=1)
        cell7.value = "Total de Menores de 2 años: " +' ' + str(total_menores_2)

#*************  HOJA DE DATOS ANTROPOMETRICOS DEL MENORES  *********************

    #*********  Registro de Datos antropometricos Menores  *************
        worksheet = workbook.worksheets[1]
        worksheet.merge_cells('A4:W6')
        fourth_cell = worksheet['A4']
        fourth_cell.value = "REGISTRO DATOS ANTROPOMETRICOS DEL MENOR"
        fourth_cell.font  = Font(name = 'Tahoma', size = 16, bold = True, color="333399")
        fourth_cell.alignment = Alignment(horizontal="center", vertical="center")      

        imc_menores = AntropMenor.objects.filter(proyecto_id=proyecto,jornada=jornada).order_by('-cedula_bef','cedula_id').select_related('cedula')

        titulos = ['PROYECTO','REPRESENTANTE','CÉDULA','NOMBRE','APELLIDO','SEXO','FECHA NAC.','EDAD','MESES','PESO ACTUAL','TALLA ACTUAL','FECHA JORNADA','EDAD','MESES','PESO','TALLA','CBI','PTR','PSE','CC','IMC','DIAGNOSTICO PESO','DIAGNOSTICO TALLA']
        row_num = 7
        thin = Side(border_style="thin", color="000000")
        double = Side(border_style="double", color="000000")

    #********* asigna el titulo a las columnas  ************************
        for col_num, column_title in enumerate(titulos, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.fill = PatternFill("solid", fgColor="E2D9F3")
            cell.border = Border(top=thin, left=thin, right=thin, bottom=double)
            cell.font  = Font(bold=True, size = 14, color="333399")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            adjusted_width = (len(cell.value) + 10) * 1.2
            worksheet.column_dimensions[get_column_letter(col_num)].width = adjusted_width

    #**************  Agrega la data a las celdas
        xCedula = 0

        for imc in imc_menores:
            row_num += 1
            datos = [imc.proyecto,imc.cedula_bef,imc.cedula.cedula,imc.cedula.nombre,imc.cedula.apellido,imc.cedula.sexo,imc.cedula.fecha_nac.strftime('%d-%m-%Y'),imc.edad,imc.meses,imc.cedula.peso_actual,imc.cedula.talla_actual,imc.jornada,imc.edad,imc.meses,imc.peso,imc.talla,imc.cbi,imc.ptr,imc.pse,imc.cc,imc.imc,imc.diagnostico,imc.diagnostico_talla]

            if imc.cedula.cedula != xCedula:
                for col_num, cell_value in enumerate(datos, 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.value = str(cell_value)
                    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                    cell.alignment = Alignment(horizontal="center")
                    if col_num == 1 or col_num == 2 or col_num == 4 or col_num == 5 or col_num == 22 or col_num == 23:
                        cell.alignment = Alignment(horizontal="left")

                xCedula = imc.cedula.cedula
            else:
                for col_num, cell_value in enumerate(datos, 1):
                    if col_num > 11:
                        cell = worksheet.cell(row=row_num, column=col_num)
                        cell.value = str(cell_value)
                        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                        cell.alignment = Alignment(horizontal="center")
                        if col_num == 22 or col_num == 23:
                            cell.alignment = Alignment(horizontal="left")



#*************  HOJA DE DATOS ANTROPOMETRICOS EMBARAZADAS  *********************

    #*********  Registro de Datos antropometricos Embarazadas  *************
        worksheet = workbook.worksheets[2]
        worksheet.merge_cells('A4:P6')
        fourth_cell = worksheet['A4']
        fourth_cell.value = "REGISTRO DATOS ANTROPOMETRICOS EMBARAZADAS"
        fourth_cell.font  = Font(name = 'Tahoma', size = 16, bold = True, color="333399")
        fourth_cell.alignment = Alignment(horizontal="center", vertical="center")      

        imc_embarazos = AntropBef.objects.filter(proyecto=proyecto, jornada=jornada, embarazo_lactando = 'EMBARAZADA').order_by('-cedula_bef').select_related('cedula_bef')

        titulos = ['PROYECTO','CÉDULA','NOMBRE','APELLIDO','FECHA NAC.','FECHA JORNADA','EDAD','MESES','MESES EMBARAZO','PESO','TALLA','CBI','IMC','DIAGNOSTICO PESO','RIESGO','OBSERVACIONES']
        row_num = 7
        thin = Side(border_style="thin", color="000000")
        double = Side(border_style="double", color="000000")

    #********* asigna el titulo a las columnas  ************************
        for col_num, column_title in enumerate(titulos, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.fill = PatternFill("solid", fgColor="E2D9F3")
            cell.border = Border(top=thin, left=thin, right=thin, bottom=double)
            cell.font  = Font(bold=True, size = 14, color="333399")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            adjusted_width = (len(cell.value) + 10) * 1.2
            worksheet.column_dimensions[get_column_letter(col_num)].width = adjusted_width

    #**************  Agrega la data a las celdas
        xCedula = 0

        for imc in imc_embarazos:
            row_num += 1
            datos = [imc.cedula_bef.proyecto,imc.cedula_bef.cedula,imc.cedula_bef.nombre,imc.cedula_bef.apellido,imc.cedula_bef.fecha_nac.strftime('%d-%m-%Y'),imc.jornada,imc.edad,imc.meses,imc.tiempo_gestacion,imc.peso,imc.talla,imc.cbi,imc.imc,imc.diagnostico,imc.riesgo,imc.observacion]

            if imc.cedula_bef.cedula != xCedula:
                for col_num, cell_value in enumerate(datos, 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.value = str(cell_value)
                    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                    cell.alignment = Alignment(horizontal="center")
                    if col_num == 1 or col_num == 2 or col_num == 3 or col_num == 4:
                        cell.alignment = Alignment(horizontal="left")

                xCedula = imc.cedula_bef.cedula
            else:
                for col_num, cell_value in enumerate(datos, 1):
                    if col_num > 7:
                        cell = worksheet.cell(row=row_num, column=col_num)
                        cell.value = str(cell_value)
                        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                        cell.alignment = Alignment(horizontal="center")


#*************  HOJA DE DATOS ANTROPOMETRICOS LACTANTES  *********************

    #*********  Registro de Datos antropometricos Lactante  *************
        worksheet = workbook.worksheets[3]
        worksheet.merge_cells('A4:P6')
        fourth_cell = worksheet['A4']
        fourth_cell.value = "REGISTRO DATOS ANTROPOMETRICOS LACTANTES"
        fourth_cell.font  = Font(name = 'Tahoma', size = 16, bold = True, color="333399")
        fourth_cell.alignment = Alignment(horizontal="center", vertical="center")      

        imc_embarazos = AntropBef.objects.filter(proyecto=proyecto,jornada=jornada, embarazo_lactando = 'LACTANDO').order_by('-cedula_bef').select_related('cedula_bef')

        titulos = ['PROYECTO','CÉDULA','NOMBRE','APELLIDO','FECHA NAC.','FECHA JORNADA','EDAD','MESES','MESES LACTANDO','PESO','TALLA','CBI','IMC','DIAGNOSTICO PESO','RIESGO','OBSERVACIONES']
        row_num = 7
        thin = Side(border_style="thin", color="000000")
        double = Side(border_style="double", color="000000")

    #********* asigna el titulo a las columnas  ************************
        for col_num, column_title in enumerate(titulos, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.fill = PatternFill("solid", fgColor="E2D9F3")
            cell.border = Border(top=thin, left=thin, right=thin, bottom=double)
            cell.font  = Font(bold=True, size = 14, color="333399")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            adjusted_width = (len(cell.value) + 10) * 1.2
            worksheet.column_dimensions[get_column_letter(col_num)].width = adjusted_width

    #**************  Agrega la data a las celdas
        xCedula = 0

        for imc in imc_embarazos:
            row_num += 1
            datos = [imc.cedula_bef.proyecto,imc.cedula_bef.cedula,imc.cedula_bef.nombre,imc.cedula_bef.apellido,imc.cedula_bef.fecha_nac.strftime('%d-%m-%Y'),imc.jornada,imc.edad,imc.meses,imc.tiempo_gestacion,imc.peso,imc.talla,imc.cbi,imc.imc,imc.diagnostico,imc.riesgo,imc.observacion]

            if imc.cedula_bef.cedula != xCedula:
                for col_num, cell_value in enumerate(datos, 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.value = str(cell_value)
                    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                    cell.alignment = Alignment(horizontal="center")
                    if col_num == 1 or col_num == 2 or col_num == 3 or col_num == 4:
                        cell.alignment = Alignment(horizontal="left")

                xCedula = imc.cedula_bef.cedula
            else:
                for col_num, cell_value in enumerate(datos, 1):
                    if col_num > 7:
                        cell = worksheet.cell(row=row_num, column=col_num)
                        cell.value = str(cell_value)
                        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                        cell.alignment = Alignment(horizontal="center")


#*************  HOJA DE DATOS ENTREGA MEDICAMENTOS PRODUCTOS  *********************

    #*********  Registro de Datos DATOS ENTREGA MEDICAMENTOS  *************
        worksheet = workbook.worksheets[4]
        worksheet.merge_cells('A4:H6')
        fourth_cell = worksheet['A4']
        fourth_cell.value = "REGISTRO DE ENTREGA DE MEDICAMENTOS Y PRODUCTOS"
        fourth_cell.font  = Font(name = 'Tahoma', size = 16, bold = True, color="333399")
        fourth_cell.alignment = Alignment(horizontal="center", vertical="center")      

        medicamentos = Medicamento.objects.filter(proyecto=proyecto,jornada=jornada).order_by('-cedula_bef').select_related('cedula_bef')

        titulos = ['PROYECTO','CÉDULA','NOMBRE','APELLIDO','FECHA JORNADA','NOMBRE DEL PRODUCTO','DESCRIPCION DEL PRODUCTO','CANTIDAD']
        row_num = 7
        thin = Side(border_style="thin", color="000000")
        double = Side(border_style="double", color="000000")

    #********* asigna el titulo a las columnas  ************************
        for col_num, column_title in enumerate(titulos, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.fill = PatternFill("solid", fgColor="E2D9F3")
            cell.border = Border(top=thin, left=thin, right=thin, bottom=double)
            cell.font  = Font(bold=True, size = 14, color="333399")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            adjusted_width = (len(cell.value) + 10) * 1.2
            worksheet.column_dimensions[get_column_letter(col_num)].width = adjusted_width

    #**************  Agrega la data a las celdas

        for med in medicamentos:
            row_num += 1
            datos = [med.cedula_bef.proyecto,med.cedula_bef.cedula,med.cedula_bef.nombre,med.cedula_bef.apellido,med.jornada,med.nombre,med.descripcion,med.cantidad]

            for col_num, cell_value in enumerate(datos, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = str(cell_value)
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                cell.alignment = Alignment(horizontal="center")
                if col_num == 1 or col_num == 2 or col_num == 3 or col_num == 4:
                    cell.alignment = Alignment(horizontal="left")


    #*********  Establecer el nombre del Archivo *******
        nombre_archvo = "Reporte_por_Jornmada.xlsx"
    
    #*********  Definir el tipo de respuesta que se va a dar ***********
        response = HttpResponse(content_type = "application/ms-excel")
        contenido = "attachment; filename = {0}".format(nombre_archvo)
        response["Content-Disposition"] = contenido
        workbook.save(response)
        return response
    


class estadistica_nutricional_proyecto(TemplateView):
    def get(self, request, *args, **kwargs):
        proyecto = request.GET.get('proyecto')
        nutricional = Nutricional.objects.filter(proyecto=proyecto)
        nombre_proyecto = get_object_or_404(Proyecto, id=proyecto)
        total_encuestados = nutricional.count()
        wb = Workbook()
        ws = wb.active
        ws.title = "ESTADISTICA NUTRICIONAL"
        thin = Side(border_style="thin", color="000000")

    #*********  Crear encabezado en la hoja  *************

        for column in ('A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U'): 
            ws.column_dimensions[column].width = 20
    

        fecha = datetime.now()
        fecha_fin = fecha.strftime('%d-%m-%Y - hora: %H:%m')
        ws.merge_cells('A1:C1')
        first_cell = ws['A1']
        first_cell.value = "Fecha: " + fecha_fin
        
        ws.merge_cells('A2:E2')
        second_cell = ws['A2']
        second_cell.value = "Bokitas"
        second_cell.font  = Font(name = 'Tw Cen MT Condensed Extra Bold', size = 52, bold = True, color="6F42C1")

        ws.merge_cells('A3:E3')
        third_cell = ws['A3']
        third_cell.value = "Bokitas Fundation    www.bokitas.org"
        third_cell.font  = Font(name = 'Tw Cen MT Condensed Extra Bold', size = 12, bold = True, color="6F42C1")
        
#***********************  DATOS ESTADISTICA NUTRICIONAL  ********************************

        #*********  Tutulo Principal  *************
        ws.merge_cells('A4:P5')
        fourth_cell = ws['A4']
        fourth_cell.value = "ESTADISTICA NUTRICIONAL DEL PROYECTO: " + str(nombre_proyecto)
        fourth_cell.font  = Font(name = 'Tahoma', size = 14, bold = True, color="333399")
        fourth_cell.fill = PatternFill("solid", fgColor="E2D9F3")
        fourth_cell.alignment = Alignment(horizontal="center", vertical="center") 

        ws.merge_cells('A6:P6')
        fourth_cell = ws['A6']
        fourth_cell.value = "TOTAL DE ENCUESTADOS: " + str(total_encuestados)
        fourth_cell.font  = Font(name = 'Tahoma', size = 14, bold = True, color="333399")
        fourth_cell.fill = PatternFill("solid", fgColor="E2D9F3")
        fourth_cell.alignment = Alignment(horizontal="center", vertical="center") 
    
    
#**********************       TITULO DE LA PRIMERA PREGUNTA    *********************
        ws.merge_cells('A8:P9')
        fourth_cell = ws['A8']
        fourth_cell.value = "RESPONSABLE DE HACER EL MERCADO, COCINAR Y FRECUENCIA DE COMPRA DE ALIMENTOS"
        fourth_cell.font  = Font(name = 'Tahoma', size = 12, bold = True, color="333399")
        fourth_cell.fill = PatternFill("solid", fgColor="E6E6FA")
        fourth_cell.alignment = Alignment(horizontal="center", vertical="center")      

    #********* asigna el titulo a los graficos  ************************
        ws.merge_cells('D11:G11')
        five_cell = ws['D11']
        five_cell.value = "RESPONSABLE DE HACER EL MERCADO Y COCINAR"
        five_cell.font  = Font(name = 'Tahoma', size = 12, bold = True, color="333399")
        five_cell.fill = PatternFill("solid", fgColor="C6E2FF")
        five_cell.alignment = Alignment(horizontal="center", vertical="center")  

        ws.merge_cells('J11:M11')
        six_cell = ws['J11']
        six_cell.value = "FRECUENCIA DE LA COMPRA DE ALIMENTOS"
        six_cell.font  = Font(name = 'Tahoma', size = 12, bold = True, color="333399")
        six_cell.fill = PatternFill("solid", fgColor="C6E2FF")
        six_cell.alignment = Alignment(horizontal="center", vertical="center") 

    #**************  Obtener el total de las encuestas  ***************

        padre_mercado = nutricional.values("mercado_lorealiza").filter(mercado_lorealiza__icontains="Padre").count()
        madre_mercado = nutricional.values("mercado_lorealiza").filter(mercado_lorealiza__icontains="Madre").count()
        abuelo_mercado = nutricional.values("mercado_lorealiza").filter(mercado_lorealiza__icontains="Abuelo(a)").count()
        tio_mercado = nutricional.values("mercado_lorealiza").filter(mercado_lorealiza__icontains="Tio(a)").count()
        otros_mercado = nutricional.values("mercado_lorealiza").filter(mercado_lorealiza__icontains="Otros").count()

        padre_cocina = nutricional.values("cocina_lorealiza").filter(cocina_lorealiza__icontains="Padre").count()
        madre_cocina = nutricional.values("cocina_lorealiza").filter(cocina_lorealiza__icontains="Madre").count()
        abuelo_cocina = nutricional.values("cocina_lorealiza").filter(cocina_lorealiza__icontains="Abuelo(a)").count()
        tio_cocina = nutricional.values("cocina_lorealiza").filter(cocina_lorealiza__icontains="Tio(a)").count()
        otros_cocina = nutricional.values("cocina_lorealiza").filter(cocina_lorealiza__icontains="Otros").count()
        
        fecuencia_diario = nutricional.values("frecuencia").filter(frecuencia__icontains="diario").count()
        fecuencia_2_3 = nutricional.values("frecuencia").filter(frecuencia__icontains="2-3 días").count()
        fecuencia_semanal = nutricional.values("frecuencia").filter(frecuencia__icontains="Semanal").count()
        fecuencia_quincenal = nutricional.values("frecuencia").filter(frecuencia__icontains="Quincenal").count()


    #**************  Agrega la data a las celdas  ***************************

    #**************  RESPONSABLE DE HACER EL MERCADO Y COCINAR  *********
        datos = [
                ('Integrantes', 'Quíen Hace Mercado', 'Quíen Cocina'),
                ('Madre', madre_mercado, madre_cocina),
                ('Padre', padre_mercado, padre_cocina),
                ('Abuelos', abuelo_mercado, abuelo_cocina),
                ('Tios', tio_mercado, tio_cocina),
                ('Otros', otros_mercado, otros_cocina),
            ]
        for row, cell_value in enumerate(datos, 12):
            ws.cell(row=row+1, column=13)
            for col_num, cell_value in enumerate(cell_value, 4):
                cell = ws.cell(row=row+1, column=col_num)
                cell.value = cell_value
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                cell.alignment = Alignment(horizontal="center", vertical="center")
        
        chart1 = BarChart3D()
        chart1.type = "col"
        chart1.style = 10
        chart1.title = "RESPONSABLE DE HACER EL MERCADO Y COCINAR"
        chart1.y_axis.title = 'Número de Casos'
        chart1.x_axis.title = 'Integrantes'
        chart1.width = 15  # Ancho en pulgadas  
        chart1.height = 8  # Altura en pulgadas

        data = Reference(ws, min_col=5, min_row=13, max_row=18, max_col=6)
        cats = Reference(ws, min_col=4, min_row=14, max_row=18)
        chart1.add_data(data, titles_from_data=True)
        chart1.set_categories(cats)
    
        ws.add_chart(chart1, "D19")


    #**************  FRECUENCIA DE LA COMPRA DE ALIMENTOS  *********

        #********* asigna los Datos a las celdas  ************************
        datos = [
                ('Compras', 'Frecuencia de la Compra'),
                ('Diaria', fecuencia_diario),
                ('2-3 días', fecuencia_2_3),
                ('Semanal', fecuencia_semanal),
                ('Quincenal', fecuencia_quincenal),
            ]
        
        for row, cell_value in enumerate(datos, 12):
            ws.cell(row=row+1, column=13)
            for col_num, cell_value in enumerate(cell_value, 11):
                cell = ws.cell(row=row+1, column=col_num)
                cell.value = cell_value
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                cell.alignment = Alignment(horizontal="center", vertical="center")

        chart2 = BarChart3D()
        chart2.type = "col"
        chart2.style = 10
        chart2.title = "FRECUENCIA DE LA COMPRA DE ALIMENTOS"
        chart2.y_axis.title = 'Número de Casos'
        chart2.x_axis.title = 'Frecuencia de la Compra'
        chart2.width = 15  # Ancho en pulgadas  
        chart2.height = 8  # Altura en pulgadas

        data = Reference(ws, min_col=12, min_row=13, max_row=17)
        cats = Reference(ws, min_col=11, min_row=14, max_row=17)
        chart2.add_data(data, titles_from_data=True)
        chart2.set_categories(cats)
    
        ws.add_chart(chart2, "J19")


#**********************       TITULO DE LA SEGUNDA PREGUNTA    *********************
        ws.merge_cells('A35:P36')
        fourth_cell = ws['A35']
        fourth_cell.value = "APETITO, COMIDAS Y MERIENDAS"
        fourth_cell.font  = Font(name = 'Tahoma', size = 12, bold = True, color="333399")
        fourth_cell.fill = PatternFill("solid", fgColor="E6E6FA")
        fourth_cell.alignment = Alignment(horizontal="center", vertical="center")      

    #********* asigna el titulo a los graficos  ************************
        ws.merge_cells('B38:E38')
        five_cell = ws['B38']
        five_cell.value = "COMO ES EL APETITO"
        five_cell.font  = Font(name = 'Tahoma', size = 12, bold = True, color="333399")
        five_cell.fill = PatternFill("solid", fgColor="C6E2FF")
        five_cell.alignment = Alignment(horizontal="center", vertical="center")  

        ws.merge_cells('G38:J38')
        six_cell = ws['G38']
        six_cell.value = "COMIDAS AL DIA"
        six_cell.font  = Font(name = 'Tahoma', size = 12, bold = True, color="333399")
        six_cell.fill = PatternFill("solid", fgColor="C6E2FF")
        six_cell.alignment = Alignment(horizontal="center", vertical="center") 

        ws.merge_cells('L38:O38')
        seven_cell = ws['L38']
        seven_cell.value = "CUANTAS MERIENDAS"
        seven_cell.font  = Font(name = 'Tahoma', size = 12, bold = True, color="333399")
        seven_cell.fill = PatternFill("solid", fgColor="C6E2FF")
        seven_cell.alignment = Alignment(horizontal="center", vertical="center") 

    #**************  Obtener el total de las encuestas  ***************

        apetito_bueno = nutricional.values("apetito").filter(apetito__icontains="Bueno").count()
        apetito_regular = nutricional.values("apetito").filter(apetito__icontains="Regular").count()
        apetito_malo = nutricional.values("apetito").filter(apetito__icontains="Malo").count()
        apetito_elevado = nutricional.values("apetito").filter(apetito__icontains="Elevado").count()

        comidas_una = nutricional.values("cuantas_comidas").filter(cuantas_comidas__icontains="1 Comida").count()
        comidas_dos = nutricional.values("cuantas_comidas").filter(cuantas_comidas__icontains="2 Comidas").count()
        comidas_tres = nutricional.values("cuantas_comidas").filter(cuantas_comidas__icontains="3 Comidas").count()

        meriendas_una = nutricional.values("meriendas").filter(meriendas__icontains="1 Merienda").count()
        meriendas_dos = nutricional.values("meriendas").filter(meriendas__icontains="2 Meriendas").count()
        meriendas_tres = nutricional.values("meriendas").filter(meriendas__icontains="3 Meriendas").count()
        meriendas_ninguna = nutricional.values("meriendas").filter(meriendas__icontains="Ninguna").count()

    #**************  Agrega la data a las celdas  ***************************

    #**************  COMO ES EL APETITO  *********
        datos = [
                ('Apetito', 'Resultados'),
                ('Bueno', apetito_bueno),
                ('Regular', apetito_regular),
                ('Malo', apetito_malo),
                ('Elevado', apetito_elevado),
            ]
        
        for row, cell_value in enumerate(datos, 39):
            ws.cell(row=row+1, column=13)
            for col_num, cell_value in enumerate(cell_value, 3):
                cell = ws.cell(row=row+1, column=col_num)
                cell.value = cell_value
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                

        chart2 = BarChart3D()
        chart2.type = "col"
        chart2.style = 10
        chart2.title = "FRECUENCIA DEL APETITO"
        chart2.y_axis.title = 'Número de Casos'
        chart2.x_axis.title = 'Frecuencia del Apetito'
        chart2.width = 15  # Ancho en pulgadas  
        chart2.height = 8  # Altura en pulgadas

        data = Reference(ws, min_col=4, min_row=40, max_row=44)
        cats = Reference(ws, min_col=3, min_row=41, max_row=44)
        chart2.add_data(data, titles_from_data=True)
        chart2.set_categories(cats)
    
        ws.add_chart(chart2, "B46")

    #**************  COMIDAS AL DIA  *********
        datos = [
                ('Comidas', 'Resultados'),
                ('1 Comida', comidas_una),
                ('2 Comidas', comidas_dos),
                ('3 Comidas', comidas_tres),
            ]
        
        for row, cell_value in enumerate(datos, 39):
            ws.cell(row=row+1, column=13)
            for col_num, cell_value in enumerate(cell_value, 8):
                cell = ws.cell(row=row+1, column=col_num)
                cell.value = cell_value
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                

        chart2 = BarChart3D()
        chart2.type = "col"
        chart2.style = 10
        chart2.title = "COMIDAS AL DIA"
        chart2.y_axis.title = 'Número de Casos'
        chart2.x_axis.title = 'Frecuencia de Comidas al Día'
        chart2.width = 15  # Ancho en pulgadas  
        chart2.height = 8  # Altura en pulgadas

        data = Reference(ws, min_col=9, min_row=40, max_row=43)
        cats = Reference(ws, min_col=8, min_row=41, max_row=43)
        chart2.add_data(data, titles_from_data=True)
        chart2.set_categories(cats)
    
        ws.add_chart(chart2, "G46")


    #**************  CUANTAS MERIENDAS  *********
        datos = [
                ('Meriendas', 'Frecuencia de la Merienda'),
                ('1 Merienda', meriendas_una),
                ('2 Meriendas', meriendas_dos),
                ('3 Meriendas', meriendas_tres),
                ('Ninguna', meriendas_ninguna),
            ]
        
        for row, cell_value in enumerate(datos, 39):
            ws.cell(row=row+1, column=13)
            for col_num, cell_value in enumerate(cell_value, 13):
                cell = ws.cell(row=row+1, column=col_num)
                cell.value = cell_value
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                cell.alignment = Alignment(horizontal="center", vertical="center")

        chart2 = BarChart3D()
        chart2.type = "col"
        chart2.style = 10
        chart2.title = "FRECUENCIA DE LAS MERIENDAS"
        chart2.y_axis.title = 'Número de Casos'
        chart2.x_axis.title = 'Frecuencia de la Merienda'
        chart2.width = 15  # Ancho en pulgadas  
        chart2.height = 8  # Altura en pulgadas

        data = Reference(ws, min_col=14, min_row=40, max_row=44)
        cats = Reference(ws, min_col=13, min_row=41, max_row=44)
        chart2.add_data(data, titles_from_data=True)
        chart2.set_categories(cats)
    
        ws.add_chart(chart2, "L46")


#**********************       TITULO DE LA TERCESA PREGUNTA    *********************
        ws.merge_cells('A62:P63')
        fourth_cell = ws['A62']
        fourth_cell.value = "GRUPOS DE ALIMENTOS"
        fourth_cell.font  = Font(name = 'Tahoma', size = 12, bold = True, color="333399")
        fourth_cell.fill = PatternFill("solid", fgColor="E6E6FA")
        fourth_cell.alignment = Alignment(horizontal="center", vertical="center")      

    #********* asigna el titulo a los graficos  ************************
        ws.merge_cells('D65:G65')
        five_cell = ws['D65']
        five_cell.value = "CUANTOS GRUPOS ESTAN PRESENTES EN LA ALIMENTACION"
        five_cell.font  = Font(name = 'Tahoma', size = 12, bold = True, color="333399")
        five_cell.fill = PatternFill("solid", fgColor="C6E2FF")
        five_cell.alignment = Alignment(horizontal="center", vertical="center")  

        ws.merge_cells('J65:M65')
        six_cell = ws['J65']
        six_cell.value = "CUALES SON LOS GRUPOS PRESENTES"
        six_cell.font  = Font(name = 'Tahoma', size = 12, bold = True, color="333399")
        six_cell.fill = PatternFill("solid", fgColor="C6E2FF")
        six_cell.alignment = Alignment(horizontal="center", vertical="center") 

    #**************  Obtener el total de las encuestas  ***************

        grupo_1_2 = nutricional.values("cuantos_grupos").filter(cuantos_grupos__icontains="1 - 2").count()
        grupo_2_3 = nutricional.values("cuantos_grupos").filter(cuantos_grupos__icontains="2 - 3").count()
        grupo_3_4 = nutricional.values("cuantos_grupos").filter(cuantos_grupos__icontains="3 - 4").count()
        grupo_4_5 = nutricional.values("cuantos_grupos").filter(cuantos_grupos__icontains="4 - 5").count()
        grupo_5_6 = nutricional.values("cuantos_grupos").filter(cuantos_grupos__icontains="5 - 6").count()

        grupo_1 = nutricional.values("cuantos_grupos").filter(tipo_grupos__icontains="Cereales y Leguminosas").count()
        grupo_2 = nutricional.values("cuantos_grupos").filter(tipo_grupos__icontains="Frutas y Verduras").count()
        grupo_3 = nutricional.values("cuantos_grupos").filter(tipo_grupos__icontains="Leche Yogures y Quesos").count()
        grupo_4 = nutricional.values("cuantos_grupos").filter(tipo_grupos__icontains="Grasas y Aceites").count()
        grupo_5 = nutricional.values("cuantos_grupos").filter(tipo_grupos__icontains="Dulces").count()
        grupo_6 = nutricional.values("cuantos_grupos").filter(tipo_grupos__icontains="Carnes y Huevos").count()

    #**************  Agrega la data a las celdas  ***************************

    #**************  CUANTOS GRUPOS ESTAN PRESENTES EN LA ALIMENTACION  *********
        datos = [
                ('Grupos', 'Presentes'),
                ('1 - 2', grupo_1_2),
                ('2 - 3', grupo_2_3),
                ('3 - 4', grupo_3_4),
                ('4 - 5', grupo_4_5),
                ('5 - 6', grupo_5_6),
            ]
        for row, cell_value in enumerate(datos, 66):
            ws.cell(row=row+1, column=13)
            for col_num, cell_value in enumerate(cell_value, 5):
                cell = ws.cell(row=row+1, column=col_num)
                cell.value = cell_value
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                cell.alignment = Alignment(horizontal="center", vertical="center")
        
        chart1 = BarChart3D()
        chart1.type = "col"
        chart1.style = 10
        chart1.title = "GRUPOS PRESENTES EN LA ALIMENTACION"
        chart1.y_axis.title = 'Número de Casos'
        chart1.x_axis.title = 'Grupos de Alimentos'
        chart1.width = 15  # Ancho en pulgadas  
        chart1.height = 8  # Altura en pulgadas

        data = Reference(ws, min_col=6, min_row=67, max_row=72)
        cats = Reference(ws, min_col=5, min_row=68, max_row=72)
        chart1.add_data(data, titles_from_data=True)
        chart1.set_categories(cats)
    
        ws.add_chart(chart1, "D73")


    #**************  FRECUENCIA DEL CONSUMO DE LOS GRUPOS  *********

        #********* asigna los Datos a las celdas  ************************
        datos = [
                ('Grupos', 'Grupos Presentes'),
                ('Grupo 1', grupo_1),
                ('Grupo 2', grupo_2),
                ('Grupo 3', grupo_3),
                ('Grupo 4', grupo_4),
                ('Grupo 5', grupo_5),
                ('Grupo 6', grupo_6),
            ]
        
        for row, cell_value in enumerate(datos, 65):
            ws.cell(row=row+1, column=13)
            for col_num, cell_value in enumerate(cell_value, 11):
                cell = ws.cell(row=row+1, column=col_num)
                cell.value = cell_value
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                cell.alignment = Alignment(horizontal="center", vertical="center")

        chart2 = BarChart3D()
        chart2.type = "col"
        chart2.style = 10
        chart2.title = "CUALES GRUPOS SE PRESENTAN EN LA ALIMENTACION"
        chart2.y_axis.title = 'Número de Casos'
        chart2.x_axis.title = 'Frecuencia de la Grupos'
        chart2.width = 15  # Ancho en pulgadas  
        chart2.height = 8  # Altura en pulgadas

        data = Reference(ws, min_col=12, min_row=66, max_row=72)
        cats = Reference(ws, min_col=11, min_row=67, max_row=72)
        chart2.add_data(data, titles_from_data=True)
        chart2.set_categories(cats)
    
        ws.add_chart(chart2, "J73")


#**********************       TITULO DE LA CUARTA PREGUNTA    *********************
        ws.merge_cells('A89:P90')
        fourth_cell = ws['A89']
        fourth_cell.value = "FRECUENCIA DEL CONSUMO DE LOS ALIMENTOS"
        fourth_cell.font  = Font(name = 'Tahoma', size = 12, bold = True, color="333399")
        fourth_cell.fill = PatternFill("solid", fgColor="E6E6FA")
        fourth_cell.alignment = Alignment(horizontal="center", vertical="center")      

    #********* asigna el titulo a los graficos  ************************
        ws.merge_cells('E92:L92')
        five_cell = ws['E92']
        five_cell.value = "FRECUENCIA DEL CONSUMO"
        five_cell.font  = Font(name = 'Tahoma', size = 12, bold = True, color="333399")
        five_cell.fill = PatternFill("solid", fgColor="C6E2FF")
        five_cell.alignment = Alignment(horizontal="center", vertical="center")  

    #**************  Obtener el total de las encuestas  ***************

        cereales_1 = nutricional.values("cereales").filter(cereales__icontains="0/7 días").count()
        cereales_2 = nutricional.values("cereales").filter(cereales__icontains="1-2/7 días").count()
        cereales_3 = nutricional.values("cereales").filter(cereales__icontains="3-5/7 días").count()
        cereales_4 = nutricional.values("cereales").filter(cereales__icontains="7/7 días").count()
        cereales_5 = nutricional.values("cereales").filter(cereales__icontains="1-2/15 días").count()
        cereales_6 = nutricional.values("cereales").filter(cereales__icontains="1-2/30 días").count()

        vegetales_1 = nutricional.values("vegetales").filter(vegetales__icontains="0/7 días").count()
        vegetales_2 = nutricional.values("vegetales").filter(vegetales__icontains="1-2/7 días").count()
        vegetales_3 = nutricional.values("vegetales").filter(vegetales__icontains="3-5/7 días").count()
        vegetales_4 = nutricional.values("vegetales").filter(vegetales__icontains="7/7 días").count()
        vegetales_5 = nutricional.values("vegetales").filter(vegetales__icontains="1-2/15 días").count()
        vegetales_6 = nutricional.values("vegetales").filter(vegetales__icontains="1-2/30 días").count()

        frutas_1 = nutricional.values("frutas").filter(frutas__icontains="0/7 días").count()
        frutas_2 = nutricional.values("frutas").filter(frutas__icontains="1-2/7 días").count()
        frutas_3 = nutricional.values("frutas").filter(frutas__icontains="3-5/7 días").count()
        frutas_4 = nutricional.values("frutas").filter(frutas__icontains="7/7 días").count()
        frutas_5 = nutricional.values("frutas").filter(frutas__icontains="1-2/15 días").count()
        frutas_6 = nutricional.values("frutas").filter(frutas__icontains="1-2/30 días").count()

        carnes_1 = nutricional.values("carnes").filter(carnes__icontains="0/7 días").count()
        carnes_2 = nutricional.values("carnes").filter(carnes__icontains="1-2/7 días").count()
        carnes_3 = nutricional.values("carnes").filter(carnes__icontains="3-5/7 días").count()
        carnes_4 = nutricional.values("carnes").filter(carnes__icontains="7/7 días").count()
        carnes_5 = nutricional.values("carnes").filter(carnes__icontains="1-2/15 días").count()
        carnes_6 = nutricional.values("carnes").filter(carnes__icontains="1-2/30 días").count()

        pollo_1 = nutricional.values("pollo").filter(pollo__icontains="0/7 días").count()
        pollo_2 = nutricional.values("pollo").filter(pollo__icontains="1-2/7 días").count()
        pollo_3 = nutricional.values("pollo").filter(pollo__icontains="3-5/7 días").count()
        pollo_4 = nutricional.values("pollo").filter(pollo__icontains="7/7 días").count()
        pollo_5 = nutricional.values("pollo").filter(pollo__icontains="1-2/15 días").count()
        pollo_6 = nutricional.values("pollo").filter(pollo__icontains="1-2/30 días").count()

        pescado_1 = nutricional.values("pescado").filter(pescado__icontains="0/7 días").count()
        pescado_2 = nutricional.values("pescado").filter(pescado__icontains="1-2/7 días").count()
        pescado_3 = nutricional.values("pescado").filter(pescado__icontains="3-5/7 días").count()
        pescado_4 = nutricional.values("pescado").filter(pescado__icontains="7/7 días").count()
        pescado_5 = nutricional.values("pescado").filter(pescado__icontains="1-2/15 días").count()
        pescado_6 = nutricional.values("pescado").filter(pescado__icontains="1-2/30 días").count()

        embutidos_1 = nutricional.values("embutidos").filter(embutidos__icontains="0/7 días").count()
        embutidos_2 = nutricional.values("embutidos").filter(embutidos__icontains="1-2/7 días").count()
        embutidos_3 = nutricional.values("embutidos").filter(embutidos__icontains="3-5/7 días").count()
        embutidos_4 = nutricional.values("embutidos").filter(embutidos__icontains="7/7 días").count()
        embutidos_5 = nutricional.values("embutidos").filter(embutidos__icontains="1-2/15 días").count()
        embutidos_6 = nutricional.values("embutidos").filter(embutidos__icontains="1-2/30 días").count()

        viceras_1 = nutricional.values("viceras").filter(viceras__icontains="0/7 días").count()
        viceras_2 = nutricional.values("viceras").filter(viceras__icontains="1-2/7 días").count()
        viceras_3 = nutricional.values("viceras").filter(viceras__icontains="3-5/7 días").count()
        viceras_4 = nutricional.values("viceras").filter(viceras__icontains="7/7 días").count()
        viceras_5 = nutricional.values("viceras").filter(viceras__icontains="1-2/15 días").count()
        viceras_6 = nutricional.values("viceras").filter(viceras__icontains="1-2/30 días").count()

        grasas_1 = nutricional.values("grasas").filter(grasas__icontains="0/7 días").count()
        grasas_2 = nutricional.values("grasas").filter(grasas__icontains="1-2/7 días").count()
        grasas_3 = nutricional.values("grasas").filter(grasas__icontains="3-5/7 días").count()
        grasas_4 = nutricional.values("grasas").filter(grasas__icontains="7/7 días").count()
        grasas_5 = nutricional.values("grasas").filter(grasas__icontains="1-2/15 días").count()
        grasas_6 = nutricional.values("grasas").filter(grasas__icontains="1-2/30 días").count()

        lacteos_1 = nutricional.values("lacteos").filter(lacteos__icontains="0/7 días").count()
        lacteos_2 = nutricional.values("lacteos").filter(lacteos__icontains="1-2/7 días").count()
        lacteos_3 = nutricional.values("lacteos").filter(lacteos__icontains="3-5/7 días").count()
        lacteos_4 = nutricional.values("lacteos").filter(lacteos__icontains="7/7 días").count()
        lacteos_5 = nutricional.values("lacteos").filter(lacteos__icontains="1-2/15 días").count()
        lacteos_6 = nutricional.values("lacteos").filter(lacteos__icontains="1-2/30 días").count()

        huevos_1 = nutricional.values("huevos").filter(huevos__icontains="0/7 días").count()
        huevos_2 = nutricional.values("huevos").filter(huevos__icontains="1-2/7 días").count()
        huevos_3 = nutricional.values("huevos").filter(huevos__icontains="3-5/7 días").count()
        huevos_4 = nutricional.values("huevos").filter(huevos__icontains="7/7 días").count()
        huevos_5 = nutricional.values("huevos").filter(huevos__icontains="1-2/15 días").count()
        huevos_6 = nutricional.values("huevos").filter(huevos__icontains="1-2/30 días").count()

        leguminosas_1 = nutricional.values("leguminosas").filter(leguminosas__icontains="0/7 días").count()
        leguminosas_2 = nutricional.values("leguminosas").filter(leguminosas__icontains="1-2/7 días").count()
        leguminosas_3 = nutricional.values("leguminosas").filter(leguminosas__icontains="3-5/7 días").count()
        leguminosas_4 = nutricional.values("leguminosas").filter(leguminosas__icontains="7/7 días").count()
        leguminosas_5 = nutricional.values("leguminosas").filter(leguminosas__icontains="1-2/15 días").count()
        leguminosas_6 = nutricional.values("leguminosas").filter(leguminosas__icontains="1-2/30 días").count()

        tuberculos_1 = nutricional.values("tuberculos").filter(tuberculos__icontains="0/7 días").count()
        tuberculos_2 = nutricional.values("tuberculos").filter(tuberculos__icontains="1-2/7 días").count()
        tuberculos_3 = nutricional.values("tuberculos").filter(tuberculos__icontains="3-5/7 días").count()
        tuberculos_4 = nutricional.values("tuberculos").filter(tuberculos__icontains="7/7 días").count()
        tuberculos_5 = nutricional.values("tuberculos").filter(tuberculos__icontains="1-2/15 días").count()
        tuberculos_6 = nutricional.values("tuberculos").filter(tuberculos__icontains="1-2/30 días").count()

        charcuteria_1 = nutricional.values("charcuteria").filter(charcuteria__icontains="0/7 días").count()
        charcuteria_2 = nutricional.values("charcuteria").filter(charcuteria__icontains="1-2/7 días").count()
        charcuteria_3 = nutricional.values("charcuteria").filter(charcuteria__icontains="3-5/7 días").count()
        charcuteria_4 = nutricional.values("charcuteria").filter(charcuteria__icontains="7/7 días").count()
        charcuteria_5 = nutricional.values("charcuteria").filter(charcuteria__icontains="1-2/15 días").count()
        charcuteria_6 = nutricional.values("charcuteria").filter(charcuteria__icontains="1-2/30 días").count()

    #**************  Agrega la data a las celdas  ***************************

    #**************  RESPONSABLE DE HACER EL MERCADO Y COCINAR  *********
        datos = [
                ('Alimentos', '0/7 días', '1-2/7 días', '3-5/7 días', '7/7 días', '1-2/15 días', '1-2/30 días'),
                ('Cereales', cereales_1, cereales_2, cereales_3, cereales_4, cereales_5, cereales_6),
                ('Vegetales', vegetales_1, vegetales_2, vegetales_3, vegetales_4, vegetales_5, vegetales_6),
                ('Frutas', frutas_1, frutas_2, frutas_3, frutas_4, frutas_5, frutas_6),
                ('Carnes', carnes_1, carnes_2, carnes_3, carnes_4, carnes_5, carnes_6),
                ('Pollo', pollo_1, pollo_2, pollo_3, pollo_4, pollo_5, pollo_6),
                ('Pescado', pescado_1, pescado_2, pescado_3, pescado_4, pescado_5, pescado_6),
                ('Embutidos', embutidos_1, embutidos_2, embutidos_3, embutidos_4, embutidos_5, embutidos_6),
                ('Viceras', viceras_1, viceras_2, viceras_3, viceras_4, viceras_5, viceras_6),
                ('Grasas', grasas_1, grasas_2, grasas_3, grasas_4, grasas_5, grasas_6),
                ('Lacteos', lacteos_1, lacteos_2, lacteos_3, lacteos_4, lacteos_5, lacteos_6),
                ('Huevos', huevos_1, huevos_2, huevos_3, huevos_4, huevos_5, huevos_6),
                ('Leguminosas', leguminosas_1, leguminosas_2, leguminosas_3, leguminosas_4, leguminosas_5, leguminosas_6),
                ('Tuberculos', tuberculos_1, tuberculos_2, tuberculos_3, tuberculos_4, tuberculos_5, tuberculos_6),
                ('Charcuteria', charcuteria_1, charcuteria_2, charcuteria_3, charcuteria_4, charcuteria_5, charcuteria_6),
            ]
        for row, cell_value in enumerate(datos, 93):
            ws.cell(row=row+1, column=13)
            for col_num, cell_value in enumerate(cell_value, 6):
                cell = ws.cell(row=row+1, column=col_num)
                cell.value = cell_value
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                cell.alignment = Alignment(horizontal="center", vertical="center")
        
        chart1 = BarChart3D()
        chart1.type = "col"
        chart1.style = 10
        chart1.title = "FRECUENCIA DEL CONSUMO DE ALIMENTOS"
        chart1.y_axis.title = 'Número de Casos'
        chart1.x_axis.title = 'Alimentos'
        chart1.width = 44  # Ancho en pulgadas  
        chart1.height = 12  # Altura en pulgadas

        data = Reference(ws, min_col=7, min_row=94, max_row=109, max_col=12)
        cats = Reference(ws, min_col=6, min_row=95, max_row=109)
        chart1.add_data(data, titles_from_data=True)
        chart1.set_categories(cats)
    
        ws.add_chart(chart1, "C110")


#**********************       TITULO DE LA QUINTA PREGUNTA    *********************
        ws.merge_cells('A135:P136')
        fourth_cell = ws['A135']
        fourth_cell.value = "POCO CONSUMO DE ALIMENTOS"
        fourth_cell.font  = Font(name = 'Tahoma', size = 12, bold = True, color="333399")
        fourth_cell.fill = PatternFill("solid", fgColor="E6E6FA")
        fourth_cell.alignment = Alignment(horizontal="center", vertical="center")      

    #********* asigna el titulo a los graficos  ************************

        ws.merge_cells('G138:J138')
        six_cell = ws['G138']
        six_cell.value = "POCO CONSUMO DE VEGETALES FRUTAS Y VICERAS"
        six_cell.font  = Font(name = 'Tahoma', size = 12, bold = True, color="333399")
        six_cell.fill = PatternFill("solid", fgColor="C6E2FF")
        six_cell.alignment = Alignment(horizontal="center", vertical="center") 

    #**************  Obtener el total de las encuestas  ***************

        poco_vegetales_1 = nutricional.values("poco_vegetales").filter(poco_vegetales__icontains="Rechazo").count()
        poco_vegetales_2 = nutricional.values("poco_vegetales").filter(poco_vegetales__icontains="prepararlo").count()
        poco_vegetales_3 = nutricional.values("poco_vegetales").filter(poco_vegetales__icontains="costosos").count()
        poco_vegetales_4 = nutricional.values("poco_vegetales").filter(poco_vegetales__icontains="acostumbrados").count()

        poco_frutas_1 = nutricional.values("poco_frutas").filter(poco_frutas__icontains="Rechazo").count()
        poco_frutas_2 = nutricional.values("poco_frutas").filter(poco_frutas__icontains="prepararlo").count()
        poco_frutas_3 = nutricional.values("poco_frutas").filter(poco_frutas__icontains="costosos").count()
        poco_frutas_4 = nutricional.values("poco_frutas").filter(poco_frutas__icontains="acostumbrados").count()

        poco_viceras_1 = nutricional.values("poco_viceras").filter(poco_viceras__icontains="Rechazo").count()
        poco_viceras_2 = nutricional.values("poco_viceras").filter(poco_viceras__icontains="prepararlo").count()
        poco_viceras_3 = nutricional.values("poco_viceras").filter(poco_viceras__icontains="costosos").count()
        poco_viceras_4 = nutricional.values("poco_viceras").filter(poco_viceras__icontains="acostumbrados").count()
        

    #**************  Agrega la data a las celdas  ***************************

    #**************  CONSUMO DE VEGETALES, FRUTAS Y VICERAS  *********
        datos = [
                ('Poco Consumo', 'Vegetales', 'Frutas', 'Viceras'),
                ('Por Rechazo', poco_vegetales_1, poco_frutas_1, poco_viceras_1),
                ('No sabe prepararlo', poco_vegetales_2, poco_frutas_2, poco_viceras_2),
                ('Están muy costosos', poco_vegetales_3, poco_frutas_3, poco_viceras_3),
                ('No están acostumbrados', poco_vegetales_4, poco_frutas_4, poco_viceras_4),
            ]
        
        for row, cell_value in enumerate(datos, 139):
            ws.cell(row=row+1, column=13)
            for col_num, cell_value in enumerate(cell_value, 7):
                cell = ws.cell(row=row+1, column=col_num)
                cell.value = cell_value
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                
        chart2 = BarChart3D()
        chart2.type = "col"
        chart2.style = 10
        chart2.title = "POCO CONSUMO DE ALIMENTOS"
        chart2.y_axis.title = 'Número de Casos'
        chart2.x_axis.title = 'Causas de poco consumo'
        chart2.width = 15  # Ancho en pulgadas  
        chart2.height = 8  # Altura en pulgadas

        data = Reference(ws, min_col=8, min_row=140, max_row=144, max_col=10)
        cats = Reference(ws, min_col=7, min_row=141, max_row=144)
        chart2.add_data(data, titles_from_data=True)
        chart2.set_categories(cats)
    
        ws.add_chart(chart2, "G146")


#**********************       TITULO DE LA SEXTA PREGUNTA    *********************
        ws.merge_cells('A162:P163')
        fourth_cell = ws['A162']
        fourth_cell.value = "AYUDA ECONOMICA Y BONOS"
        fourth_cell.font  = Font(name = 'Tahoma', size = 12, bold = True, color="333399")
        fourth_cell.fill = PatternFill("solid", fgColor="E6E6FA")
        fourth_cell.alignment = Alignment(horizontal="center", vertical="center")      

    #********* asigna el titulo a los graficos  ************************

        ws.merge_cells('G165:J165')
        six_cell = ws['G165']
        six_cell.value = "RECIBE ALGUNA AYUDA ECONOMICA Y ALGUN BONO"
        six_cell.font  = Font(name = 'Tahoma', size = 12, bold = True, color="333399")
        six_cell.fill = PatternFill("solid", fgColor="C6E2FF")
        six_cell.alignment = Alignment(horizontal="center", vertical="center") 

    #**************  Obtener el total de las encuestas  ***************

        recibe_bono = nutricional.values("bonos").filter(bonos__icontains="Si").count()
        recibe_clap = nutricional.values("clap").filter(clap__icontains="Si").count()
        recibe_iglesia = nutricional.values("iglesia").filter(iglesia__icontains="Si").count()
        recibe_familiar = nutricional.values("familiar").filter(familiar__icontains="Si").count()
        recibe_pensionado = nutricional.values("pensionado").filter(pensionado__icontains="Si").count()

    #**************  Agrega la data a las celdas  ***************************

    #**************  RECIBE ALGUNA AYUDA ECONOMICA O BONO  *********
        datos = [
                ('Ayuda Economica', 'Recibidos'),
                ('Algún Bono', recibe_bono),
                ('CALP', recibe_clap),
                ('Iglesia', recibe_iglesia),
                ('Familiar', recibe_familiar),
                ('Pensionado', recibe_pensionado),
            ]
        
        for row, cell_value in enumerate(datos, 166):
            ws.cell(row=row+1, column=13)
            for col_num, cell_value in enumerate(cell_value, 8):
                cell = ws.cell(row=row+1, column=col_num)
                cell.value = cell_value
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                cell.alignment = Alignment(horizontal="center", vertical="center")           

        chart2 = BarChart3D()
        chart2.type = "col"
        chart2.style = 10
        chart2.title = "AYUDA ECONOMICA Y ALGUN BONO"
        chart2.y_axis.title = 'Número de Casos'
        chart2.x_axis.title = 'Ayudas y Bonos'
        chart2.width = 15  # Ancho en pulgadas  
        chart2.height = 8  # Altura en pulgadas

        data = Reference(ws, min_col=9, min_row=167, max_row=172)
        cats = Reference(ws, min_col=8, min_row=168, max_row=172)
        chart2.add_data(data, titles_from_data=True)
        chart2.set_categories(cats)
    
        ws.add_chart(chart2, "G173")


#**********************       TITULO DE LA SEXTA PREGUNTA    *********************
        ws.merge_cells('A189:P190')
        fourth_cell = ws['A189']
        fourth_cell.value = "ACTIVIDAD DEPORTIVA"
        fourth_cell.font  = Font(name = 'Tahoma', size = 12, bold = True, color="333399")
        fourth_cell.fill = PatternFill("solid", fgColor="E6E6FA")
        fourth_cell.alignment = Alignment(horizontal="center", vertical="center")      

    #********* asigna el titulo a los graficos  ************************
        ws.merge_cells('B192:E192')
        five_cell = ws['B192']
        five_cell.value = "PRACTICA ALGUN DEPORTE"
        five_cell.font  = Font(name = 'Tahoma', size = 12, bold = True, color="333399")
        five_cell.fill = PatternFill("solid", fgColor="C6E2FF")
        five_cell.alignment = Alignment(horizontal="center", vertical="center")  

        ws.merge_cells('G192:J192')
        six_cell = ws['G192']
        six_cell.value = "TIEMPO DE LA ACTIVIDAD"
        six_cell.font  = Font(name = 'Tahoma', size = 12, bold = True, color="333399")
        six_cell.fill = PatternFill("solid", fgColor="C6E2FF")
        six_cell.alignment = Alignment(horizontal="center", vertical="center") 

        ws.merge_cells('L192:O192')
        seven_cell = ws['L192']
        seven_cell.value = "TIPO DE ACTIVIDAD"
        seven_cell.font  = Font(name = 'Tahoma', size = 12, bold = True, color="333399")
        seven_cell.fill = PatternFill("solid", fgColor="C6E2FF")
        seven_cell.alignment = Alignment(horizontal="center", vertical="center") 

    #**************  Obtener el total de las encuestas  ***************

        practica_si = nutricional.values("practica_deporte").filter(practica_deporte__icontains="Si").count()
        practica_no = nutricional.values("practica_deporte").filter(practica_deporte__icontains="No").count()

        tiempo_1 = nutricional.values("tiempo").filter(tiempo__icontains="30 min de 1-3 / 7").count()
        tiempo_2 = nutricional.values("tiempo").filter(tiempo__icontains="30 min de 4-7 / 7").count()
        tiempo_3 = nutricional.values("tiempo").filter(tiempo__icontains="más de 30 min 1-3 / 7").count()
        tiempo_4 = nutricional.values("tiempo").filter(tiempo__icontains="más de 30 min 4-7 / 7").count()
        tiempo_5 = nutricional.values("tiempo").filter(tiempo__icontains="No precisa datos").count()

        actividad_1 = nutricional.values("actividad").filter(actividad__icontains="Juegos").count()
        actividad_2 = nutricional.values("actividad").filter(actividad__icontains="Práctica").count()
        actividad_3 = nutricional.values("actividad").filter(actividad__icontains="Ambas").count()

    #**************  Agrega la data a las celdas  ***************************

    #**************  PRACTICA ALGUN DEPORTE  *********
        datos = [
                ('Practica', 'Resultados'),
                ('Si', practica_si),
                ('No', practica_no),
            ]
        
        for row, cell_value in enumerate(datos, 193):
            ws.cell(row=row+1, column=13)
            for col_num, cell_value in enumerate(cell_value, 3):
                cell = ws.cell(row=row+1, column=col_num)
                cell.value = cell_value
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                

        chart2 = BarChart3D()
        chart2.type = "col"
        chart2.style = 10
        chart2.title = "PRACTICA DEPORTES"
        chart2.y_axis.title = 'Número de Casos'
        chart2.x_axis.title = 'Practica algún Deporte'
        chart2.width = 15  # Ancho en pulgadas  
        chart2.height = 8  # Altura en pulgadas

        data = Reference(ws, min_col=4, min_row=194, max_row=196)
        cats = Reference(ws, min_col=3, min_row=195, max_row=196)
        chart2.add_data(data, titles_from_data=True)
        chart2.set_categories(cats)
    
        ws.add_chart(chart2, "B200")

    #**************  TIEMPO DE LA ACTIVIDAD  *********
        datos = [
                ('Tiempos', 'Resultados'),
                ('30 min de 1-3 / 7', tiempo_1),
                ('30 min de 4-7 / 7', tiempo_2),
                ('más de 30 min 1-3 / 7', tiempo_3),
                ('más de 30 min 4-7 / 7', tiempo_4),
                ('No precisa datos', tiempo_5),
            ]
        
        for row, cell_value in enumerate(datos, 193):
            ws.cell(row=row+1, column=13)
            for col_num, cell_value in enumerate(cell_value, 8):
                cell = ws.cell(row=row+1, column=col_num)
                cell.value = cell_value
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                

        chart2 = BarChart3D()
        chart2.type = "col"
        chart2.style = 10
        chart2.title = "TIEMPO DE LA ACTIVIDAD"
        chart2.y_axis.title = 'Número de Casos'
        chart2.x_axis.title = 'Frecuencia de la Actividad'
        chart2.width = 15  # Ancho en pulgadas  
        chart2.height = 8  # Altura en pulgadas

        data = Reference(ws, min_col=9, min_row=194, max_row=199)
        cats = Reference(ws, min_col=8, min_row=195, max_row=199)
        chart2.add_data(data, titles_from_data=True)
        chart2.set_categories(cats)
    
        ws.add_chart(chart2, "G200")

    #**************  TIPO DE ACTIVIDAD  *********
        datos = [
                ('Tipo de Actividad', 'Frecuencia'),
                ('Juegos Activos', meriendas_una),
                ('Práctica Deportiva', meriendas_dos),
                ('Ambas Actividades', meriendas_tres),
            ]
        
        for row, cell_value in enumerate(datos, 193):
            ws.cell(row=row+1, column=13)
            for col_num, cell_value in enumerate(cell_value, 13):
                cell = ws.cell(row=row+1, column=col_num)
                cell.value = cell_value
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                cell.alignment = Alignment(horizontal="center", vertical="center")

        chart2 = BarChart3D()
        chart2.type = "col"
        chart2.style = 10
        chart2.title = "TIPO DE ACTIVIDAD"
        chart2.y_axis.title = 'Número de Casos'
        chart2.x_axis.title = 'Tipo de la Actividad'
        chart2.width = 15  # Ancho en pulgadas  
        chart2.height = 8  # Altura en pulgadas

        data = Reference(ws, min_col=14, min_row=194, max_row=197)
        cats = Reference(ws, min_col=13, min_row=195, max_row=197)
        chart2.add_data(data, titles_from_data=True)
        chart2.set_categories(cats)
    
        ws.add_chart(chart2, "L200")


#**********************       TITULO DE LA SEPTIMA PREGUNTA    *********************
        ws.merge_cells('A216:P217')
        fourth_cell = ws['A216']
        fourth_cell.value = "CONSUMO DE AGUA Y SERVICIOS BASICOS"
        fourth_cell.font  = Font(name = 'Tahoma', size = 12, bold = True, color="333399")
        fourth_cell.fill = PatternFill("solid", fgColor="E6E6FA")
        fourth_cell.alignment = Alignment(horizontal="center", vertical="center")      

    #********* asigna el titulo a los graficos  ************************
        ws.merge_cells('D219:G219')
        five_cell = ws['D219']
        five_cell.value = "COMO ES EL CONSUMO DEL AGUA"
        five_cell.font  = Font(name = 'Tahoma', size = 12, bold = True, color="333399")
        five_cell.fill = PatternFill("solid", fgColor="C6E2FF")
        five_cell.alignment = Alignment(horizontal="center", vertical="center")  

        ws.merge_cells('J219:M219')
        six_cell = ws['J219']
        six_cell.value = "CUAL SERVICIO PRESENTAN FALLAS"
        six_cell.font  = Font(name = 'Tahoma', size = 12, bold = True, color="333399")
        six_cell.fill = PatternFill("solid", fgColor="C6E2FF")
        six_cell.alignment = Alignment(horizontal="center", vertical="center") 

    #**************  Obtener el total de las encuestas  ***************

        agua_tuberia = nutricional.values("agua").filter(agua__icontains="Tuberia").count()
        agua_filtrada = nutricional.values("agua").filter(agua__icontains="Filtrada").count()
        agua_botellon = nutricional.values("agua").filter(agua__icontains="Botellon").count()
        agua_hervida = nutricional.values("agua").filter(agua__icontains="Hervida").count()
        agua_tabletas = nutricional.values("agua").filter(agua__icontains="Tabletas").count()

        falla_agua = nutricional.values("falla_servicio").filter(falla_servicio__icontains="agua").count()
        falla_gas = nutricional.values("falla_servicio").filter(falla_servicio__icontains="Gas").count()
        falla_electricidad = nutricional.values("falla_servicio").filter(falla_servicio__icontains="Eléctricidad").count()
        falla_telefonia = nutricional.values("falla_servicio").filter(falla_servicio__icontains="Telefonía").count()
        falla_aseo = nutricional.values("falla_servicio").filter(falla_servicio__icontains="Aseo").count()

    #**************  Agrega la data a las celdas  ***************************

    #**************  COMO ES EL CONSUMO DEL AGUA  *********
        datos = [
                ('Consumo del Agua', 'Consumo'),
                ('Tuberia', agua_tuberia),
                ('Filtrada', agua_filtrada),
                ('Botellon', agua_botellon),
                ('Hervida', agua_hervida),
                ('Tabletas', agua_tabletas),
            ]
        for row, cell_value in enumerate(datos, 220):
            ws.cell(row=row+1, column=13)
            for col_num, cell_value in enumerate(cell_value, 5):
                cell = ws.cell(row=row+1, column=col_num)
                cell.value = cell_value
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                cell.alignment = Alignment(horizontal="center", vertical="center")
        
        chart1 = BarChart3D()
        chart1.type = "col"
        chart1.style = 10
        chart1.title = "CONSUMO DEL AGUA"
        chart1.y_axis.title = 'Número de Casos'
        chart1.x_axis.title = 'Consumo del Agua'
        chart1.width = 15  # Ancho en pulgadas  
        chart1.height = 8  # Altura en pulgadas

        data = Reference(ws, min_col=6, min_row=221, max_row=226)
        cats = Reference(ws, min_col=5, min_row=222, max_row=226)
        chart1.add_data(data, titles_from_data=True)
        chart1.set_categories(cats)
    
        ws.add_chart(chart1, "D227")


    #**************  CUAL SERVICIO PRESENTA FALLAS  *********

        #********* asigna los Datos a las celdas  ************************
        datos = [
                ('Servicio', 'Presentan Falla'),
                ('Agua', falla_agua),
                ('Gas', falla_gas),
                ('Eléctricidad', falla_electricidad),
                ('Telefonía Internet', falla_telefonia),
                ('Aseo Urbano', falla_aseo),
            ]
        
        for row, cell_value in enumerate(datos, 220):
            ws.cell(row=row+1, column=13)
            for col_num, cell_value in enumerate(cell_value, 11):
                cell = ws.cell(row=row+1, column=col_num)
                cell.value = cell_value
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                cell.alignment = Alignment(horizontal="center", vertical="center")

        chart2 = BarChart3D()
        chart2.type = "col"
        chart2.style = 10
        chart2.title = "SERVICIO QUE PRESENTAN FALLAS"
        chart2.y_axis.title = 'Número de Casos'
        chart2.x_axis.title = 'Servicios que Presentan Falla'
        chart2.width = 15  # Ancho en pulgadas  
        chart2.height = 8  # Altura en pulgadas

        data = Reference(ws, min_col=12, min_row=221, max_row=226)
        cats = Reference(ws, min_col=11, min_row=222, max_row=226)
        chart2.add_data(data, titles_from_data=True)
        chart2.set_categories(cats)
    
        ws.add_chart(chart2, "J227")


#******************    COMPRA GAS AGUA ALMACENA AGUA Y DONDE LO ALMACENA    ********************  

    #********* asigna el titulo a los graficos  ************************
        ws.merge_cells('B244:D244')
        five_cell = ws['B244']
        five_cell.value = "COMPRA DE GAS"
        five_cell.font  = Font(name = 'Tahoma', size = 12, bold = True, color="333399")
        five_cell.fill = PatternFill("solid", fgColor="C6E2FF")
        five_cell.alignment = Alignment(horizontal="center", vertical="center")  

        ws.merge_cells('F244:H244')
        six_cell = ws['F244']
        six_cell.value = "COMPRA DE AGUA"
        six_cell.font  = Font(name = 'Tahoma', size = 12, bold = True, color="333399")
        six_cell.fill = PatternFill("solid", fgColor="C6E2FF")
        six_cell.alignment = Alignment(horizontal="center", vertical="center") 

        ws.merge_cells('J244:L244')
        seven_cell = ws['J244']
        seven_cell.value = "ALMACENAMIENTO DE AGUA"
        seven_cell.font  = Font(name = 'Tahoma', size = 12, bold = True, color="333399")
        seven_cell.fill = PatternFill("solid", fgColor="C6E2FF")
        seven_cell.alignment = Alignment(horizontal="center", vertical="center") 

        ws.merge_cells('N244:P244')
        eight_cell = ws['N244']
        eight_cell.value = "DONDE SE ALMACENA"
        eight_cell.font  = Font(name = 'Tahoma', size = 12, bold = True, color="333399")
        eight_cell.fill = PatternFill("solid", fgColor="C6E2FF")
        eight_cell.alignment = Alignment(horizontal="center", vertical="center") 

    #**************  Obtener el total de las encuestas  ***************

        gas_si = nutricional.values("compra_gas").filter(compra_gas__icontains="Si").count()
        gas_no = nutricional.values("compra_gas").filter(compra_gas__icontains="No").count()

        agua_si = nutricional.values("compra_agua").filter(compra_agua__icontains="Si").count()
        agua_no = nutricional.values("compra_agua").filter(compra_agua__icontains="No").count()

        almacena_si = nutricional.values("almacena_agua").filter(almacena_agua__icontains="Si").count()
        almacena_no = nutricional.values("almacena_agua").filter(almacena_agua__icontains="No").count()

        donde_pipote = nutricional.values("donde_almacena").filter(donde_almacena__icontains="Pipote").count()
        donde_tanque = nutricional.values("donde_almacena").filter(donde_almacena__icontains="Tanque").count()
        donde_otros = nutricional.values("donde_almacena").filter(donde_almacena__icontains="Otros").count()

    #**************  Agrega la data a las celdas  ***************************

    #**************  COMPRA DE GAS  *********
        datos = [
                ('Compra Gas', 'Resultados'),
                ('Si', gas_si),
                ('No', gas_no),
            ]
        
        for row, cell_value in enumerate(datos, 245):
            ws.cell(row=row+1, column=13)
            for col_num, cell_value in enumerate(cell_value, 2):
                cell = ws.cell(row=row+1, column=col_num)
                cell.value = cell_value
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                

        chart2 = BarChart3D()
        chart2.type = "col"
        chart2.style = 10
        chart2.title = "COMPRA DE GAS"
        chart2.y_axis.title = 'Número de Casos'
        chart2.x_axis.title = 'Compran Gas'
        chart2.width = 11  # Ancho en pulgadas  
        chart2.height = 8  # Altura en pulgadas

        data = Reference(ws, min_col=3, min_row=246, max_row=248)
        cats = Reference(ws, min_col=2, min_row=247, max_row=248)
        chart2.add_data(data, titles_from_data=True)
        chart2.set_categories(cats)
    
        ws.add_chart(chart2, "B250")

    #**************  COMPRA DE AGUA  *********
        datos = [
                ('Compra Agua', 'Resultados'),
                ('Si', agua_si),
                ('No', agua_no),
            ]
        
        for row, cell_value in enumerate(datos, 245):
            ws.cell(row=row+1, column=13)
            for col_num, cell_value in enumerate(cell_value, 6):
                cell = ws.cell(row=row+1, column=col_num)
                cell.value = cell_value
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                

        chart2 = BarChart3D()
        chart2.type = "col"
        chart2.style = 10
        chart2.title = "COMPRA DE AGUA"
        chart2.y_axis.title = 'Número de Casos'
        chart2.x_axis.title = 'Compran Agua'
        chart2.width = 11  # Ancho en pulgadas  
        chart2.height = 8  # Altura en pulgadas

        data = Reference(ws, min_col=7, min_row=246, max_row=248)
        cats = Reference(ws, min_col=6, min_row=247, max_row=248)
        chart2.add_data(data, titles_from_data=True)
        chart2.set_categories(cats)
    
        ws.add_chart(chart2, "F250")

    #**************  ALMACENAMIENTO DE AGUA  *********
        datos = [
                ('Almacena Agua', 'Resultados'),
                ('Si', almacena_si),
                ('No', almacena_no),
            ]
        
        for row, cell_value in enumerate(datos, 245):
            ws.cell(row=row+1, column=13)
            for col_num, cell_value in enumerate(cell_value, 10):
                cell = ws.cell(row=row+1, column=col_num)
                cell.value = cell_value
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                cell.alignment = Alignment(horizontal="center", vertical="center")

        chart2 = BarChart3D()
        chart2.type = "col"
        chart2.style = 10
        chart2.title = "ALMACENAMIENTO DE AGUA"
        chart2.y_axis.title = 'Número de Casos'
        chart2.x_axis.title = 'Almacenamiento del Agua'
        chart2.width = 11  # Ancho en pulgadas  
        chart2.height = 8  # Altura en pulgadas

        data = Reference(ws, min_col=11, min_row=246, max_row=248)
        cats = Reference(ws, min_col=10, min_row=247, max_row=248)
        chart2.add_data(data, titles_from_data=True)
        chart2.set_categories(cats)
    
        ws.add_chart(chart2, "J250")

            #**************  DONDE SE ALMACENA  *********
        datos = [
                ('Donde Almacenan', 'Resultados'),
                ('Pipote', donde_pipote),
                ('Tanque', donde_tanque),
                ('Otros', donde_otros),
            ]
        
        for row, cell_value in enumerate(datos, 245):
            ws.cell(row=row+1, column=13)
            for col_num, cell_value in enumerate(cell_value, 14):
                cell = ws.cell(row=row+1, column=col_num)
                cell.value = cell_value
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                cell.alignment = Alignment(horizontal="center", vertical="center")

        chart2 = BarChart3D()
        chart2.type = "col"
        chart2.style = 10
        chart2.title = "DONDE SE ALMACENA"
        chart2.y_axis.title = 'Número de Casos'
        chart2.x_axis.title = 'Almacenamiento del Agua'
        chart2.width = 11  # Ancho en pulgadas  
        chart2.height = 8  # Altura en pulgadas

        data = Reference(ws, min_col=15, min_row=246, max_row=249)
        cats = Reference(ws, min_col=14, min_row=247, max_row=249)
        chart2.add_data(data, titles_from_data=True)
        chart2.set_categories(cats)
    
        ws.add_chart(chart2, "N250")


#**********************       TITULO DE LA OPTAVA PREGUNTA    *********************
        ws.merge_cells('A266:P267')
        fourth_cell = ws['A266']
        fourth_cell.value = "EDUCACION NUTRICIONAL"
        fourth_cell.font  = Font(name = 'Tahoma', size = 12, bold = True, color="333399")
        fourth_cell.fill = PatternFill("solid", fgColor="E6E6FA")
        fourth_cell.alignment = Alignment(horizontal="center", vertical="center")      

    #********* asigna el titulo a los graficos  ************************
        ws.merge_cells('B269:E269')
        five_cell = ws['B269']
        five_cell.value = "COONOCE LOS GRUPOS DE ALIMENTOS"
        five_cell.font  = Font(name = 'Tahoma', size = 12, bold = True, color="333399")
        five_cell.fill = PatternFill("solid", fgColor="C6E2FF")
        five_cell.alignment = Alignment(horizontal="center", vertical="center")  

        ws.merge_cells('G269:J269')
        six_cell = ws['G269']
        six_cell.value = "CONOCE DE LA PROTEINA VEGETAL"
        six_cell.font  = Font(name = 'Tahoma', size = 12, bold = True, color="333399")
        six_cell.fill = PatternFill("solid", fgColor="C6E2FF")
        six_cell.alignment = Alignment(horizontal="center", vertical="center") 

        ws.merge_cells('L269:O269')
        seven_cell = ws['L269']
        seven_cell.value = "CONOCE SOBRE LA DESNUTRICION"
        seven_cell.font  = Font(name = 'Tahoma', size = 12, bold = True, color="333399")
        seven_cell.fill = PatternFill("solid", fgColor="C6E2FF")
        seven_cell.alignment = Alignment(horizontal="center", vertical="center") 

    #**************  Obtener el total de las encuestas  ***************

        conoce_grupos_si = nutricional.values("conoce_grupos").filter(conoce_grupos__icontains="Si").count()
        conoce_grupos_no = nutricional.values("conoce_grupos").filter(conoce_grupos__icontains="No").count()
        conoce_grupos_poca = nutricional.values("conoce_grupos").filter(conoce_grupos__icontains="Poca").count()

        conoce_vegetales_si = nutricional.values("conoce_calidad").filter(conoce_calidad__icontains="Si").count()
        conoce_vegetales_no = nutricional.values("conoce_calidad").filter(conoce_calidad__icontains="No").count()
        conoce_vegetales_poca = nutricional.values("conoce_calidad").filter(conoce_calidad__icontains="Poca").count()

        conoce_desnutricion_si = nutricional.values("conoce_desnutricion").filter(conoce_desnutricion__icontains="Si").count()
        conoce_desnutricion_no = nutricional.values("conoce_desnutricion").filter(conoce_desnutricion__icontains="No").count()
        conoce_desnutricion_poca = nutricional.values("conoce_desnutricion").filter(conoce_desnutricion__icontains="Poca").count()


    #**************  Agrega la data a las celdas  ***************************

    #**************  COONOCE LOS GRUPOS DE ALIMENTOS  *********
        datos = [
                ('Conoce los Grupos', 'Resultados'),
                ('Si', conoce_grupos_si),
                ('No', conoce_grupos_no),
                ('Poca Información', conoce_grupos_poca),
            ]
        
        for row, cell_value in enumerate(datos, 270):
            ws.cell(row=row+1, column=13)
            for col_num, cell_value in enumerate(cell_value, 3):
                cell = ws.cell(row=row+1, column=col_num)
                cell.value = cell_value
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                

        chart2 = BarChart3D()
        chart2.type = "col"
        chart2.style = 10
        chart2.title = "COONOCE LOS GRUPOS DE ALIMENTOS"
        chart2.y_axis.title = 'Número de Casos'
        chart2.x_axis.title = 'Conoce los Grupos de Alimentos'
        chart2.width = 15  # Ancho en pulgadas  
        chart2.height = 8  # Altura en pulgadas

        data = Reference(ws, min_col=4, min_row=271, max_row=274)
        cats = Reference(ws, min_col=3, min_row=272, max_row=274)
        chart2.add_data(data, titles_from_data=True)
        chart2.set_categories(cats)
    
        ws.add_chart(chart2, "B276")

    #**************  CONOCE DE LA PROTEINA VEGETAL  *********
        datos = [
                ('Conoce la Proteina', 'Resultados'),
                ('Si', conoce_vegetales_si),
                ('No', conoce_vegetales_no),
                ('Poca Información', conoce_vegetales_poca),
            ]
        
        for row, cell_value in enumerate(datos, 270):
            ws.cell(row=row+1, column=13)
            for col_num, cell_value in enumerate(cell_value, 8):
                cell = ws.cell(row=row+1, column=col_num)
                cell.value = cell_value
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                

        chart2 = BarChart3D()
        chart2.type = "col"
        chart2.style = 10
        chart2.title = "CALIDAD DE LA PROTEINA VEGETAL"
        chart2.y_axis.title = 'Número de Casos'
        chart2.x_axis.title = 'Conoce la Calidad de la Proteina Vegetal'
        chart2.width = 15  # Ancho en pulgadas  
        chart2.height = 8  # Altura en pulgadas

        data = Reference(ws, min_col=9, min_row=271, max_row=274)
        cats = Reference(ws, min_col=8, min_row=272, max_row=274)
        chart2.add_data(data, titles_from_data=True)
        chart2.set_categories(cats)
    
        ws.add_chart(chart2, "G276")

    #**************  CONOCE SOBRE LA DESNUTRICION  *********
        datos = [
                ('Conoce la Desnutrición', 'Resultados'),
                ('Si', conoce_desnutricion_si),
                ('No', conoce_desnutricion_no),
                ('Poca Información', conoce_desnutricion_poca),
            ]
        
        for row, cell_value in enumerate(datos, 270):
            ws.cell(row=row+1, column=13)
            for col_num, cell_value in enumerate(cell_value, 13):
                cell = ws.cell(row=row+1, column=col_num)
                cell.value = cell_value
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                cell.alignment = Alignment(horizontal="center", vertical="center")

        chart2 = BarChart3D()
        chart2.type = "col"
        chart2.style = 10
        chart2.title = "CONOCE SOBRE LA DESNUTRICION"
        chart2.y_axis.title = 'Número de Casos'
        chart2.x_axis.title = 'Conoce sobre la desnutrición y su consecuencia'
        chart2.width = 15  # Ancho en pulgadas  
        chart2.height = 8  # Altura en pulgadas

        data = Reference(ws, min_col=14, min_row=271, max_row=274)
        cats = Reference(ws, min_col=13, min_row=272, max_row=274)
        chart2.add_data(data, titles_from_data=True)
        chart2.set_categories(cats)
    
        ws.add_chart(chart2, "L276")


#**********************       TITULO DE LA NOVENA PREGUNTA    *********************
        ws.merge_cells('A292:P293')
        fourth_cell = ws['A292']
        fourth_cell.value = "EDUCACION PARA EMBARAZADAS Y LACTANTES"
        fourth_cell.font  = Font(name = 'Tahoma', size = 12, bold = True, color="333399")
        fourth_cell.fill = PatternFill("solid", fgColor="E6E6FA")
        fourth_cell.alignment = Alignment(horizontal="center", vertical="center")      

    #********* asigna el titulo a los graficos  ************************
        ws.merge_cells('E295:L295')
        five_cell = ws['E295']
        five_cell.value = "INFORMACION DE ORIENTACION PARA LAS EMBARAZADAS Y LACTANTES"
        five_cell.font  = Font(name = 'Tahoma', size = 12, bold = True, color="333399")
        five_cell.fill = PatternFill("solid", fgColor="C6E2FF")
        five_cell.alignment = Alignment(horizontal="center", vertical="center")  

    #**************  Obtener el total de las encuestas  ***************

        beneficio_si = nutricional.values("conoce_beneficio").filter(conoce_beneficio__icontains="Si").count()
        beneficio_no = nutricional.values("conoce_beneficio").filter(conoce_beneficio__icontains="No").count()
        beneficio_poca = nutricional.values("conoce_beneficio").filter(conoce_beneficio__icontains="Poca").count()
        beneficio_aplica = nutricional.values("conoce_beneficio").filter(conoce_beneficio__icontains="Aplica").count()

        amamantar_si = nutricional.values("desea_amamantar").filter(conoce_beneficio__icontains="Si").count()
        amamantar_no = nutricional.values("desea_amamantar").filter(conoce_beneficio__icontains="No").count()
        amamantar_poca = nutricional.values("desea_amamantar").filter(conoce_beneficio__icontains="Poca").count()
        amamantar_aplica = nutricional.values("desea_amamantar").filter(conoce_beneficio__icontains="Aplica").count()

        dificultad_si = nutricional.values("dificultad_amamantar").filter(dificultad_amamantar__icontains="Si").count()
        dificultad_no = nutricional.values("dificultad_amamantar").filter(dificultad_amamantar__icontains="No").count()
        dificultad_poca = nutricional.values("dificultad_amamantar").filter(dificultad_amamantar__icontains="Poca").count()
        dificultad_aplica = nutricional.values("dificultad_amamantar").filter(dificultad_amamantar__icontains="Aplica").count()

        orientacion_si = nutricional.values("desea_orientacion").filter(desea_orientacion__icontains="Si").count()
        orientacion_no = nutricional.values("desea_orientacion").filter(desea_orientacion__icontains="No").count()
        orientacion_poca = nutricional.values("desea_orientacion").filter(desea_orientacion__icontains="Poca").count()
        orientacion_aplica = nutricional.values("desea_orientacion").filter(desea_orientacion__icontains="Aplica").count()

        conocimiento_si = nutricional.values("desea_conocimiento").filter(desea_conocimiento__icontains="Si").count()
        conocimiento_no = nutricional.values("desea_conocimiento").filter(desea_conocimiento__icontains="No").count()
        conocimiento_poca = nutricional.values("desea_conocimiento").filter(desea_conocimiento__icontains="Poca").count()
        conocimiento_aplica = nutricional.values("desea_conocimiento").filter(desea_conocimiento__icontains="Aplica").count()

    #**************  Agrega la data a las celdas  ***************************

    #**************  EDUCACION PARA EMBARAZADAS Y LACTANTES  *********
        datos = [
                ('Educación', 'Si', 'No', 'Poca Información', 'No Aplica'),
                ('Beneficio Lactancia', beneficio_si, beneficio_no, beneficio_poca, beneficio_aplica),
                ('Desea Amamantar', amamantar_si, amamantar_no, amamantar_poca, amamantar_aplica),
                ('Dificultad de Amamantar', dificultad_si, dificultad_no, dificultad_poca, dificultad_aplica),
                ('Desea Orientación', orientacion_si, orientacion_no, orientacion_poca, orientacion_aplica),
                ('Mas Conocimiento', conocimiento_si, conocimiento_no, conocimiento_poca, conocimiento_aplica),
            ]
        for row, cell_value in enumerate(datos, 296):
            ws.cell(row=row+1, column=13)
            for col_num, cell_value in enumerate(cell_value, 7):
                cell = ws.cell(row=row+1, column=col_num)
                cell.value = cell_value
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                cell.alignment = Alignment(horizontal="center", vertical="center")
        
        chart1 = BarChart3D()
        chart1.type = "col"
        chart1.style = 10
        chart1.title = "EDUCACION PARA EMBARAZADAS Y LACTANTES"
        chart1.y_axis.title = 'Número de Casos'
        chart1.x_axis.title = 'EDUCACION PARA EMBARAZADAS Y LACTANTES'
        chart1.width = 37  # Ancho en pulgadas  
        chart1.height = 12  # Altura en pulgadas

        data = Reference(ws, min_col=8, min_row=297, max_row=302, max_col=11)
        cats = Reference(ws, min_col=7, min_row=298, max_row=302)
        chart1.add_data(data, titles_from_data=True)
        chart1.set_categories(cats)
    
        ws.add_chart(chart1, "D304")


#*-*-*-*-*-*-*-*-*-**-*    final de la seccion de la hoja   *-*-*-*-*-*-*-*-*

    #*********  Establecer el nombre del Archivo *******
        nombre_archvo = "Reporte_Nutricional_Proyecto.xlsx"
    
    #*********  Definir el tipo de respuesta que se va a dar ***********
        response = HttpResponse(content_type = "application/ms-excel")
        contenido = "attachment; filename = {0}".format(nombre_archvo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response
    


class estadistica_nutricional_jornada(TemplateView):
     def get(self, request, *args, **kwargs):
        proyecto = request.GET.get('proyecto')
        jornada = request.GET.get('jornada')
        nutricional = Nutricional.objects.filter(proyecto=proyecto,jornada=jornada)
        nombre_proyecto = get_object_or_404(Proyecto, id=proyecto)
        nombre_jornada = get_object_or_404(Jornada, id=jornada)
        total_encuestados = nutricional.count()
        wb = Workbook()
        ws = wb.active
        ws.title = "ESTADISTICA NUTRICIONAL"
        thin = Side(border_style="thin", color="000000")

    #*********  Crear encabezado en la hoja  *************

        for column in ('A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U'): 
            ws.column_dimensions[column].width = 20
    

        fecha = datetime.now()
        fecha_fin = fecha.strftime('%d-%m-%Y - hora: %H:%m')
        ws.merge_cells('A1:C1')
        first_cell = ws['A1']
        first_cell.value = "Fecha: " + fecha_fin
        
        ws.merge_cells('A2:E2')
        second_cell = ws['A2']
        second_cell.value = "Bokitas"
        second_cell.font  = Font(name = 'Tw Cen MT Condensed Extra Bold', size = 52, bold = True, color="6F42C1")

        ws.merge_cells('A3:E3')
        third_cell = ws['A3']
        third_cell.value = "Bokitas Fundation    www.bokitas.org"
        third_cell.font  = Font(name = 'Tw Cen MT Condensed Extra Bold', size = 12, bold = True, color="6F42C1")
        
#***********************  DATOS ESTADISTICA NUTRICIONAL  ********************************

        #*********  Tutulo Principal  *************
        ws.merge_cells('A4:P5')
        fourth_cell = ws['A4']
        fourth_cell.value = "ESTADISTICA NUTRICIONAL DEL PROYECTO: " + str(nombre_proyecto) + " DE LA JORNADA: " + str(nombre_jornada.jornada)
        fourth_cell.font  = Font(name = 'Tahoma', size = 14, bold = True, color="333399")
        fourth_cell.fill = PatternFill("solid", fgColor="E2D9F3")
        fourth_cell.alignment = Alignment(horizontal="center", vertical="center") 

        ws.merge_cells('A6:P6')
        fourth_cell = ws['A6']
        fourth_cell.value = "TOTAL DE ENCUESTADOS: " + str(total_encuestados)
        fourth_cell.font  = Font(name = 'Tahoma', size = 14, bold = True, color="333399")
        fourth_cell.fill = PatternFill("solid", fgColor="E2D9F3")
        fourth_cell.alignment = Alignment(horizontal="center", vertical="center") 
    
    
#**********************       TITULO DE LA PRIMERA PREGUNTA    *********************
        ws.merge_cells('A8:P9')
        fourth_cell = ws['A8']
        fourth_cell.value = "RESPONSABLE DE HACER EL MERCADO, COCINAR Y FRECUENCIA DE COMPRA DE ALIMENTOS"
        fourth_cell.font  = Font(name = 'Tahoma', size = 12, bold = True, color="333399")
        fourth_cell.fill = PatternFill("solid", fgColor="E6E6FA")
        fourth_cell.alignment = Alignment(horizontal="center", vertical="center")      

    #********* asigna el titulo a los graficos  ************************
        ws.merge_cells('D11:G11')
        five_cell = ws['D11']
        five_cell.value = "RESPONSABLE DE HACER EL MERCADO Y COCINAR"
        five_cell.font  = Font(name = 'Tahoma', size = 12, bold = True, color="333399")
        five_cell.fill = PatternFill("solid", fgColor="C6E2FF")
        five_cell.alignment = Alignment(horizontal="center", vertical="center")  

        ws.merge_cells('J11:M11')
        six_cell = ws['J11']
        six_cell.value = "FRECUENCIA DE LA COMPRA DE ALIMENTOS"
        six_cell.font  = Font(name = 'Tahoma', size = 12, bold = True, color="333399")
        six_cell.fill = PatternFill("solid", fgColor="C6E2FF")
        six_cell.alignment = Alignment(horizontal="center", vertical="center") 

    #**************  Obtener el total de las encuestas  ***************

        padre_mercado = nutricional.values("mercado_lorealiza").filter(mercado_lorealiza__icontains="Padre").count()
        madre_mercado = nutricional.values("mercado_lorealiza").filter(mercado_lorealiza__icontains="Madre").count()
        abuelo_mercado = nutricional.values("mercado_lorealiza").filter(mercado_lorealiza__icontains="Abuelo(a)").count()
        tio_mercado = nutricional.values("mercado_lorealiza").filter(mercado_lorealiza__icontains="Tio(a)").count()
        otros_mercado = nutricional.values("mercado_lorealiza").filter(mercado_lorealiza__icontains="Otros").count()

        padre_cocina = nutricional.values("cocina_lorealiza").filter(cocina_lorealiza__icontains="Padre").count()
        madre_cocina = nutricional.values("cocina_lorealiza").filter(cocina_lorealiza__icontains="Madre").count()
        abuelo_cocina = nutricional.values("cocina_lorealiza").filter(cocina_lorealiza__icontains="Abuelo(a)").count()
        tio_cocina = nutricional.values("cocina_lorealiza").filter(cocina_lorealiza__icontains="Tio(a)").count()
        otros_cocina = nutricional.values("cocina_lorealiza").filter(cocina_lorealiza__icontains="Otros").count()
        
        fecuencia_diario = nutricional.values("frecuencia").filter(frecuencia__icontains="diario").count()
        fecuencia_2_3 = nutricional.values("frecuencia").filter(frecuencia__icontains="2-3 días").count()
        fecuencia_semanal = nutricional.values("frecuencia").filter(frecuencia__icontains="Semanal").count()
        fecuencia_quincenal = nutricional.values("frecuencia").filter(frecuencia__icontains="Quincenal").count()


    #**************  Agrega la data a las celdas  ***************************

    #**************  RESPONSABLE DE HACER EL MERCADO Y COCINAR  *********
        datos = [
                ('Integrantes', 'Quíen Hace Mercado', 'Quíen Cocina'),
                ('Madre', madre_mercado, madre_cocina),
                ('Padre', padre_mercado, padre_cocina),
                ('Abuelos', abuelo_mercado, abuelo_cocina),
                ('Tios', tio_mercado, tio_cocina),
                ('Otros', otros_mercado, otros_cocina),
            ]
        for row, cell_value in enumerate(datos, 12):
            ws.cell(row=row+1, column=13)
            for col_num, cell_value in enumerate(cell_value, 4):
                cell = ws.cell(row=row+1, column=col_num)
                cell.value = cell_value
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                cell.alignment = Alignment(horizontal="center", vertical="center")
        
        chart1 = BarChart3D()
        chart1.type = "col"
        chart1.style = 10
        chart1.title = "RESPONSABLE DE HACER EL MERCADO Y COCINAR"
        chart1.y_axis.title = 'Número de Casos'
        chart1.x_axis.title = 'Integrantes'
        chart1.width = 15  # Ancho en pulgadas  
        chart1.height = 8  # Altura en pulgadas

        data = Reference(ws, min_col=5, min_row=13, max_row=18, max_col=6)
        cats = Reference(ws, min_col=4, min_row=14, max_row=18)
        chart1.add_data(data, titles_from_data=True)
        chart1.set_categories(cats)
    
        ws.add_chart(chart1, "D19")


    #**************  FRECUENCIA DE LA COMPRA DE ALIMENTOS  *********

        #********* asigna los Datos a las celdas  ************************
        datos = [
                ('Compras', 'Frecuencia de la Compra'),
                ('Diaria', fecuencia_diario),
                ('2-3 días', fecuencia_2_3),
                ('Semanal', fecuencia_semanal),
                ('Quincenal', fecuencia_quincenal),
            ]
        
        for row, cell_value in enumerate(datos, 12):
            ws.cell(row=row+1, column=13)
            for col_num, cell_value in enumerate(cell_value, 11):
                cell = ws.cell(row=row+1, column=col_num)
                cell.value = cell_value
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                cell.alignment = Alignment(horizontal="center", vertical="center")

        chart2 = BarChart3D()
        chart2.type = "col"
        chart2.style = 10
        chart2.title = "FRECUENCIA DE LA COMPRA DE ALIMENTOS"
        chart2.y_axis.title = 'Número de Casos'
        chart2.x_axis.title = 'Frecuencia de la Compra'
        chart2.width = 15  # Ancho en pulgadas  
        chart2.height = 8  # Altura en pulgadas

        data = Reference(ws, min_col=12, min_row=13, max_row=17)
        cats = Reference(ws, min_col=11, min_row=14, max_row=17)
        chart2.add_data(data, titles_from_data=True)
        chart2.set_categories(cats)
    
        ws.add_chart(chart2, "J19")


#**********************       TITULO DE LA SEGUNDA PREGUNTA    *********************
        ws.merge_cells('A35:P36')
        fourth_cell = ws['A35']
        fourth_cell.value = "APETITO, COMIDAS Y MERIENDAS"
        fourth_cell.font  = Font(name = 'Tahoma', size = 12, bold = True, color="333399")
        fourth_cell.fill = PatternFill("solid", fgColor="E6E6FA")
        fourth_cell.alignment = Alignment(horizontal="center", vertical="center")      

    #********* asigna el titulo a los graficos  ************************
        ws.merge_cells('B38:E38')
        five_cell = ws['B38']
        five_cell.value = "COMO ES EL APETITO"
        five_cell.font  = Font(name = 'Tahoma', size = 12, bold = True, color="333399")
        five_cell.fill = PatternFill("solid", fgColor="C6E2FF")
        five_cell.alignment = Alignment(horizontal="center", vertical="center")  

        ws.merge_cells('G38:J38')
        six_cell = ws['G38']
        six_cell.value = "COMIDAS AL DIA"
        six_cell.font  = Font(name = 'Tahoma', size = 12, bold = True, color="333399")
        six_cell.fill = PatternFill("solid", fgColor="C6E2FF")
        six_cell.alignment = Alignment(horizontal="center", vertical="center") 

        ws.merge_cells('L38:O38')
        seven_cell = ws['L38']
        seven_cell.value = "CUANTAS MERIENDAS"
        seven_cell.font  = Font(name = 'Tahoma', size = 12, bold = True, color="333399")
        seven_cell.fill = PatternFill("solid", fgColor="C6E2FF")
        seven_cell.alignment = Alignment(horizontal="center", vertical="center") 

    #**************  Obtener el total de las encuestas  ***************

        apetito_bueno = nutricional.values("apetito").filter(apetito__icontains="Bueno").count()
        apetito_regular = nutricional.values("apetito").filter(apetito__icontains="Regular").count()
        apetito_malo = nutricional.values("apetito").filter(apetito__icontains="Malo").count()
        apetito_elevado = nutricional.values("apetito").filter(apetito__icontains="Elevado").count()

        comidas_una = nutricional.values("cuantas_comidas").filter(cuantas_comidas__icontains="1 Comida").count()
        comidas_dos = nutricional.values("cuantas_comidas").filter(cuantas_comidas__icontains="2 Comidas").count()
        comidas_tres = nutricional.values("cuantas_comidas").filter(cuantas_comidas__icontains="3 Comidas").count()

        meriendas_una = nutricional.values("meriendas").filter(meriendas__icontains="1 Merienda").count()
        meriendas_dos = nutricional.values("meriendas").filter(meriendas__icontains="2 Meriendas").count()
        meriendas_tres = nutricional.values("meriendas").filter(meriendas__icontains="3 Meriendas").count()
        meriendas_ninguna = nutricional.values("meriendas").filter(meriendas__icontains="Ninguna").count()

    #**************  Agrega la data a las celdas  ***************************

    #**************  COMO ES EL APETITO  *********
        datos = [
                ('Apetito', 'Resultados'),
                ('Bueno', apetito_bueno),
                ('Regular', apetito_regular),
                ('Malo', apetito_malo),
                ('Elevado', apetito_elevado),
            ]
        
        for row, cell_value in enumerate(datos, 39):
            ws.cell(row=row+1, column=13)
            for col_num, cell_value in enumerate(cell_value, 3):
                cell = ws.cell(row=row+1, column=col_num)
                cell.value = cell_value
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                

        chart2 = BarChart3D()
        chart2.type = "col"
        chart2.style = 10
        chart2.title = "FRECUENCIA DEL APETITO"
        chart2.y_axis.title = 'Número de Casos'
        chart2.x_axis.title = 'Frecuencia del Apetito'
        chart2.width = 15  # Ancho en pulgadas  
        chart2.height = 8  # Altura en pulgadas

        data = Reference(ws, min_col=4, min_row=40, max_row=44)
        cats = Reference(ws, min_col=3, min_row=41, max_row=44)
        chart2.add_data(data, titles_from_data=True)
        chart2.set_categories(cats)
    
        ws.add_chart(chart2, "B46")

    #**************  COMIDAS AL DIA  *********
        datos = [
                ('Comidas', 'Resultados'),
                ('1 Comida', comidas_una),
                ('2 Comidas', comidas_dos),
                ('3 Comidas', comidas_tres),
            ]
        
        for row, cell_value in enumerate(datos, 39):
            ws.cell(row=row+1, column=13)
            for col_num, cell_value in enumerate(cell_value, 8):
                cell = ws.cell(row=row+1, column=col_num)
                cell.value = cell_value
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                

        chart2 = BarChart3D()
        chart2.type = "col"
        chart2.style = 10
        chart2.title = "COMIDAS AL DIA"
        chart2.y_axis.title = 'Número de Casos'
        chart2.x_axis.title = 'Frecuencia de Comidas al Día'
        chart2.width = 15  # Ancho en pulgadas  
        chart2.height = 8  # Altura en pulgadas

        data = Reference(ws, min_col=9, min_row=40, max_row=43)
        cats = Reference(ws, min_col=8, min_row=41, max_row=43)
        chart2.add_data(data, titles_from_data=True)
        chart2.set_categories(cats)
    
        ws.add_chart(chart2, "G46")


    #**************  CUANTAS MERIENDAS  *********
        datos = [
                ('Meriendas', 'Frecuencia de la Merienda'),
                ('1 Merienda', meriendas_una),
                ('2 Meriendas', meriendas_dos),
                ('3 Meriendas', meriendas_tres),
                ('Ninguna', meriendas_ninguna),
            ]
        
        for row, cell_value in enumerate(datos, 39):
            ws.cell(row=row+1, column=13)
            for col_num, cell_value in enumerate(cell_value, 13):
                cell = ws.cell(row=row+1, column=col_num)
                cell.value = cell_value
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                cell.alignment = Alignment(horizontal="center", vertical="center")

        chart2 = BarChart3D()
        chart2.type = "col"
        chart2.style = 10
        chart2.title = "FRECUENCIA DE LAS MERIENDAS"
        chart2.y_axis.title = 'Número de Casos'
        chart2.x_axis.title = 'Frecuencia de la Merienda'
        chart2.width = 15  # Ancho en pulgadas  
        chart2.height = 8  # Altura en pulgadas

        data = Reference(ws, min_col=14, min_row=40, max_row=44)
        cats = Reference(ws, min_col=13, min_row=41, max_row=44)
        chart2.add_data(data, titles_from_data=True)
        chart2.set_categories(cats)
    
        ws.add_chart(chart2, "L46")


#**********************       TITULO DE LA TERCESA PREGUNTA    *********************
        ws.merge_cells('A62:P63')
        fourth_cell = ws['A62']
        fourth_cell.value = "GRUPOS DE ALIMENTOS"
        fourth_cell.font  = Font(name = 'Tahoma', size = 12, bold = True, color="333399")
        fourth_cell.fill = PatternFill("solid", fgColor="E6E6FA")
        fourth_cell.alignment = Alignment(horizontal="center", vertical="center")      

    #********* asigna el titulo a los graficos  ************************
        ws.merge_cells('D65:G65')
        five_cell = ws['D65']
        five_cell.value = "CUANTOS GRUPOS ESTAN PRESENTES EN LA ALIMENTACION"
        five_cell.font  = Font(name = 'Tahoma', size = 12, bold = True, color="333399")
        five_cell.fill = PatternFill("solid", fgColor="C6E2FF")
        five_cell.alignment = Alignment(horizontal="center", vertical="center")  

        ws.merge_cells('J65:M65')
        six_cell = ws['J65']
        six_cell.value = "CUALES SON LOS GRUPOS PRESENTES"
        six_cell.font  = Font(name = 'Tahoma', size = 12, bold = True, color="333399")
        six_cell.fill = PatternFill("solid", fgColor="C6E2FF")
        six_cell.alignment = Alignment(horizontal="center", vertical="center") 

    #**************  Obtener el total de las encuestas  ***************

        grupo_1_2 = nutricional.values("cuantos_grupos").filter(cuantos_grupos__icontains="1 - 2").count()
        grupo_2_3 = nutricional.values("cuantos_grupos").filter(cuantos_grupos__icontains="2 - 3").count()
        grupo_3_4 = nutricional.values("cuantos_grupos").filter(cuantos_grupos__icontains="3 - 4").count()
        grupo_4_5 = nutricional.values("cuantos_grupos").filter(cuantos_grupos__icontains="4 - 5").count()
        grupo_5_6 = nutricional.values("cuantos_grupos").filter(cuantos_grupos__icontains="5 - 6").count()

        grupo_1 = nutricional.values("cuantos_grupos").filter(tipo_grupos__icontains="Cereales y Leguminosas").count()
        grupo_2 = nutricional.values("cuantos_grupos").filter(tipo_grupos__icontains="Frutas y Verduras").count()
        grupo_3 = nutricional.values("cuantos_grupos").filter(tipo_grupos__icontains="Leche Yogures y Quesos").count()
        grupo_4 = nutricional.values("cuantos_grupos").filter(tipo_grupos__icontains="Grasas y Aceites").count()
        grupo_5 = nutricional.values("cuantos_grupos").filter(tipo_grupos__icontains="Dulces").count()
        grupo_6 = nutricional.values("cuantos_grupos").filter(tipo_grupos__icontains="Carnes y Huevos").count()

    #**************  Agrega la data a las celdas  ***************************

    #**************  CUANTOS GRUPOS ESTAN PRESENTES EN LA ALIMENTACION  *********
        datos = [
                ('Grupos', 'Presentes'),
                ('1 - 2', grupo_1_2),
                ('2 - 3', grupo_2_3),
                ('3 - 4', grupo_3_4),
                ('4 - 5', grupo_4_5),
                ('5 - 6', grupo_5_6),
            ]
        for row, cell_value in enumerate(datos, 66):
            ws.cell(row=row+1, column=13)
            for col_num, cell_value in enumerate(cell_value, 5):
                cell = ws.cell(row=row+1, column=col_num)
                cell.value = cell_value
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                cell.alignment = Alignment(horizontal="center", vertical="center")
        
        chart1 = BarChart3D()
        chart1.type = "col"
        chart1.style = 10
        chart1.title = "GRUPOS PRESENTES EN LA ALIMENTACION"
        chart1.y_axis.title = 'Número de Casos'
        chart1.x_axis.title = 'Grupos de Alimentos'
        chart1.width = 15  # Ancho en pulgadas  
        chart1.height = 8  # Altura en pulgadas

        data = Reference(ws, min_col=6, min_row=67, max_row=72)
        cats = Reference(ws, min_col=5, min_row=68, max_row=72)
        chart1.add_data(data, titles_from_data=True)
        chart1.set_categories(cats)
    
        ws.add_chart(chart1, "D73")


    #**************  FRECUENCIA DEL CONSUMO DE LOS GRUPOS  *********

        #********* asigna los Datos a las celdas  ************************
        datos = [
                ('Grupos', 'Grupos Presentes'),
                ('Grupo 1', grupo_1),
                ('Grupo 2', grupo_2),
                ('Grupo 3', grupo_3),
                ('Grupo 4', grupo_4),
                ('Grupo 5', grupo_5),
                ('Grupo 6', grupo_6),
            ]
        
        for row, cell_value in enumerate(datos, 65):
            ws.cell(row=row+1, column=13)
            for col_num, cell_value in enumerate(cell_value, 11):
                cell = ws.cell(row=row+1, column=col_num)
                cell.value = cell_value
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                cell.alignment = Alignment(horizontal="center", vertical="center")

        chart2 = BarChart3D()
        chart2.type = "col"
        chart2.style = 10
        chart2.title = "CUALES GRUPOS SE PRESENTAN EN LA ALIMENTACION"
        chart2.y_axis.title = 'Número de Casos'
        chart2.x_axis.title = 'Frecuencia de la Grupos'
        chart2.width = 15  # Ancho en pulgadas  
        chart2.height = 8  # Altura en pulgadas

        data = Reference(ws, min_col=12, min_row=66, max_row=72)
        cats = Reference(ws, min_col=11, min_row=67, max_row=72)
        chart2.add_data(data, titles_from_data=True)
        chart2.set_categories(cats)
    
        ws.add_chart(chart2, "J73")


#**********************       TITULO DE LA CUARTA PREGUNTA    *********************
        ws.merge_cells('A89:P90')
        fourth_cell = ws['A89']
        fourth_cell.value = "FRECUENCIA DEL CONSUMO DE LOS ALIMENTOS"
        fourth_cell.font  = Font(name = 'Tahoma', size = 12, bold = True, color="333399")
        fourth_cell.fill = PatternFill("solid", fgColor="E6E6FA")
        fourth_cell.alignment = Alignment(horizontal="center", vertical="center")      

    #********* asigna el titulo a los graficos  ************************
        ws.merge_cells('E92:L92')
        five_cell = ws['E92']
        five_cell.value = "FRECUENCIA DEL CONSUMO"
        five_cell.font  = Font(name = 'Tahoma', size = 12, bold = True, color="333399")
        five_cell.fill = PatternFill("solid", fgColor="C6E2FF")
        five_cell.alignment = Alignment(horizontal="center", vertical="center")  

    #**************  Obtener el total de las encuestas  ***************

        cereales_1 = nutricional.values("cereales").filter(cereales__icontains="0/7 días").count()
        cereales_2 = nutricional.values("cereales").filter(cereales__icontains="1-2/7 días").count()
        cereales_3 = nutricional.values("cereales").filter(cereales__icontains="3-5/7 días").count()
        cereales_4 = nutricional.values("cereales").filter(cereales__icontains="7/7 días").count()
        cereales_5 = nutricional.values("cereales").filter(cereales__icontains="1-2/15 días").count()
        cereales_6 = nutricional.values("cereales").filter(cereales__icontains="1-2/30 días").count()

        vegetales_1 = nutricional.values("vegetales").filter(vegetales__icontains="0/7 días").count()
        vegetales_2 = nutricional.values("vegetales").filter(vegetales__icontains="1-2/7 días").count()
        vegetales_3 = nutricional.values("vegetales").filter(vegetales__icontains="3-5/7 días").count()
        vegetales_4 = nutricional.values("vegetales").filter(vegetales__icontains="7/7 días").count()
        vegetales_5 = nutricional.values("vegetales").filter(vegetales__icontains="1-2/15 días").count()
        vegetales_6 = nutricional.values("vegetales").filter(vegetales__icontains="1-2/30 días").count()

        frutas_1 = nutricional.values("frutas").filter(frutas__icontains="0/7 días").count()
        frutas_2 = nutricional.values("frutas").filter(frutas__icontains="1-2/7 días").count()
        frutas_3 = nutricional.values("frutas").filter(frutas__icontains="3-5/7 días").count()
        frutas_4 = nutricional.values("frutas").filter(frutas__icontains="7/7 días").count()
        frutas_5 = nutricional.values("frutas").filter(frutas__icontains="1-2/15 días").count()
        frutas_6 = nutricional.values("frutas").filter(frutas__icontains="1-2/30 días").count()

        carnes_1 = nutricional.values("carnes").filter(carnes__icontains="0/7 días").count()
        carnes_2 = nutricional.values("carnes").filter(carnes__icontains="1-2/7 días").count()
        carnes_3 = nutricional.values("carnes").filter(carnes__icontains="3-5/7 días").count()
        carnes_4 = nutricional.values("carnes").filter(carnes__icontains="7/7 días").count()
        carnes_5 = nutricional.values("carnes").filter(carnes__icontains="1-2/15 días").count()
        carnes_6 = nutricional.values("carnes").filter(carnes__icontains="1-2/30 días").count()

        pollo_1 = nutricional.values("pollo").filter(pollo__icontains="0/7 días").count()
        pollo_2 = nutricional.values("pollo").filter(pollo__icontains="1-2/7 días").count()
        pollo_3 = nutricional.values("pollo").filter(pollo__icontains="3-5/7 días").count()
        pollo_4 = nutricional.values("pollo").filter(pollo__icontains="7/7 días").count()
        pollo_5 = nutricional.values("pollo").filter(pollo__icontains="1-2/15 días").count()
        pollo_6 = nutricional.values("pollo").filter(pollo__icontains="1-2/30 días").count()

        pescado_1 = nutricional.values("pescado").filter(pescado__icontains="0/7 días").count()
        pescado_2 = nutricional.values("pescado").filter(pescado__icontains="1-2/7 días").count()
        pescado_3 = nutricional.values("pescado").filter(pescado__icontains="3-5/7 días").count()
        pescado_4 = nutricional.values("pescado").filter(pescado__icontains="7/7 días").count()
        pescado_5 = nutricional.values("pescado").filter(pescado__icontains="1-2/15 días").count()
        pescado_6 = nutricional.values("pescado").filter(pescado__icontains="1-2/30 días").count()

        embutidos_1 = nutricional.values("embutidos").filter(embutidos__icontains="0/7 días").count()
        embutidos_2 = nutricional.values("embutidos").filter(embutidos__icontains="1-2/7 días").count()
        embutidos_3 = nutricional.values("embutidos").filter(embutidos__icontains="3-5/7 días").count()
        embutidos_4 = nutricional.values("embutidos").filter(embutidos__icontains="7/7 días").count()
        embutidos_5 = nutricional.values("embutidos").filter(embutidos__icontains="1-2/15 días").count()
        embutidos_6 = nutricional.values("embutidos").filter(embutidos__icontains="1-2/30 días").count()

        viceras_1 = nutricional.values("viceras").filter(viceras__icontains="0/7 días").count()
        viceras_2 = nutricional.values("viceras").filter(viceras__icontains="1-2/7 días").count()
        viceras_3 = nutricional.values("viceras").filter(viceras__icontains="3-5/7 días").count()
        viceras_4 = nutricional.values("viceras").filter(viceras__icontains="7/7 días").count()
        viceras_5 = nutricional.values("viceras").filter(viceras__icontains="1-2/15 días").count()
        viceras_6 = nutricional.values("viceras").filter(viceras__icontains="1-2/30 días").count()

        grasas_1 = nutricional.values("grasas").filter(grasas__icontains="0/7 días").count()
        grasas_2 = nutricional.values("grasas").filter(grasas__icontains="1-2/7 días").count()
        grasas_3 = nutricional.values("grasas").filter(grasas__icontains="3-5/7 días").count()
        grasas_4 = nutricional.values("grasas").filter(grasas__icontains="7/7 días").count()
        grasas_5 = nutricional.values("grasas").filter(grasas__icontains="1-2/15 días").count()
        grasas_6 = nutricional.values("grasas").filter(grasas__icontains="1-2/30 días").count()

        lacteos_1 = nutricional.values("lacteos").filter(lacteos__icontains="0/7 días").count()
        lacteos_2 = nutricional.values("lacteos").filter(lacteos__icontains="1-2/7 días").count()
        lacteos_3 = nutricional.values("lacteos").filter(lacteos__icontains="3-5/7 días").count()
        lacteos_4 = nutricional.values("lacteos").filter(lacteos__icontains="7/7 días").count()
        lacteos_5 = nutricional.values("lacteos").filter(lacteos__icontains="1-2/15 días").count()
        lacteos_6 = nutricional.values("lacteos").filter(lacteos__icontains="1-2/30 días").count()

        huevos_1 = nutricional.values("huevos").filter(huevos__icontains="0/7 días").count()
        huevos_2 = nutricional.values("huevos").filter(huevos__icontains="1-2/7 días").count()
        huevos_3 = nutricional.values("huevos").filter(huevos__icontains="3-5/7 días").count()
        huevos_4 = nutricional.values("huevos").filter(huevos__icontains="7/7 días").count()
        huevos_5 = nutricional.values("huevos").filter(huevos__icontains="1-2/15 días").count()
        huevos_6 = nutricional.values("huevos").filter(huevos__icontains="1-2/30 días").count()

        leguminosas_1 = nutricional.values("leguminosas").filter(leguminosas__icontains="0/7 días").count()
        leguminosas_2 = nutricional.values("leguminosas").filter(leguminosas__icontains="1-2/7 días").count()
        leguminosas_3 = nutricional.values("leguminosas").filter(leguminosas__icontains="3-5/7 días").count()
        leguminosas_4 = nutricional.values("leguminosas").filter(leguminosas__icontains="7/7 días").count()
        leguminosas_5 = nutricional.values("leguminosas").filter(leguminosas__icontains="1-2/15 días").count()
        leguminosas_6 = nutricional.values("leguminosas").filter(leguminosas__icontains="1-2/30 días").count()

        tuberculos_1 = nutricional.values("tuberculos").filter(tuberculos__icontains="0/7 días").count()
        tuberculos_2 = nutricional.values("tuberculos").filter(tuberculos__icontains="1-2/7 días").count()
        tuberculos_3 = nutricional.values("tuberculos").filter(tuberculos__icontains="3-5/7 días").count()
        tuberculos_4 = nutricional.values("tuberculos").filter(tuberculos__icontains="7/7 días").count()
        tuberculos_5 = nutricional.values("tuberculos").filter(tuberculos__icontains="1-2/15 días").count()
        tuberculos_6 = nutricional.values("tuberculos").filter(tuberculos__icontains="1-2/30 días").count()

        charcuteria_1 = nutricional.values("charcuteria").filter(charcuteria__icontains="0/7 días").count()
        charcuteria_2 = nutricional.values("charcuteria").filter(charcuteria__icontains="1-2/7 días").count()
        charcuteria_3 = nutricional.values("charcuteria").filter(charcuteria__icontains="3-5/7 días").count()
        charcuteria_4 = nutricional.values("charcuteria").filter(charcuteria__icontains="7/7 días").count()
        charcuteria_5 = nutricional.values("charcuteria").filter(charcuteria__icontains="1-2/15 días").count()
        charcuteria_6 = nutricional.values("charcuteria").filter(charcuteria__icontains="1-2/30 días").count()

    #**************  Agrega la data a las celdas  ***************************

    #**************  RESPONSABLE DE HACER EL MERCADO Y COCINAR  *********
        datos = [
                ('Alimentos', '0/7 días', '1-2/7 días', '3-5/7 días', '7/7 días', '1-2/15 días', '1-2/30 días'),
                ('Cereales', cereales_1, cereales_2, cereales_3, cereales_4, cereales_5, cereales_6),
                ('Vegetales', vegetales_1, vegetales_2, vegetales_3, vegetales_4, vegetales_5, vegetales_6),
                ('Frutas', frutas_1, frutas_2, frutas_3, frutas_4, frutas_5, frutas_6),
                ('Carnes', carnes_1, carnes_2, carnes_3, carnes_4, carnes_5, carnes_6),
                ('Pollo', pollo_1, pollo_2, pollo_3, pollo_4, pollo_5, pollo_6),
                ('Pescado', pescado_1, pescado_2, pescado_3, pescado_4, pescado_5, pescado_6),
                ('Embutidos', embutidos_1, embutidos_2, embutidos_3, embutidos_4, embutidos_5, embutidos_6),
                ('Viceras', viceras_1, viceras_2, viceras_3, viceras_4, viceras_5, viceras_6),
                ('Grasas', grasas_1, grasas_2, grasas_3, grasas_4, grasas_5, grasas_6),
                ('Lacteos', lacteos_1, lacteos_2, lacteos_3, lacteos_4, lacteos_5, lacteos_6),
                ('Huevos', huevos_1, huevos_2, huevos_3, huevos_4, huevos_5, huevos_6),
                ('Leguminosas', leguminosas_1, leguminosas_2, leguminosas_3, leguminosas_4, leguminosas_5, leguminosas_6),
                ('Tuberculos', tuberculos_1, tuberculos_2, tuberculos_3, tuberculos_4, tuberculos_5, tuberculos_6),
                ('Charcuteria', charcuteria_1, charcuteria_2, charcuteria_3, charcuteria_4, charcuteria_5, charcuteria_6),
            ]
        for row, cell_value in enumerate(datos, 93):
            ws.cell(row=row+1, column=13)
            for col_num, cell_value in enumerate(cell_value, 6):
                cell = ws.cell(row=row+1, column=col_num)
                cell.value = cell_value
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                cell.alignment = Alignment(horizontal="center", vertical="center")
        
        chart1 = BarChart3D()
        chart1.type = "col"
        chart1.style = 10
        chart1.title = "FRECUENCIA DEL CONSUMO DE ALIMENTOS"
        chart1.y_axis.title = 'Número de Casos'
        chart1.x_axis.title = 'Alimentos'
        chart1.width = 44  # Ancho en pulgadas  
        chart1.height = 12  # Altura en pulgadas

        data = Reference(ws, min_col=7, min_row=94, max_row=109, max_col=12)
        cats = Reference(ws, min_col=6, min_row=95, max_row=109)
        chart1.add_data(data, titles_from_data=True)
        chart1.set_categories(cats)
    
        ws.add_chart(chart1, "C110")


#**********************       TITULO DE LA QUINTA PREGUNTA    *********************
        ws.merge_cells('A135:P136')
        fourth_cell = ws['A135']
        fourth_cell.value = "POCO CONSUMO DE ALIMENTOS"
        fourth_cell.font  = Font(name = 'Tahoma', size = 12, bold = True, color="333399")
        fourth_cell.fill = PatternFill("solid", fgColor="E6E6FA")
        fourth_cell.alignment = Alignment(horizontal="center", vertical="center")      

    #********* asigna el titulo a los graficos  ************************

        ws.merge_cells('G138:J138')
        six_cell = ws['G138']
        six_cell.value = "POCO CONSUMO DE VEGETALES FRUTAS Y VICERAS"
        six_cell.font  = Font(name = 'Tahoma', size = 12, bold = True, color="333399")
        six_cell.fill = PatternFill("solid", fgColor="C6E2FF")
        six_cell.alignment = Alignment(horizontal="center", vertical="center") 

    #**************  Obtener el total de las encuestas  ***************

        poco_vegetales_1 = nutricional.values("poco_vegetales").filter(poco_vegetales__icontains="Rechazo").count()
        poco_vegetales_2 = nutricional.values("poco_vegetales").filter(poco_vegetales__icontains="prepararlo").count()
        poco_vegetales_3 = nutricional.values("poco_vegetales").filter(poco_vegetales__icontains="costosos").count()
        poco_vegetales_4 = nutricional.values("poco_vegetales").filter(poco_vegetales__icontains="acostumbrados").count()

        poco_frutas_1 = nutricional.values("poco_frutas").filter(poco_frutas__icontains="Rechazo").count()
        poco_frutas_2 = nutricional.values("poco_frutas").filter(poco_frutas__icontains="prepararlo").count()
        poco_frutas_3 = nutricional.values("poco_frutas").filter(poco_frutas__icontains="costosos").count()
        poco_frutas_4 = nutricional.values("poco_frutas").filter(poco_frutas__icontains="acostumbrados").count()

        poco_viceras_1 = nutricional.values("poco_viceras").filter(poco_viceras__icontains="Rechazo").count()
        poco_viceras_2 = nutricional.values("poco_viceras").filter(poco_viceras__icontains="prepararlo").count()
        poco_viceras_3 = nutricional.values("poco_viceras").filter(poco_viceras__icontains="costosos").count()
        poco_viceras_4 = nutricional.values("poco_viceras").filter(poco_viceras__icontains="acostumbrados").count()
        

    #**************  Agrega la data a las celdas  ***************************

    #**************  CONSUMO DE VEGETALES, FRUTAS Y VICERAS  *********
        datos = [
                ('Poco Consumo', 'Vegetales', 'Frutas', 'Viceras'),
                ('Por Rechazo', poco_vegetales_1, poco_frutas_1, poco_viceras_1),
                ('No sabe prepararlo', poco_vegetales_2, poco_frutas_2, poco_viceras_2),
                ('Están muy costosos', poco_vegetales_3, poco_frutas_3, poco_viceras_3),
                ('No están acostumbrados', poco_vegetales_4, poco_frutas_4, poco_viceras_4),
            ]
        
        for row, cell_value in enumerate(datos, 139):
            ws.cell(row=row+1, column=13)
            for col_num, cell_value in enumerate(cell_value, 7):
                cell = ws.cell(row=row+1, column=col_num)
                cell.value = cell_value
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                
        chart2 = BarChart3D()
        chart2.type = "col"
        chart2.style = 10
        chart2.title = "POCO CONSUMO DE ALIMENTOS"
        chart2.y_axis.title = 'Número de Casos'
        chart2.x_axis.title = 'Causas de poco consumo'
        chart2.width = 15  # Ancho en pulgadas  
        chart2.height = 8  # Altura en pulgadas

        data = Reference(ws, min_col=8, min_row=140, max_row=144, max_col=10)
        cats = Reference(ws, min_col=7, min_row=141, max_row=144)
        chart2.add_data(data, titles_from_data=True)
        chart2.set_categories(cats)
    
        ws.add_chart(chart2, "G146")


#**********************       TITULO DE LA SEXTA PREGUNTA    *********************
        ws.merge_cells('A162:P163')
        fourth_cell = ws['A162']
        fourth_cell.value = "AYUDA ECONOMICA Y BONOS"
        fourth_cell.font  = Font(name = 'Tahoma', size = 12, bold = True, color="333399")
        fourth_cell.fill = PatternFill("solid", fgColor="E6E6FA")
        fourth_cell.alignment = Alignment(horizontal="center", vertical="center")      

    #********* asigna el titulo a los graficos  ************************

        ws.merge_cells('G165:J165')
        six_cell = ws['G165']
        six_cell.value = "RECIBE ALGUNA AYUDA ECONOMICA Y ALGUN BONO"
        six_cell.font  = Font(name = 'Tahoma', size = 12, bold = True, color="333399")
        six_cell.fill = PatternFill("solid", fgColor="C6E2FF")
        six_cell.alignment = Alignment(horizontal="center", vertical="center") 

    #**************  Obtener el total de las encuestas  ***************

        recibe_bono = nutricional.values("bonos").filter(bonos__icontains="Si").count()
        recibe_clap = nutricional.values("clap").filter(clap__icontains="Si").count()
        recibe_iglesia = nutricional.values("iglesia").filter(iglesia__icontains="Si").count()
        recibe_familiar = nutricional.values("familiar").filter(familiar__icontains="Si").count()
        recibe_pensionado = nutricional.values("pensionado").filter(pensionado__icontains="Si").count()

    #**************  Agrega la data a las celdas  ***************************

    #**************  RECIBE ALGUNA AYUDA ECONOMICA O BONO  *********
        datos = [
                ('Ayuda Economica', 'Recibidos'),
                ('Algún Bono', recibe_bono),
                ('CALP', recibe_clap),
                ('Iglesia', recibe_iglesia),
                ('Familiar', recibe_familiar),
                ('Pensionado', recibe_pensionado),
            ]
        
        for row, cell_value in enumerate(datos, 166):
            ws.cell(row=row+1, column=13)
            for col_num, cell_value in enumerate(cell_value, 8):
                cell = ws.cell(row=row+1, column=col_num)
                cell.value = cell_value
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                cell.alignment = Alignment(horizontal="center", vertical="center")           

        chart2 = BarChart3D()
        chart2.type = "col"
        chart2.style = 10
        chart2.title = "AYUDA ECONOMICA Y ALGUN BONO"
        chart2.y_axis.title = 'Número de Casos'
        chart2.x_axis.title = 'Ayudas y Bonos'
        chart2.width = 15  # Ancho en pulgadas  
        chart2.height = 8  # Altura en pulgadas

        data = Reference(ws, min_col=9, min_row=167, max_row=172)
        cats = Reference(ws, min_col=8, min_row=168, max_row=172)
        chart2.add_data(data, titles_from_data=True)
        chart2.set_categories(cats)
    
        ws.add_chart(chart2, "G173")


#**********************       TITULO DE LA SEXTA PREGUNTA    *********************
        ws.merge_cells('A189:P190')
        fourth_cell = ws['A189']
        fourth_cell.value = "ACTIVIDAD DEPORTIVA"
        fourth_cell.font  = Font(name = 'Tahoma', size = 12, bold = True, color="333399")
        fourth_cell.fill = PatternFill("solid", fgColor="E6E6FA")
        fourth_cell.alignment = Alignment(horizontal="center", vertical="center")      

    #********* asigna el titulo a los graficos  ************************
        ws.merge_cells('B192:E192')
        five_cell = ws['B192']
        five_cell.value = "PRACTICA ALGUN DEPORTE"
        five_cell.font  = Font(name = 'Tahoma', size = 12, bold = True, color="333399")
        five_cell.fill = PatternFill("solid", fgColor="C6E2FF")
        five_cell.alignment = Alignment(horizontal="center", vertical="center")  

        ws.merge_cells('G192:J192')
        six_cell = ws['G192']
        six_cell.value = "TIEMPO DE LA ACTIVIDAD"
        six_cell.font  = Font(name = 'Tahoma', size = 12, bold = True, color="333399")
        six_cell.fill = PatternFill("solid", fgColor="C6E2FF")
        six_cell.alignment = Alignment(horizontal="center", vertical="center") 

        ws.merge_cells('L192:O192')
        seven_cell = ws['L192']
        seven_cell.value = "TIPO DE ACTIVIDAD"
        seven_cell.font  = Font(name = 'Tahoma', size = 12, bold = True, color="333399")
        seven_cell.fill = PatternFill("solid", fgColor="C6E2FF")
        seven_cell.alignment = Alignment(horizontal="center", vertical="center") 

    #**************  Obtener el total de las encuestas  ***************

        practica_si = nutricional.values("practica_deporte").filter(practica_deporte__icontains="Si").count()
        practica_no = nutricional.values("practica_deporte").filter(practica_deporte__icontains="No").count()

        tiempo_1 = nutricional.values("tiempo").filter(tiempo__icontains="30 min de 1-3 / 7").count()
        tiempo_2 = nutricional.values("tiempo").filter(tiempo__icontains="30 min de 4-7 / 7").count()
        tiempo_3 = nutricional.values("tiempo").filter(tiempo__icontains="más de 30 min 1-3 / 7").count()
        tiempo_4 = nutricional.values("tiempo").filter(tiempo__icontains="más de 30 min 4-7 / 7").count()
        tiempo_5 = nutricional.values("tiempo").filter(tiempo__icontains="No precisa datos").count()

        actividad_1 = nutricional.values("actividad").filter(actividad__icontains="Juegos").count()
        actividad_2 = nutricional.values("actividad").filter(actividad__icontains="Práctica").count()
        actividad_3 = nutricional.values("actividad").filter(actividad__icontains="Ambas").count()

    #**************  Agrega la data a las celdas  ***************************

    #**************  PRACTICA ALGUN DEPORTE  *********
        datos = [
                ('Practica', 'Resultados'),
                ('Si', practica_si),
                ('No', practica_no),
            ]
        
        for row, cell_value in enumerate(datos, 193):
            ws.cell(row=row+1, column=13)
            for col_num, cell_value in enumerate(cell_value, 3):
                cell = ws.cell(row=row+1, column=col_num)
                cell.value = cell_value
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                

        chart2 = BarChart3D()
        chart2.type = "col"
        chart2.style = 10
        chart2.title = "PRACTICA DEPORTES"
        chart2.y_axis.title = 'Número de Casos'
        chart2.x_axis.title = 'Practica algún Deporte'
        chart2.width = 15  # Ancho en pulgadas  
        chart2.height = 8  # Altura en pulgadas

        data = Reference(ws, min_col=4, min_row=194, max_row=196)
        cats = Reference(ws, min_col=3, min_row=195, max_row=196)
        chart2.add_data(data, titles_from_data=True)
        chart2.set_categories(cats)
    
        ws.add_chart(chart2, "B200")

    #**************  TIEMPO DE LA ACTIVIDAD  *********
        datos = [
                ('Tiempos', 'Resultados'),
                ('30 min de 1-3 / 7', tiempo_1),
                ('30 min de 4-7 / 7', tiempo_2),
                ('más de 30 min 1-3 / 7', tiempo_3),
                ('más de 30 min 4-7 / 7', tiempo_4),
                ('No precisa datos', tiempo_5),
            ]
        
        for row, cell_value in enumerate(datos, 193):
            ws.cell(row=row+1, column=13)
            for col_num, cell_value in enumerate(cell_value, 8):
                cell = ws.cell(row=row+1, column=col_num)
                cell.value = cell_value
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                

        chart2 = BarChart3D()
        chart2.type = "col"
        chart2.style = 10
        chart2.title = "TIEMPO DE LA ACTIVIDAD"
        chart2.y_axis.title = 'Número de Casos'
        chart2.x_axis.title = 'Frecuencia de la Actividad'
        chart2.width = 15  # Ancho en pulgadas  
        chart2.height = 8  # Altura en pulgadas

        data = Reference(ws, min_col=9, min_row=194, max_row=199)
        cats = Reference(ws, min_col=8, min_row=195, max_row=199)
        chart2.add_data(data, titles_from_data=True)
        chart2.set_categories(cats)
    
        ws.add_chart(chart2, "G200")

    #**************  TIPO DE ACTIVIDAD  *********
        datos = [
                ('Tipo de Actividad', 'Frecuencia'),
                ('Juegos Activos', meriendas_una),
                ('Práctica Deportiva', meriendas_dos),
                ('Ambas Actividades', meriendas_tres),
            ]
        
        for row, cell_value in enumerate(datos, 193):
            ws.cell(row=row+1, column=13)
            for col_num, cell_value in enumerate(cell_value, 13):
                cell = ws.cell(row=row+1, column=col_num)
                cell.value = cell_value
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                cell.alignment = Alignment(horizontal="center", vertical="center")

        chart2 = BarChart3D()
        chart2.type = "col"
        chart2.style = 10
        chart2.title = "TIPO DE ACTIVIDAD"
        chart2.y_axis.title = 'Número de Casos'
        chart2.x_axis.title = 'Tipo de la Actividad'
        chart2.width = 15  # Ancho en pulgadas  
        chart2.height = 8  # Altura en pulgadas

        data = Reference(ws, min_col=14, min_row=194, max_row=197)
        cats = Reference(ws, min_col=13, min_row=195, max_row=197)
        chart2.add_data(data, titles_from_data=True)
        chart2.set_categories(cats)
    
        ws.add_chart(chart2, "L200")


#**********************       TITULO DE LA SEPTIMA PREGUNTA    *********************
        ws.merge_cells('A216:P217')
        fourth_cell = ws['A216']
        fourth_cell.value = "CONSUMO DE AGUA Y SERVICIOS BASICOS"
        fourth_cell.font  = Font(name = 'Tahoma', size = 12, bold = True, color="333399")
        fourth_cell.fill = PatternFill("solid", fgColor="E6E6FA")
        fourth_cell.alignment = Alignment(horizontal="center", vertical="center")      

    #********* asigna el titulo a los graficos  ************************
        ws.merge_cells('D219:G219')
        five_cell = ws['D219']
        five_cell.value = "COMO ES EL CONSUMO DEL AGUA"
        five_cell.font  = Font(name = 'Tahoma', size = 12, bold = True, color="333399")
        five_cell.fill = PatternFill("solid", fgColor="C6E2FF")
        five_cell.alignment = Alignment(horizontal="center", vertical="center")  

        ws.merge_cells('J219:M219')
        six_cell = ws['J219']
        six_cell.value = "CUAL SERVICIO PRESENTAN FALLAS"
        six_cell.font  = Font(name = 'Tahoma', size = 12, bold = True, color="333399")
        six_cell.fill = PatternFill("solid", fgColor="C6E2FF")
        six_cell.alignment = Alignment(horizontal="center", vertical="center") 

    #**************  Obtener el total de las encuestas  ***************

        agua_tuberia = nutricional.values("agua").filter(agua__icontains="Tuberia").count()
        agua_filtrada = nutricional.values("agua").filter(agua__icontains="Filtrada").count()
        agua_botellon = nutricional.values("agua").filter(agua__icontains="Botellon").count()
        agua_hervida = nutricional.values("agua").filter(agua__icontains="Hervida").count()
        agua_tabletas = nutricional.values("agua").filter(agua__icontains="Tabletas").count()

        falla_agua = nutricional.values("falla_servicio").filter(falla_servicio__icontains="agua").count()
        falla_gas = nutricional.values("falla_servicio").filter(falla_servicio__icontains="Gas").count()
        falla_electricidad = nutricional.values("falla_servicio").filter(falla_servicio__icontains="Eléctricidad").count()
        falla_telefonia = nutricional.values("falla_servicio").filter(falla_servicio__icontains="Telefonía").count()
        falla_aseo = nutricional.values("falla_servicio").filter(falla_servicio__icontains="Aseo").count()

    #**************  Agrega la data a las celdas  ***************************

    #**************  COMO ES EL CONSUMO DEL AGUA  *********
        datos = [
                ('Consumo del Agua', 'Consumo'),
                ('Tuberia', agua_tuberia),
                ('Filtrada', agua_filtrada),
                ('Botellon', agua_botellon),
                ('Hervida', agua_hervida),
                ('Tabletas', agua_tabletas),
            ]
        for row, cell_value in enumerate(datos, 220):
            ws.cell(row=row+1, column=13)
            for col_num, cell_value in enumerate(cell_value, 5):
                cell = ws.cell(row=row+1, column=col_num)
                cell.value = cell_value
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                cell.alignment = Alignment(horizontal="center", vertical="center")
        
        chart1 = BarChart3D()
        chart1.type = "col"
        chart1.style = 10
        chart1.title = "CONSUMO DEL AGUA"
        chart1.y_axis.title = 'Número de Casos'
        chart1.x_axis.title = 'Consumo del Agua'
        chart1.width = 15  # Ancho en pulgadas  
        chart1.height = 8  # Altura en pulgadas

        data = Reference(ws, min_col=6, min_row=221, max_row=226)
        cats = Reference(ws, min_col=5, min_row=222, max_row=226)
        chart1.add_data(data, titles_from_data=True)
        chart1.set_categories(cats)
    
        ws.add_chart(chart1, "D227")


    #**************  CUAL SERVICIO PRESENTA FALLAS  *********

        #********* asigna los Datos a las celdas  ************************
        datos = [
                ('Servicio', 'Presentan Falla'),
                ('Agua', falla_agua),
                ('Gas', falla_gas),
                ('Eléctricidad', falla_electricidad),
                ('Telefonía Internet', falla_telefonia),
                ('Aseo Urbano', falla_aseo),
            ]
        
        for row, cell_value in enumerate(datos, 220):
            ws.cell(row=row+1, column=13)
            for col_num, cell_value in enumerate(cell_value, 11):
                cell = ws.cell(row=row+1, column=col_num)
                cell.value = cell_value
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                cell.alignment = Alignment(horizontal="center", vertical="center")

        chart2 = BarChart3D()
        chart2.type = "col"
        chart2.style = 10
        chart2.title = "SERVICIO QUE PRESENTAN FALLAS"
        chart2.y_axis.title = 'Número de Casos'
        chart2.x_axis.title = 'Servicios que Presentan Falla'
        chart2.width = 15  # Ancho en pulgadas  
        chart2.height = 8  # Altura en pulgadas

        data = Reference(ws, min_col=12, min_row=221, max_row=226)
        cats = Reference(ws, min_col=11, min_row=222, max_row=226)
        chart2.add_data(data, titles_from_data=True)
        chart2.set_categories(cats)
    
        ws.add_chart(chart2, "J227")


#******************    COMPRA GAS AGUA ALMACENA AGUA Y DONDE LO ALMACENA    ********************  

    #********* asigna el titulo a los graficos  ************************
        ws.merge_cells('B244:D244')
        five_cell = ws['B244']
        five_cell.value = "COMPRA DE GAS"
        five_cell.font  = Font(name = 'Tahoma', size = 12, bold = True, color="333399")
        five_cell.fill = PatternFill("solid", fgColor="C6E2FF")
        five_cell.alignment = Alignment(horizontal="center", vertical="center")  

        ws.merge_cells('F244:H244')
        six_cell = ws['F244']
        six_cell.value = "COMPRA DE AGUA"
        six_cell.font  = Font(name = 'Tahoma', size = 12, bold = True, color="333399")
        six_cell.fill = PatternFill("solid", fgColor="C6E2FF")
        six_cell.alignment = Alignment(horizontal="center", vertical="center") 

        ws.merge_cells('J244:L244')
        seven_cell = ws['J244']
        seven_cell.value = "ALMACENAMIENTO DE AGUA"
        seven_cell.font  = Font(name = 'Tahoma', size = 12, bold = True, color="333399")
        seven_cell.fill = PatternFill("solid", fgColor="C6E2FF")
        seven_cell.alignment = Alignment(horizontal="center", vertical="center") 

        ws.merge_cells('N244:P244')
        eight_cell = ws['N244']
        eight_cell.value = "DONDE SE ALMACENA"
        eight_cell.font  = Font(name = 'Tahoma', size = 12, bold = True, color="333399")
        eight_cell.fill = PatternFill("solid", fgColor="C6E2FF")
        eight_cell.alignment = Alignment(horizontal="center", vertical="center") 

    #**************  Obtener el total de las encuestas  ***************

        gas_si = nutricional.values("compra_gas").filter(compra_gas__icontains="Si").count()
        gas_no = nutricional.values("compra_gas").filter(compra_gas__icontains="No").count()

        agua_si = nutricional.values("compra_agua").filter(compra_agua__icontains="Si").count()
        agua_no = nutricional.values("compra_agua").filter(compra_agua__icontains="No").count()

        almacena_si = nutricional.values("almacena_agua").filter(almacena_agua__icontains="Si").count()
        almacena_no = nutricional.values("almacena_agua").filter(almacena_agua__icontains="No").count()

        donde_pipote = nutricional.values("donde_almacena").filter(donde_almacena__icontains="Pipote").count()
        donde_tanque = nutricional.values("donde_almacena").filter(donde_almacena__icontains="Tanque").count()
        donde_otros = nutricional.values("donde_almacena").filter(donde_almacena__icontains="Otros").count()

    #**************  Agrega la data a las celdas  ***************************

    #**************  COMPRA DE GAS  *********
        datos = [
                ('Compra Gas', 'Resultados'),
                ('Si', gas_si),
                ('No', gas_no),
            ]
        
        for row, cell_value in enumerate(datos, 245):
            ws.cell(row=row+1, column=13)
            for col_num, cell_value in enumerate(cell_value, 2):
                cell = ws.cell(row=row+1, column=col_num)
                cell.value = cell_value
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                

        chart2 = BarChart3D()
        chart2.type = "col"
        chart2.style = 10
        chart2.title = "COMPRA DE GAS"
        chart2.y_axis.title = 'Número de Casos'
        chart2.x_axis.title = 'Compran Gas'
        chart2.width = 11  # Ancho en pulgadas  
        chart2.height = 8  # Altura en pulgadas

        data = Reference(ws, min_col=3, min_row=246, max_row=248)
        cats = Reference(ws, min_col=2, min_row=247, max_row=248)
        chart2.add_data(data, titles_from_data=True)
        chart2.set_categories(cats)
    
        ws.add_chart(chart2, "B250")

    #**************  COMPRA DE AGUA  *********
        datos = [
                ('Compra Agua', 'Resultados'),
                ('Si', agua_si),
                ('No', agua_no),
            ]
        
        for row, cell_value in enumerate(datos, 245):
            ws.cell(row=row+1, column=13)
            for col_num, cell_value in enumerate(cell_value, 6):
                cell = ws.cell(row=row+1, column=col_num)
                cell.value = cell_value
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                

        chart2 = BarChart3D()
        chart2.type = "col"
        chart2.style = 10
        chart2.title = "COMPRA DE AGUA"
        chart2.y_axis.title = 'Número de Casos'
        chart2.x_axis.title = 'Compran Agua'
        chart2.width = 11  # Ancho en pulgadas  
        chart2.height = 8  # Altura en pulgadas

        data = Reference(ws, min_col=7, min_row=246, max_row=248)
        cats = Reference(ws, min_col=6, min_row=247, max_row=248)
        chart2.add_data(data, titles_from_data=True)
        chart2.set_categories(cats)
    
        ws.add_chart(chart2, "F250")

    #**************  ALMACENAMIENTO DE AGUA  *********
        datos = [
                ('Almacena Agua', 'Resultados'),
                ('Si', almacena_si),
                ('No', almacena_no),
            ]
        
        for row, cell_value in enumerate(datos, 245):
            ws.cell(row=row+1, column=13)
            for col_num, cell_value in enumerate(cell_value, 10):
                cell = ws.cell(row=row+1, column=col_num)
                cell.value = cell_value
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                cell.alignment = Alignment(horizontal="center", vertical="center")

        chart2 = BarChart3D()
        chart2.type = "col"
        chart2.style = 10
        chart2.title = "ALMACENAMIENTO DE AGUA"
        chart2.y_axis.title = 'Número de Casos'
        chart2.x_axis.title = 'Almacenamiento del Agua'
        chart2.width = 11  # Ancho en pulgadas  
        chart2.height = 8  # Altura en pulgadas

        data = Reference(ws, min_col=11, min_row=246, max_row=248)
        cats = Reference(ws, min_col=10, min_row=247, max_row=248)
        chart2.add_data(data, titles_from_data=True)
        chart2.set_categories(cats)
    
        ws.add_chart(chart2, "J250")

            #**************  DONDE SE ALMACENA  *********
        datos = [
                ('Donde Almacenan', 'Resultados'),
                ('Pipote', donde_pipote),
                ('Tanque', donde_tanque),
                ('Otros', donde_otros),
            ]
        
        for row, cell_value in enumerate(datos, 245):
            ws.cell(row=row+1, column=13)
            for col_num, cell_value in enumerate(cell_value, 14):
                cell = ws.cell(row=row+1, column=col_num)
                cell.value = cell_value
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                cell.alignment = Alignment(horizontal="center", vertical="center")

        chart2 = BarChart3D()
        chart2.type = "col"
        chart2.style = 10
        chart2.title = "DONDE SE ALMACENA"
        chart2.y_axis.title = 'Número de Casos'
        chart2.x_axis.title = 'Almacenamiento del Agua'
        chart2.width = 11  # Ancho en pulgadas  
        chart2.height = 8  # Altura en pulgadas

        data = Reference(ws, min_col=15, min_row=246, max_row=249)
        cats = Reference(ws, min_col=14, min_row=247, max_row=249)
        chart2.add_data(data, titles_from_data=True)
        chart2.set_categories(cats)
    
        ws.add_chart(chart2, "N250")


#**********************       TITULO DE LA OPTAVA PREGUNTA    *********************
        ws.merge_cells('A266:P267')
        fourth_cell = ws['A266']
        fourth_cell.value = "EDUCACION NUTRICIONAL"
        fourth_cell.font  = Font(name = 'Tahoma', size = 12, bold = True, color="333399")
        fourth_cell.fill = PatternFill("solid", fgColor="E6E6FA")
        fourth_cell.alignment = Alignment(horizontal="center", vertical="center")      

    #********* asigna el titulo a los graficos  ************************
        ws.merge_cells('B269:E269')
        five_cell = ws['B269']
        five_cell.value = "COONOCE LOS GRUPOS DE ALIMENTOS"
        five_cell.font  = Font(name = 'Tahoma', size = 12, bold = True, color="333399")
        five_cell.fill = PatternFill("solid", fgColor="C6E2FF")
        five_cell.alignment = Alignment(horizontal="center", vertical="center")  

        ws.merge_cells('G269:J269')
        six_cell = ws['G269']
        six_cell.value = "CONOCE DE LA PROTEINA VEGETAL"
        six_cell.font  = Font(name = 'Tahoma', size = 12, bold = True, color="333399")
        six_cell.fill = PatternFill("solid", fgColor="C6E2FF")
        six_cell.alignment = Alignment(horizontal="center", vertical="center") 

        ws.merge_cells('L269:O269')
        seven_cell = ws['L269']
        seven_cell.value = "CONOCE SOBRE LA DESNUTRICION"
        seven_cell.font  = Font(name = 'Tahoma', size = 12, bold = True, color="333399")
        seven_cell.fill = PatternFill("solid", fgColor="C6E2FF")
        seven_cell.alignment = Alignment(horizontal="center", vertical="center") 

    #**************  Obtener el total de las encuestas  ***************

        conoce_grupos_si = nutricional.values("conoce_grupos").filter(conoce_grupos__icontains="Si").count()
        conoce_grupos_no = nutricional.values("conoce_grupos").filter(conoce_grupos__icontains="No").count()
        conoce_grupos_poca = nutricional.values("conoce_grupos").filter(conoce_grupos__icontains="Poca").count()

        conoce_vegetales_si = nutricional.values("conoce_calidad").filter(conoce_calidad__icontains="Si").count()
        conoce_vegetales_no = nutricional.values("conoce_calidad").filter(conoce_calidad__icontains="No").count()
        conoce_vegetales_poca = nutricional.values("conoce_calidad").filter(conoce_calidad__icontains="Poca").count()

        conoce_desnutricion_si = nutricional.values("conoce_desnutricion").filter(conoce_desnutricion__icontains="Si").count()
        conoce_desnutricion_no = nutricional.values("conoce_desnutricion").filter(conoce_desnutricion__icontains="No").count()
        conoce_desnutricion_poca = nutricional.values("conoce_desnutricion").filter(conoce_desnutricion__icontains="Poca").count()


    #**************  Agrega la data a las celdas  ***************************

    #**************  COONOCE LOS GRUPOS DE ALIMENTOS  *********
        datos = [
                ('Conoce los Grupos', 'Resultados'),
                ('Si', conoce_grupos_si),
                ('No', conoce_grupos_no),
                ('Poca Información', conoce_grupos_poca),
            ]
        
        for row, cell_value in enumerate(datos, 270):
            ws.cell(row=row+1, column=13)
            for col_num, cell_value in enumerate(cell_value, 3):
                cell = ws.cell(row=row+1, column=col_num)
                cell.value = cell_value
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                

        chart2 = BarChart3D()
        chart2.type = "col"
        chart2.style = 10
        chart2.title = "COONOCE LOS GRUPOS DE ALIMENTOS"
        chart2.y_axis.title = 'Número de Casos'
        chart2.x_axis.title = 'Conoce los Grupos de Alimentos'
        chart2.width = 15  # Ancho en pulgadas  
        chart2.height = 8  # Altura en pulgadas

        data = Reference(ws, min_col=4, min_row=271, max_row=274)
        cats = Reference(ws, min_col=3, min_row=272, max_row=274)
        chart2.add_data(data, titles_from_data=True)
        chart2.set_categories(cats)
    
        ws.add_chart(chart2, "B276")

    #**************  CONOCE DE LA PROTEINA VEGETAL  *********
        datos = [
                ('Conoce la Proteina', 'Resultados'),
                ('Si', conoce_vegetales_si),
                ('No', conoce_vegetales_no),
                ('Poca Información', conoce_vegetales_poca),
            ]
        
        for row, cell_value in enumerate(datos, 270):
            ws.cell(row=row+1, column=13)
            for col_num, cell_value in enumerate(cell_value, 8):
                cell = ws.cell(row=row+1, column=col_num)
                cell.value = cell_value
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                

        chart2 = BarChart3D()
        chart2.type = "col"
        chart2.style = 10
        chart2.title = "CALIDAD DE LA PROTEINA VEGETAL"
        chart2.y_axis.title = 'Número de Casos'
        chart2.x_axis.title = 'Conoce la Calidad de la Proteina Vegetal'
        chart2.width = 15  # Ancho en pulgadas  
        chart2.height = 8  # Altura en pulgadas

        data = Reference(ws, min_col=9, min_row=271, max_row=274)
        cats = Reference(ws, min_col=8, min_row=272, max_row=274)
        chart2.add_data(data, titles_from_data=True)
        chart2.set_categories(cats)
    
        ws.add_chart(chart2, "G276")

    #**************  CONOCE SOBRE LA DESNUTRICION  *********
        datos = [
                ('Conoce la Desnutrición', 'Resultados'),
                ('Si', conoce_desnutricion_si),
                ('No', conoce_desnutricion_no),
                ('Poca Información', conoce_desnutricion_poca),
            ]
        
        for row, cell_value in enumerate(datos, 270):
            ws.cell(row=row+1, column=13)
            for col_num, cell_value in enumerate(cell_value, 13):
                cell = ws.cell(row=row+1, column=col_num)
                cell.value = cell_value
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                cell.alignment = Alignment(horizontal="center", vertical="center")

        chart2 = BarChart3D()
        chart2.type = "col"
        chart2.style = 10
        chart2.title = "CONOCE SOBRE LA DESNUTRICION"
        chart2.y_axis.title = 'Número de Casos'
        chart2.x_axis.title = 'Conoce sobre la desnutrición y su consecuencia'
        chart2.width = 15  # Ancho en pulgadas  
        chart2.height = 8  # Altura en pulgadas

        data = Reference(ws, min_col=14, min_row=271, max_row=274)
        cats = Reference(ws, min_col=13, min_row=272, max_row=274)
        chart2.add_data(data, titles_from_data=True)
        chart2.set_categories(cats)
    
        ws.add_chart(chart2, "L276")


#**********************       TITULO DE LA NOVENA PREGUNTA    *********************
        ws.merge_cells('A292:P293')
        fourth_cell = ws['A292']
        fourth_cell.value = "EDUCACION PARA EMBARAZADAS Y LACTANTES"
        fourth_cell.font  = Font(name = 'Tahoma', size = 12, bold = True, color="333399")
        fourth_cell.fill = PatternFill("solid", fgColor="E6E6FA")
        fourth_cell.alignment = Alignment(horizontal="center", vertical="center")      

    #********* asigna el titulo a los graficos  ************************
        ws.merge_cells('E295:L295')
        five_cell = ws['E295']
        five_cell.value = "INFORMACION DE ORIENTACION PARA LAS EMBARAZADAS Y LACTANTES"
        five_cell.font  = Font(name = 'Tahoma', size = 12, bold = True, color="333399")
        five_cell.fill = PatternFill("solid", fgColor="C6E2FF")
        five_cell.alignment = Alignment(horizontal="center", vertical="center")  

    #**************  Obtener el total de las encuestas  ***************

        beneficio_si = nutricional.values("conoce_beneficio").filter(conoce_beneficio__icontains="Si").count()
        beneficio_no = nutricional.values("conoce_beneficio").filter(conoce_beneficio__icontains="No").count()
        beneficio_poca = nutricional.values("conoce_beneficio").filter(conoce_beneficio__icontains="Poca").count()
        beneficio_aplica = nutricional.values("conoce_beneficio").filter(conoce_beneficio__icontains="Aplica").count()

        amamantar_si = nutricional.values("desea_amamantar").filter(conoce_beneficio__icontains="Si").count()
        amamantar_no = nutricional.values("desea_amamantar").filter(conoce_beneficio__icontains="No").count()
        amamantar_poca = nutricional.values("desea_amamantar").filter(conoce_beneficio__icontains="Poca").count()
        amamantar_aplica = nutricional.values("desea_amamantar").filter(conoce_beneficio__icontains="Aplica").count()

        dificultad_si = nutricional.values("dificultad_amamantar").filter(dificultad_amamantar__icontains="Si").count()
        dificultad_no = nutricional.values("dificultad_amamantar").filter(dificultad_amamantar__icontains="No").count()
        dificultad_poca = nutricional.values("dificultad_amamantar").filter(dificultad_amamantar__icontains="Poca").count()
        dificultad_aplica = nutricional.values("dificultad_amamantar").filter(dificultad_amamantar__icontains="Aplica").count()

        orientacion_si = nutricional.values("desea_orientacion").filter(desea_orientacion__icontains="Si").count()
        orientacion_no = nutricional.values("desea_orientacion").filter(desea_orientacion__icontains="No").count()
        orientacion_poca = nutricional.values("desea_orientacion").filter(desea_orientacion__icontains="Poca").count()
        orientacion_aplica = nutricional.values("desea_orientacion").filter(desea_orientacion__icontains="Aplica").count()

        conocimiento_si = nutricional.values("desea_conocimiento").filter(desea_conocimiento__icontains="Si").count()
        conocimiento_no = nutricional.values("desea_conocimiento").filter(desea_conocimiento__icontains="No").count()
        conocimiento_poca = nutricional.values("desea_conocimiento").filter(desea_conocimiento__icontains="Poca").count()
        conocimiento_aplica = nutricional.values("desea_conocimiento").filter(desea_conocimiento__icontains="Aplica").count()

    #**************  Agrega la data a las celdas  ***************************

    #**************  EDUCACION PARA EMBARAZADAS Y LACTANTES  *********
        datos = [
                ('Educación', 'Si', 'No', 'Poca Información', 'No Aplica'),
                ('Beneficio Lactancia', beneficio_si, beneficio_no, beneficio_poca, beneficio_aplica),
                ('Desea Amamantar', amamantar_si, amamantar_no, amamantar_poca, amamantar_aplica),
                ('Dificultad de Amamantar', dificultad_si, dificultad_no, dificultad_poca, dificultad_aplica),
                ('Desea Orientación', orientacion_si, orientacion_no, orientacion_poca, orientacion_aplica),
                ('Mas Conocimiento', conocimiento_si, conocimiento_no, conocimiento_poca, conocimiento_aplica),
            ]
        for row, cell_value in enumerate(datos, 296):
            ws.cell(row=row+1, column=13)
            for col_num, cell_value in enumerate(cell_value, 7):
                cell = ws.cell(row=row+1, column=col_num)
                cell.value = cell_value
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                cell.alignment = Alignment(horizontal="center", vertical="center")
        
        chart1 = BarChart3D()
        chart1.type = "col"
        chart1.style = 10
        chart1.title = "EDUCACION PARA EMBARAZADAS Y LACTANTES"
        chart1.y_axis.title = 'Número de Casos'
        chart1.x_axis.title = 'EDUCACION PARA EMBARAZADAS Y LACTANTES'
        chart1.width = 37  # Ancho en pulgadas  
        chart1.height = 12  # Altura en pulgadas

        data = Reference(ws, min_col=8, min_row=297, max_row=302, max_col=11)
        cats = Reference(ws, min_col=7, min_row=298, max_row=302)
        chart1.add_data(data, titles_from_data=True)
        chart1.set_categories(cats)
    
        ws.add_chart(chart1, "D304")


#*-*-*-*-*-*-*-*-*-**-*    final de la seccion de la hoja   *-*-*-*-*-*-*-*-*

    #*********  Establecer el nombre del Archivo *******
        nombre_archvo = "Reporte_Nutricional_Jornada.xlsx"
    
    #*********  Definir el tipo de respuesta que se va a dar ***********
        response = HttpResponse(content_type = "application/ms-excel")
        contenido = "attachment; filename = {0}".format(nombre_archvo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response
    

